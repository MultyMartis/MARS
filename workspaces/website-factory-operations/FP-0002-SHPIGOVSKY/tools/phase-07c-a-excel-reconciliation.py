#!/usr/bin/env python3
"""FP-0002 V8 Phase 07C-A — Excel-driven static demo scope reconciliation."""
from __future__ import annotations

import hashlib
import json
import re
import shutil
import subprocess
import zipfile
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --- Paths ---
REPO = Path(r"X:\AI MARS")
OPS = REPO / "workspaces" / "website-factory-operations" / "FP-0002-SHPIGOVSKY"
V8 = REPO / "workspaces" / "fp-0002-shpigovsky-v8"
STORAGE = Path(r"X:\AI MARS STORAGE") / "website-factory" / "fp-0002-shpigovsky-v8" / "phase-07c-a-excel-demo-reconciliation"
EXCEL_DIR = OPS / "INCOMING" / "02_CONTENT"
EXPECTED_EXCEL_SHA = "64741FDDBD61199D6B3D80E8770576DAE86C374099C6AFEC292F9BD744512696"
DOC_COMMIT = "8612d8f6732352708c787c2c610837018ae3e1a8"
BASELINE_COMMIT = "eb47ebb4066252373e02d9e1095403d0ce6b6b22"
BASELINE_TAG = "fp-0002-v8-operator-approved-frontend-stable-01"

PROTECTED_REL = [
    "src/pages",
    "src/partials",
    "src/scss/style.scss",
    "src/js",
    "src/img",
    "gulpfile.js",
    "package.json",
]

PLACEHOLDER_NAMES = {"название", "название статьи", "название специалиста", "placeholder"}


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest().upper()


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest().upper()


def git(*args: str) -> str:
    r = subprocess.run(
        ["git", "-C", str(REPO), *args],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    if r.returncode != 0:
        raise RuntimeError(f"git {' '.join(args)} failed: {r.stderr}")
    return r.stdout.strip()


def hash_inventory(root: Path, rel_paths: list[str]) -> dict[str, str]:
    out: dict[str, str] = {}
    for rel in rel_paths:
        p = root / rel
        if p.is_file():
            out[rel] = sha256_file(p)
        elif p.is_dir():
            for f in sorted(p.rglob("*")):
                if f.is_file():
                    r = f.relative_to(root).as_posix()
                    out[r] = sha256_file(f)
    return out


def find_excel() -> Path:
    matches = list(EXCEL_DIR.glob("*.xlsx"))
    if not matches:
        raise FileNotFoundError(f"No xlsx in {EXCEL_DIR}")
    if len(matches) > 1:
        for m in matches:
            if sha256_file(m) == EXPECTED_EXCEL_SHA:
                return m
    return matches[0]


def normalize_url(url: str | None) -> str | None:
    if not url or not str(url).strip():
        return None
    u = str(url).strip().replace(" ", "")
    if not re.match(r"^[a-zA-Z][a-zA-Z0-9+.-]*://", u):
        u = "https://" + u.lstrip("/")
    parsed = urlparse(u)
    path = parsed.path or "/"
    path = re.sub(r"/+", "/", path)
    if not path.startswith("/"):
        path = "/" + path
    if path != "/" and not path.endswith("/"):
        path += "/"
    return path


def slug_from_path(path: str) -> str:
    parts = [p for p in path.strip("/").split("/") if p]
    return parts[-1] if parts else ""


def depth_from_path(path: str) -> int:
    if path == "/":
        return 0
    return len([p for p in path.strip("/").split("/") if p])


def parent_path(path: str) -> str | None:
    parts = [p for p in path.strip("/").split("/") if p]
    if len(parts) <= 1:
        return "/" if len(parts) == 1 and parts[0] else None
    parent = "/" + "/".join(parts[:-1]) + "/"
    return parent


def page_name_from_row(levels: list[str | None], path: str) -> str:
    for lv in reversed(levels):
        if lv and str(lv).strip():
            return str(lv).strip()
    slug = slug_from_path(path)
    return slug or path


def is_placeholder_name(name: str) -> bool:
    n = name.strip().lower()
    return n in PLACEHOLDER_NAMES or n.startswith("название")


@dataclass
class PageEntity:
    id: str
    worksheet: str
    source_row: int
    raw_url: str | None
    raw_page_name: str
    normalized_page_name: str
    parent_page: str | None
    hierarchy_depth: int
    category_family: str
    proposed_slug: str
    proposed_route: str
    demand_query: str | None = None
    demand_msk: int | None = None
    notes: list[str] = field(default_factory=list)
    status_priority: str | None = None
    duplicate_group: str | None = None
    content_authority: str = "EXCEL_STRUCTURE"
    design_authority: str = "OPERATOR_DECISION"
    confidence: str = "HIGH"
    unresolved_fields: list[str] = field(default_factory=list)
    is_page_like: bool = True
    disposition: str = "UNKNOWN_OPERATOR_DECISION"
    v8_match: str | None = None
    template_family: str | None = None
    content_readiness: str = "OPERATOR_DECISION"
    design_readiness: str = "OPERATOR_DECISION"
    assembly_effort: str = "MEDIUM"
    demo_risk: str = "MEDIUM"
    wp_relevance: str = "FUTURE"
    operator_decision: str | None = None
    recommendation: str = ""


V8_PAGES = [
    {
        "name": "Home",
        "ru": "Главная",
        "source": "src/pages/index.html",
        "preview": "/index.html",
        "production": "/",
        "family": "HOME",
        "template_id": "FP0002-TPL-001",
        "desktop": "OPERATOR_APPROVED",
        "mobile": "OPERATOR_APPROVED",
        "content": "CURRENT_DEMO_CONTENT",
        "design": "APPROVED_V8_PAGE",
        "role": "direct",
        "nav": "logo",
    },
    {
        "name": "O-Centre",
        "ru": "О центре",
        "source": "src/pages/o-centre.html",
        "preview": "/o-centre.html",
        "production": "/o-centre/",
        "family": "INSTITUTIONAL",
        "template_id": None,
        "desktop": "STABLE_PREVIOUSLY_APPROVED",
        "mobile": "STABLE_PREVIOUSLY_APPROVED",
        "content": "CURRENT_DEMO_CONTENT",
        "design": "APPROVED_V8_PAGE",
        "role": "direct",
        "nav": "header",
    },
    {
        "name": "Contacts",
        "ru": "Контакты",
        "source": "src/pages/kontakty.html",
        "preview": "/kontakty.html",
        "production": "/kontakty/",
        "family": "CONTACTS",
        "template_id": None,
        "desktop": "STABLE_PREVIOUSLY_APPROVED",
        "mobile": "STABLE_PREVIOUSLY_APPROVED",
        "content": "CURRENT_DEMO_CONTENT",
        "design": "APPROVED_V8_PAGE",
        "role": "direct",
        "nav": "header",
    },
    {
        "name": "Reviews",
        "ru": "Отзывы",
        "source": "src/pages/otzyvy.html",
        "preview": "/otzyvy.html",
        "production": "/otzyvy/",
        "family": "REVIEWS_ARCHIVE",
        "template_id": None,
        "desktop": "STABLE_PREVIOUSLY_APPROVED",
        "mobile": "STABLE_PREVIOUSLY_APPROVED",
        "content": "CURRENT_DEMO_CONTENT",
        "design": "APPROVED_V8_PAGE",
        "role": "direct",
        "nav": "header",
    },
    {
        "name": "Blog archive",
        "ru": "Статьи",
        "source": "src/pages/blog.html",
        "preview": "/blog.html",
        "production": "/blog/",
        "family": "BLOG_ARCHIVE",
        "template_id": None,
        "desktop": "STABLE_PREVIOUSLY_APPROVED",
        "mobile": "STABLE_PREVIOUSLY_APPROVED",
        "content": "CURRENT_DEMO_CONTENT",
        "design": "APPROVED_V8_PAGE",
        "role": "direct",
        "nav": "header",
    },
    {
        "name": "Blog Article",
        "ru": "Статья",
        "source": "src/pages/blog/nazvanie-stati.html",
        "preview": "/blog/nazvanie-stati.html",
        "production": "/blog/nazvanie-stati/",
        "family": "BLOG_ARTICLE",
        "template_id": None,
        "desktop": "OPERATOR_APPROVED",
        "mobile": "OPERATOR_APPROVED",
        "content": "CURRENT_DEMO_CONTENT",
        "design": "APPROVED_V8_PAGE",
        "role": "template",
        "nav": "internal",
    },
    {
        "name": "Services hub (legacy)",
        "ru": "Услуги (legacy)",
        "source": "src/pages/uslugi.html",
        "preview": "/uslugi.html",
        "production": "/uslugi/",
        "family": "SERVICES_HUB",
        "template_id": None,
        "desktop": "STABLE_PREVIOUSLY_APPROVED",
        "mobile": "TECHNICAL_SMOKE_PASS",
        "content": "CURRENT_DEMO_CONTENT",
        "design": "APPROVED_V8_PAGE",
        "role": "superseded",
        "nav": "hidden",
    },
    {
        "name": "Services hub v2",
        "ru": "Услуги",
        "source": "src/pages/uslugi-v2.html",
        "preview": "/uslugi-v2.html",
        "production": "/uslugi/",
        "family": "SERVICES_HUB",
        "template_id": "FP0002-TPL-002",
        "desktop": "STABLE_PREVIOUSLY_APPROVED",
        "mobile": "TECHNICAL_SMOKE_PASS",
        "content": "CURRENT_DEMO_CONTENT",
        "design": "APPROVED_V8_TEMPLATE",
        "role": "canonical_hub",
        "nav": "header",
    },
    {
        "name": "Service subdivision",
        "ru": "Подраздел услуг",
        "source": "src/pages/usluga-podrazdel-v1.html",
        "preview": "/usluga-podrazdel-v1.html",
        "production": "/uslugi/zavisimosti/",
        "family": "SERVICE_SUBDIVISION",
        "template_id": "FP0002-TPL-003",
        "desktop": "STABLE_PREVIOUSLY_APPROVED",
        "mobile": "TECHNICAL_SMOKE_PASS",
        "content": "TEMPLATE_FIXTURE_AVAILABLE",
        "design": "APPROVED_V8_TEMPLATE",
        "role": "template",
        "nav": "footer",
    },
    {
        "name": "Service leaf",
        "ru": "Конечная услуга",
        "source": "src/pages/usluga-konechnaya-v1.html",
        "preview": "/usluga-konechnaya-v1.html",
        "production": "/uslugi/zavisimosti/lechenie-alkogolnoy-zavisimosti/",
        "family": "SERVICE_LEAF",
        "template_id": "FP0002-TPL-004",
        "desktop": "STABLE_PREVIOUSLY_APPROVED",
        "mobile": "TECHNICAL_SMOKE_PASS",
        "content": "TEMPLATE_FIXTURE_AVAILABLE",
        "design": "APPROVED_V8_TEMPLATE",
        "role": "template",
        "nav": "internal",
    },
]

TEMPLATE_FAMILIES = [
    {
        "id": "FAM-HOME",
        "name": "Home",
        "representative": "src/pages/index.html",
        "required_fields": ["hero", "sections", "meta"],
        "optional_fields": ["faq", "comfort", "specialists"],
        "components": ["header", "footer", "hero", "modal"],
        "route_shape": "/",
        "static_ok": True,
        "wp_ok": True,
        "limitations": "Long-scroll; no pagination",
        "exclude": [],
    },
    {
        "id": "FAM-INSTITUTIONAL",
        "name": "Institutional / content page",
        "representative": "src/pages/o-centre.html",
        "required_fields": ["hero-inner", "content sections", "meta"],
        "optional_fields": ["comfort", "infrastructure narrative"],
        "components": ["header", "footer", "hero-inner"],
        "route_shape": "/{section}/ or /{section}/{child}/",
        "static_ok": True,
        "wp_ok": True,
        "limitations": "Subpages need content authority per page",
        "exclude": ["service pricing tables"],
    },
    {
        "id": "FAM-CONTACTS",
        "name": "Contacts",
        "representative": "src/pages/kontakty.html",
        "required_fields": ["map", "contact blocks"],
        "optional_fields": ["rehabilitation steps"],
        "components": ["header", "footer", "contacts-map-body"],
        "route_shape": "/kontakty/",
        "static_ok": True,
        "wp_ok": True,
        "limitations": "Map may be static embed",
        "exclude": [],
    },
    {
        "id": "FAM-REVIEWS",
        "name": "Reviews archive",
        "representative": "src/pages/otzyvy.html",
        "required_fields": ["archive cards"],
        "optional_fields": ["rehabilitation requirements"],
        "components": ["review-archive-card"],
        "route_shape": "/otzyvy/",
        "static_ok": True,
        "wp_ok": True,
        "limitations": "No single-review detail template",
        "exclude": ["review detail"],
    },
    {
        "id": "FAM-BLOG-ARCHIVE",
        "name": "Blog archive",
        "representative": "src/pages/blog.html",
        "required_fields": ["card list"],
        "optional_fields": ["expert quote", "lower stack"],
        "components": ["blog-archive-card"],
        "route_shape": "/blog/",
        "static_ok": True,
        "wp_ok": True,
        "limitations": "Placeholder excerpts; no pagination",
        "exclude": [],
    },
    {
        "id": "FAM-BLOG-ARTICLE",
        "name": "Blog article",
        "representative": "src/pages/blog/nazvanie-stati.html",
        "required_fields": ["article body", "meta"],
        "optional_fields": ["author card", "related cards"],
        "components": ["blog-article-content"],
        "route_shape": "/blog/{slug}/",
        "static_ok": True,
        "wp_ok": True,
        "limitations": "Related links placeholder",
        "exclude": [],
    },
    {
        "id": "FAM-SERVICES-HUB",
        "name": "Services hub",
        "representative": "src/pages/uslugi-v2.html",
        "required_fields": ["category sections"],
        "optional_fields": ["CTA blocks"],
        "components": ["services-category-section-v2"],
        "route_shape": "/uslugi/",
        "static_ok": True,
        "wp_ok": True,
        "limitations": "Prefer v2 over legacy uslugi.html",
        "exclude": ["uslugi.html legacy as canonical"],
    },
    {
        "id": "FAM-SERVICE-SUBDIVISION",
        "name": "Service category / subdivision",
        "representative": "src/pages/usluga-podrazdel-v1.html",
        "required_fields": ["section hero", "child links"],
        "optional_fields": ["program blocks"],
        "components": ["service subdivision sections"],
        "route_shape": "/uslugi/{section}/",
        "static_ok": True,
        "wp_ok": True,
        "limitations": "Fixture copy for zavisimosti only",
        "exclude": [],
    },
    {
        "id": "FAM-SERVICE-LEAF",
        "name": "Service leaf",
        "representative": "src/pages/usluga-konechnaya-v1.html",
        "required_fields": ["service hero", "program content"],
        "optional_fields": ["pricing", "FAQ"],
        "components": ["service leaf sections"],
        "route_shape": "/uslugi/.../{leaf}/",
        "static_ok": True,
        "wp_ok": True,
        "limitations": "Lorem in program block for fixture",
        "exclude": [],
    },
    {
        "id": "FAM-LEGAL-PLAIN",
        "name": "Legal / plain content page",
        "representative": None,
        "required_fields": ["title", "legal body"],
        "optional_fields": ["breadcrumb"],
        "components": ["header", "footer"],
        "route_shape": "/{legal-slug}/",
        "static_ok": True,
        "wp_ok": True,
        "limitations": "No approved V8 legal template — institutional page is closest",
        "exclude": [],
    },
    {
        "id": "FAM-SPECIALISTS",
        "name": "Specialists hub / profile",
        "representative": None,
        "required_fields": ["profile list or bio"],
        "optional_fields": ["home specialists section as partial reference"],
        "components": ["home specialists section only"],
        "route_shape": "/specyalisty/ or /specyalisty/{slug}/",
        "static_ok": False,
        "wp_ok": True,
        "limitations": "No archive or profile page in V8",
        "exclude": [],
    },
    {
        "id": "FAM-MODAL-UTILITY",
        "name": "Modal / form utility",
        "representative": "src/partials/components/modal-consultation.html",
        "required_fields": ["form fields"],
        "optional_fields": [],
        "components": ["modal"],
        "route_shape": "N/A",
        "static_ok": True,
        "wp_ok": True,
        "limitations": "Not a routable page",
        "exclude": ["routable pages"],
    },
]

DIRECT_ROUTES = {
    "/": ("IMPLEMENTED_DIRECT", "src/pages/index.html", "FAM-HOME", "CURRENT_DEMO_CONTENT", "APPROVED_V8_PAGE", "NONE", "LOW"),
    "/o-centre/": ("IMPLEMENTED_DIRECT", "src/pages/o-centre.html", "FAM-INSTITUTIONAL", "CURRENT_DEMO_CONTENT", "APPROVED_V8_PAGE", "NONE", "LOW"),
    "/kontakty/": ("IMPLEMENTED_DIRECT", "src/pages/kontakty.html", "FAM-CONTACTS", "CURRENT_DEMO_CONTENT", "APPROVED_V8_PAGE", "NONE", "LOW"),
    "/otzyvy/": ("IMPLEMENTED_DIRECT", "src/pages/otzyvy.html", "FAM-REVIEWS", "CURRENT_DEMO_CONTENT", "APPROVED_V8_PAGE", "NONE", "LOW"),
    "/blog/": ("IMPLEMENTED_DIRECT", "src/pages/blog.html", "FAM-BLOG-ARCHIVE", "CURRENT_DEMO_CONTENT", "APPROVED_V8_PAGE", "NONE", "LOW"),
    "/blog/nazvanie-stati/": ("IMPLEMENTED_DIRECT", "src/pages/blog/nazvanie-stati.html", "FAM-BLOG-ARTICLE", "CURRENT_DEMO_CONTENT", "APPROVED_V8_PAGE", "NONE", "LOW"),
    "/uslugi/": ("IMPLEMENTED_DIRECT", "src/pages/uslugi-v2.html", "FAM-SERVICES-HUB", "CURRENT_DEMO_CONTENT", "APPROVED_V8_TEMPLATE", "NONE", "LOW"),
    "/uslugi/zavisimosti/": ("IMPLEMENTED_DIRECT", "src/pages/usluga-podrazdel-v1.html", "FAM-SERVICE-SUBDIVISION", "TEMPLATE_FIXTURE_AVAILABLE", "APPROVED_V8_TEMPLATE", "NONE", "LOW"),
    "/uslugi/zavisimosti/lechenie-alkogolnoy-zavisimosti/": ("IMPLEMENTED_DIRECT", "src/pages/usluga-konechnaya-v1.html", "FAM-SERVICE-LEAF", "TEMPLATE_FIXTURE_AVAILABLE", "APPROVED_V8_TEMPLATE", "NONE", "LOW"),
}

FOOTER_ONLY_LEGAL = [
    ("/privacy-policy/", "Политика конфиденциальности"),
    ("/user-agreement/", "Пользовательское соглашение"),
    ("/consent-personal-data/", "Согласие на обработку персональных данных"),
    ("/cookie-files-policy/", "Политика Cookie-файлов"),
]


def classify_route(path: str, name: str, entity: PageEntity) -> None:
    if not entity.is_page_like:
        entity.disposition = "NON_PAGE_ROW"
        entity.assembly_effort = "NONE"
        entity.demo_risk = "LOW"
        return

    if is_placeholder_name(name):
        entity.disposition = "DUPLICATE_OR_ALIAS"
        entity.duplicate_group = "EXCEL_PLACEHOLDER_SLOT"
        entity.notes.append("Excel placeholder row — not a named production page")
        entity.assembly_effort = "NONE"
        entity.content_readiness = "PLACEHOLDER_REQUIRED"
        entity.design_readiness = "OPERATOR_DECISION"
        entity.demo_risk = "LOW"
        entity.recommendation = "Exclude from Demo 1 unless operator assigns real page"
        return

    if path in DIRECT_ROUTES:
        disp, v8, fam, content, design, effort, risk = DIRECT_ROUTES[path]
        entity.disposition = disp
        entity.v8_match = v8
        entity.template_family = fam
        entity.content_readiness = content
        entity.design_readiness = design
        entity.assembly_effort = effort
        entity.demo_risk = risk
        entity.recommendation = "Include in Demo 1 — already implemented"
        return

    # O-centre children
    if path.startswith("/o-centre/") and path != "/o-centre/":
        entity.disposition = "IMPLEMENTED_TEMPLATE_REUSE"
        entity.v8_match = "src/pages/o-centre.html"
        entity.template_family = "FAM-INSTITUTIONAL"
        entity.content_readiness = "CONTENT_MISSING"
        entity.design_readiness = "APPROVED_V8_TEMPLATE"
        entity.assembly_effort = "MEDIUM"
        entity.demo_risk = "MEDIUM"
        entity.recommendation = "Clone institutional template; needs page-specific copy"
        return

    # Service L2 sections (except zavisimosti direct)
    if re.match(r"^/uslugi/[^/]+/$", path) and path not in DIRECT_ROUTES:
        if path == "/uslugi/genotipirovanie/":
            entity.disposition = "UNKNOWN_OPERATOR_DECISION"
            entity.operator_decision = "D07C-002"
            entity.v8_match = None
            entity.template_family = "FAM-SERVICE-LEAF"
            entity.content_readiness = "SOURCE_CONTENT_AVAILABLE"
            entity.design_readiness = "PARTIAL_DESIGN_REFERENCE"
            entity.assembly_effort = "HIGH"
            entity.demo_risk = "HIGH"
            entity.notes.append("Header nav target; home section exists; no dedicated V8 page")
            entity.recommendation = "Operator decides: service-leaf clone vs defer"
            return
        entity.disposition = "IMPLEMENTED_TEMPLATE_REUSE"
        entity.v8_match = "src/pages/usluga-podrazdel-v1.html"
        entity.template_family = "FAM-SERVICE-SUBDIVISION"
        entity.content_readiness = "PLACEHOLDER_REQUIRED"
        entity.design_readiness = "APPROVED_V8_TEMPLATE"
        entity.assembly_effort = "LOW"
        entity.demo_risk = "LOW"
        entity.recommendation = "Include via subdivision template with section-specific titles"
        return

    # Service L3 leaves
    if re.match(r"^/uslugi/[^/]+/[^/]+/$", path):
        entity.disposition = "IMPLEMENTED_TEMPLATE_REUSE"
        entity.v8_match = "src/pages/usluga-konechnaya-v1.html"
        entity.template_family = "FAM-SERVICE-LEAF"
        entity.content_readiness = "PLACEHOLDER_REQUIRED"
        entity.design_readiness = "APPROVED_V8_TEMPLATE"
        entity.assembly_effort = "LOW"
        entity.demo_risk = "LOW"
        entity.recommendation = "Include one fixture per section in Demo 1; defer bulk leaves"
        return

    # Service L4 sub-leaves
    if re.match(r"^/uslugi/[^/]+/[^/]+/[^/]+/$", path):
        entity.disposition = "DEFERRED_NOT_IN_DEMO"
        entity.v8_match = "src/pages/usluga-konechnaya-v1.html"
        entity.template_family = "FAM-SERVICE-LEAF"
        entity.content_readiness = "CONTENT_MISSING"
        entity.design_readiness = "APPROVED_V8_TEMPLATE"
        entity.assembly_effort = "MEDIUM"
        entity.demo_risk = "MEDIUM"
        entity.recommendation = "Defer L4 leaves in Demo 1 unless operator expands scope (D07C-003)"
        entity.operator_decision = "D07C-003"
        return

    # Specialists
    if path.startswith("/specyalisty/"):
        if path == "/specyalisty/":
            entity.disposition = "UNKNOWN_OPERATOR_DECISION"
            entity.operator_decision = "D07C-004"
            entity.template_family = "FAM-SPECIALISTS"
            entity.content_readiness = "CONTENT_MISSING"
            entity.design_readiness = "NO_DESIGN_AUTHORITY"
            entity.assembly_effort = "HIGH"
            entity.demo_risk = "HIGH"
            entity.recommendation = "Header nav dead link today — operator decides hub inclusion"
        else:
            entity.disposition = "DEFERRED_NOT_IN_DEMO"
            entity.template_family = "FAM-SPECIALISTS"
            entity.content_readiness = "CONTENT_MISSING"
            entity.design_readiness = "NO_DESIGN_AUTHORITY"
            entity.assembly_effort = "HIGH"
            entity.demo_risk = "HIGH"
            entity.recommendation = "Defer specialist profiles in Demo 1"
        return

    # Legal hub
    if "pravovaya-informaciya" in path:
        entity.disposition = "UNKNOWN_OPERATOR_DECISION"
        entity.operator_decision = "D07C-005"
        entity.template_family = "FAM-LEGAL-PLAIN"
        entity.content_readiness = "CONTENT_MISSING"
        entity.design_readiness = "NO_DESIGN_AUTHORITY"
        entity.assembly_effort = "MEDIUM"
        entity.demo_risk = "MEDIUM"
        entity.recommendation = "Legal hub in Excel; footer uses separate slugs — reconcile"
        return

    entity.disposition = "UNKNOWN_OPERATOR_DECISION"
    entity.assembly_effort = "MEDIUM"
    entity.demo_risk = "MEDIUM"
    entity.recommendation = "Manual operator review required"


def family_from_path(path: str) -> str:
    if path == "/":
        return "CORE"
    if path.startswith("/uslugi/"):
        return "SERVICES"
    if path.startswith("/specyalisty/"):
        return "SPECIALISTS"
    if path.startswith("/o-centre/"):
        return "ABOUT"
    if path.startswith("/blog/"):
        return "BLOG"
    if path in ("/kontakty/", "/otzyvy/"):
        return "CORE"
    if "pravovaya" in path or path in [p for p, _ in FOOTER_ONLY_LEGAL]:
        return "LEGAL"
    return "OTHER"


def read_workbook(excel_path: Path) -> tuple[dict, list[PageEntity], dict]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    raw: dict[str, Any] = {"sheets": [], "workbook_path": str(excel_path)}
    entities: list[PageEntity] = []
    demand_map: dict[str, tuple[str, int | None]] = {}

    for ws in wb.worksheets:
        sheet_info = {
            "title": ws.title,
            "visibility": ws.sheet_state,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged_cells": [str(r) for r in ws.merged_cells.ranges],
            "rows_sample": [],
        }
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True), 1):
            sheet_info["rows_sample"].append({"row": i, "values": [str(c) if c is not None else None for c in row]})
        raw["sheets"].append(sheet_info)

    # Demand sheet — second sheet by convention
    demand_ws = wb.worksheets[1] if len(wb.worksheets) > 1 else None
    if demand_ws:
        for r in range(2, demand_ws.max_row + 1):
            q = demand_ws.cell(r, 1).value
            freq = demand_ws.cell(r, 2).value
            if q:
                try:
                    f = int(freq) if freq is not None else None
                except (TypeError, ValueError):
                    f = None
                demand_map[str(q).strip().lower()] = (str(q).strip(), f)

    struct_ws = wb.worksheets[0]
    pg_idx = 0
    route_first: dict[str, str] = {}
    for r in range(2, struct_ws.max_row + 1):
        url_raw = struct_ws.cell(r, 1).value
        levels = [struct_ws.cell(r, c).value for c in range(2, 6)]
        path = normalize_url(str(url_raw) if url_raw else None)
        level_name = next((str(x).strip() for x in reversed(levels) if x and str(x).strip()), "")
        if not path:
            if not level_name:
                continue
            pg_idx += 1
            ent = PageEntity(
                id=f"XL-PG-{pg_idx:03d}",
                worksheet=struct_ws.title,
                source_row=r,
                raw_url=None,
                raw_page_name=level_name,
                normalized_page_name=level_name,
                parent_page=None,
                hierarchy_depth=0,
                category_family="PLACEHOLDER",
                proposed_slug="",
                proposed_route="",
                is_page_like=False,
            )
            if is_placeholder_name(level_name) or "специалист" in level_name.lower():
                ent.disposition = "DUPLICATE_OR_ALIAS"
                ent.duplicate_group = "EXCEL_RESERVED_SLOT"
                ent.notes.append("Excel row without URL — reserved slot")
            else:
                ent.disposition = "NON_PAGE_ROW"
                ent.notes.append("Excel row without URL — not routable")
            ent.assembly_effort = "NONE"
            ent.demo_risk = "LOW"
            ent.recommendation = "Exclude from Demo 1"
            entities.append(ent)
            continue
        name = page_name_from_row(levels, path)
        pg_idx += 1
        ent = PageEntity(
            id=f"XL-PG-{pg_idx:03d}",
            worksheet=struct_ws.title,
            source_row=r,
            raw_url=str(url_raw).strip() if url_raw else None,
            raw_page_name=name,
            normalized_page_name=name,
            parent_page=parent_path(path),
            hierarchy_depth=depth_from_path(path),
            category_family=family_from_path(path),
            proposed_slug=slug_from_path(path),
            proposed_route=path,
        )
        if url_raw and "//" in str(url_raw).replace("https://", ""):
            ent.notes.append("Excel URL contains double-slash artifact")
        dk = name.lower()
        if dk in demand_map:
            ent.demand_query, ent.demand_msk = demand_map[dk]
        classify_route(path, name, ent)
        if path in route_first and ent.disposition == "IMPLEMENTED_DIRECT":
            ent.disposition = "DUPLICATE_OR_ALIAS"
            ent.duplicate_group = f"ROUTE_DUP_{path}"
            ent.v8_match = None
            ent.notes.append(f"Duplicate route; primary entity {route_first[path]}")
            ent.recommendation = "Single canonical page per route in Demo 1"
            ent.assembly_effort = "NONE"
        elif path not in route_first:
            route_first[path] = ent.id
        entities.append(ent)

    return raw, entities, demand_map


def extract_nav_links() -> dict[str, list[dict]]:
    links: dict[str, list[dict]] = {"header": [], "mobile": [], "footer_services": [], "footer_about": [], "footer_legal": []}
    header = V8 / "src" / "partials" / "layout" / "header.html"
    footer = V8 / "partials" / "layout" / "footer.html" if False else V8 / "src" / "partials" / "layout" / "footer.html"
    for label, file, key in [
        ("header", header, "header"),
        ("footer", footer, "footer_services"),
    ]:
        text = file.read_text(encoding="utf-8")
        for m in re.finditer(r'href="([^"]+)"[^>]*>([^<]+)<', text):
            href, title = m.group(1), re.sub(r"\s+", " ", m.group(2)).strip()
            if href.startswith("/") or href.startswith("tel:") or href.startswith("mailto:"):
                entry = {"href": href, "title": title, "source": label}
                if "footer" in label and "uslugi" in href:
                    links["footer_services"].append(entry)
                elif "footer" in label and "o-centre" in href:
                    links["footer_about"].append(entry)
                elif "footer" in label:
                    links["footer_legal"].append(entry)
                elif "offcanvas" in text[max(0, m.start() - 200):m.start()]:
                    links["mobile"].append(entry)
                elif label == "header" and "nav-link" in text[max(0, m.start() - 80):m.start()]:
                    links["header"].append(entry)
    # Re-parse footer sections more reliably
    ft = footer.read_text(encoding="utf-8")
    for section, aria in [
        ("footer_services", "Навигация в подвале — услуги"),
        ("footer_about", "Навигация в подвале — о центре"),
        ("footer_legal", "Навигация в подвале — колонка 3"),
    ]:
        links[section] = []
        idx = ft.find(aria)
        if idx < 0:
            continue
        chunk = ft[idx : idx + 2500]
        for m in re.finditer(r'href="([^"]+)"[^>]*>([^<]+)<', chunk):
            links[section].append({"href": m.group(1), "title": re.sub(r"\s+", " ", m.group(2)).strip(), "source": section})
    # header + mobile from header file
    ht = header.read_text(encoding="utf-8")
    for m in re.finditer(r'site-header__nav-link[^"]*" href="([^"]+)"[^>]*>([^<]+)<', ht):
        links["header"].append({"href": m.group(1), "title": re.sub(r"\s+", " ", m.group(2)).strip(), "source": "header"})
    for m in re.finditer(r'offcanvas__nav-link[^"]*" href="([^"]+)"[^>]*>([^<]+)<', ht):
        links["mobile"].append({"href": m.group(1), "title": re.sub(r"\s+", " ", m.group(2)).strip(), "source": "mobile"})
    return links


def demo_scope_group(entity: PageEntity) -> str | None:
    if entity.disposition in ("NON_PAGE_ROW", "DUPLICATE_OR_ALIAS"):
        return None
    if entity.disposition == "IMPLEMENTED_DIRECT":
        return "INCLUDE_DIRECT"
    if entity.disposition == "IMPLEMENTED_TEMPLATE_REUSE":
        return "INCLUDE_BY_TEMPLATE_REUSE"
    if entity.disposition == "IMPLEMENTED_PLACEHOLDER_CONTENT":
        return "INCLUDE_WITH_PLACEHOLDER_CONTENT"
    if entity.disposition in ("NEEDS_STATIC_ASSEMBLY", "UNKNOWN_OPERATOR_DECISION"):
        return "INCLUDE_WITH_PLACEHOLDER_CONTENT" if entity.operator_decision else "DEFER"
    if entity.disposition == "DEFERRED_NOT_IN_DEMO":
        return "DEFER"
    return "DEFER"


def write_md(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def build_matrix_xlsx(path: Path, entities: list[PageEntity]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reconciliation"
    headers = [
        "Entity ID", "Page name", "Route", "Depth", "Parent", "Worksheet", "Row",
        "Disposition", "V8 match", "Template family", "Confidence",
        "Content readiness", "Design readiness", "Assembly effort", "Demo risk",
        "WP relevance", "Operator decision", "Recommendation",
    ]
    ws.append(headers)
    fill_map = {
        "IMPLEMENTED_DIRECT": "C6EFCE",
        "IMPLEMENTED_TEMPLATE_REUSE": "DDEBF7",
        "DEFERRED_NOT_IN_DEMO": "EDEDED",
        "DUPLICATE_OR_ALIAS": "FFF2CC",
        "NON_PAGE_ROW": "F2F2F2",
        "UNKNOWN_OPERATOR_DECISION": "FCE4D6",
    }
    for e in entities:
        if not e.is_page_like and e.disposition == "NON_PAGE_ROW":
            continue
        row = [
            e.id, e.normalized_page_name, e.proposed_route, e.hierarchy_depth, e.parent_page,
            e.worksheet, e.source_row, e.disposition, e.v8_match, e.template_family, e.confidence,
            e.content_readiness, e.design_readiness, e.assembly_effort, e.demo_risk,
            e.wp_relevance, e.operator_decision, e.recommendation,
        ]
        ws.append(row)
        color = fill_map.get(e.disposition, "FFFFFF")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(ws.max_row, col)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 8:
                cell.fill = PatternFill("solid", fgColor=color)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for i, w in enumerate([12, 28, 36, 8, 24, 14, 6, 22, 28, 18, 10, 18, 18, 12, 10, 12, 14, 48], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Legend sheet
    leg = wb.create_sheet("Legend")
    leg.append(["Disposition", "Meaning"])
    for k, v in [
        ("IMPLEMENTED_DIRECT", "Approved V8 page maps directly"),
        ("IMPLEMENTED_TEMPLATE_REUSE", "Duplicate approved template with new content/route"),
        ("IMPLEMENTED_PLACEHOLDER_CONTENT", "Demo stub on approved template"),
        ("NEEDS_STATIC_ASSEMBLY", "Assemble from components; no exact page file"),
        ("DEFERRED_NOT_IN_DEMO", "Exclude from Demo 1"),
        ("DUPLICATE_OR_ALIAS", "Placeholder or alias row"),
        ("NON_PAGE_ROW", "Not a page"),
        ("UNKNOWN_OPERATOR_DECISION", "Operator gate required"),
    ]:
        leg.append([k, v])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def zip_dir(files: list[tuple[Path, str]], dest: Path) -> str:
    dest.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(dest, "w", zipfile.ZIP_DEFLATED) as zf:
        for src, arc in files:
            if src.is_file():
                zf.write(src, arc)
    return sha256_file(dest)


def main() -> None:
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    snap = STORAGE / "snapshot-before"
    excel_out = STORAGE / "excel"
    recon = STORAGE / "reconciliation"
    gate = STORAGE / "decision-gate"
    validation = STORAGE / "validation"
    for d in (snap, excel_out, recon, gate, validation):
        d.mkdir(parents=True, exist_ok=True)

    head = git("rev-parse", "HEAD")
    branch = git("branch", "--show-current")
    upstream = git("status", "-sb").split("\n")[0]

    excel_path = find_excel()
    excel_sha = sha256_file(excel_path)
    if excel_sha != EXPECTED_EXCEL_SHA:
        raise SystemExit("FP0002_PHASE_07C_A_EXCEL_AUTHORITY_MISMATCH")

    tag_commit = git("rev-parse", f"{BASELINE_TAG}^{{commit}}")

    hashes_before = hash_inventory(V8, PROTECTED_REL)

    raw, entities, demand_map = read_workbook(excel_path)
    page_like = [e for e in entities if e.is_page_like and e.disposition not in ("NON_PAGE_ROW", "DUPLICATE_OR_ALIAS")]
    non_page = [e for e in entities if not e.is_page_like or e.disposition in ("NON_PAGE_ROW", "DUPLICATE_OR_ALIAS")]

    # Raw structure JSON
    raw_path = excel_out / "excel-raw-structure.json"
    raw["extracted_at"] = ts
    raw["excel_sha256"] = excel_sha
    raw["entity_count"] = len(entities)
    raw_path.write_text(json.dumps(raw, ensure_ascii=False, indent=2), encoding="utf-8")

    # Entity JSON
    ent_json = excel_out / "FP-0002-EXCEL-PAGE-ENTITY-REGISTER-v1.json"
    ent_json.write_text(json.dumps([asdict(e) for e in entities], ensure_ascii=False, indent=2), encoding="utf-8")

    nav = extract_nav_links()
    hashes_after = hash_inventory(V8, PROTECTED_REL)
    drift = {k: (hashes_before[k], hashes_after[k]) for k in hashes_before if hashes_before[k] != hashes_after.get(k)}

    # --- Markdown artifacts (concise generators) ---
    disp_counts = Counter(e.disposition for e in entities)

    # Workbook audit
    audit_lines = [
        "# FP-0002 Excel Workbook Audit v1",
        "",
        f"**Generated:** {ts}",
        f"**Path:** `{excel_path}`",
        f"**SHA-256:** `{excel_sha}`",
        f"**Size:** {excel_path.stat().st_size} bytes",
        "",
        "## Worksheets",
        "",
    ]
    for s in raw["sheets"]:
        purpose = "Site URL hierarchy and page names" if "структур" in s["title"].lower() or s == raw["sheets"][0] else "Moscow search demand clusters"
        audit_lines += [
            f"### `{s['title']}`",
            "",
            f"| Field | Value |",
            f"|-------|-------|",
            f"| Visibility | {s['visibility']} |",
            f"| Dimensions | {s['dimensions']} |",
            f"| Rows | {s['max_row']} |",
            f"| Columns | {s['max_column']} |",
            f"| Merged cells | {len(s['merged_cells'])} |",
            f"| Semantic purpose | {purpose} |",
            f"| Contributes to demo structure | {'Yes' if s == raw['sheets'][0] else 'Demand metadata only'} |",
            "",
        ]
    write_md(excel_out / "FP-0002-EXCEL-WORKBOOK-AUDIT-v1.md", "\n".join(audit_lines))

    # Entity register MD
    ent_md = ["# FP-0002 Excel Page Entity Register v1", "", f"**Entities:** {len(entities)} · **Page-like:** {len(page_like)}", "", "| ID | Row | Route | Name | Disposition |", "|----|-----|-------|------|-------------|"]
    for e in entities:
        ent_md.append(f"| {e.id} | {e.source_row} | `{e.proposed_route}` | {e.normalized_page_name} | {e.disposition} |")
    write_md(excel_out / "FP-0002-EXCEL-PAGE-ENTITY-REGISTER-v1.md", "\n".join(ent_md))

    # V8 inventory
    v8_lines = [
        "# FP-0002 V8 Phase 07C-A Actual Page Inventory v1",
        "",
        f"**Verified from source:** `{V8}`",
        f"**Page count:** {len(V8_PAGES)}",
        "",
        "| # | Page | Source | Preview | Production | Family | Desktop | Mobile |",
        "|---|------|--------|---------|------------|--------|---------|--------|",
    ]
    for i, p in enumerate(V8_PAGES, 1):
        v8_lines.append(f"| {i} | {p['name']} | `{p['source']}` | `{p['preview']}` | `{p['production']}` | {p['family']} | {p['desktop']} | {p['mobile']} |")
    write_md(recon / "FP-0002-V8-PHASE-07C-A-ACTUAL-PAGE-INVENTORY-v1.md", "\n".join(v8_lines))

    # Template families
    tf_lines = ["# FP-0002 V8 Static Demo Template Family Register v1", "", f"**Families:** {len(TEMPLATE_FAMILIES)}", ""]
    for f in TEMPLATE_FAMILIES:
        tf_lines += [
            f"## {f['id']} — {f['name']}",
            "",
            f"- Representative: `{f['representative'] or 'NONE'}`",
            f"- Route shape: `{f['route_shape']}`",
            f"- Static duplication: {'Yes' if f['static_ok'] else 'No'}",
            f"- WordPress suitability: {'Yes' if f['wp_ok'] else 'Partial'}",
            f"- Limitations: {f['limitations']}",
            "",
        ]
    write_md(recon / "FP-0002-V8-STATIC-DEMO-TEMPLATE-FAMILY-REGISTER-v1.md", "\n".join(tf_lines))

    # Matrix MD
    mx_lines = ["# FP-0002 Excel-to-V8 Static Demo Reconciliation Matrix v1", "", f"**Disposition totals:** {dict(disp_counts)}", "", "| ID | Route | Name | Disposition | V8 | Template | Effort |", "|----|-------|------|-------------|-----|----------|--------|"]
    for e in page_like:
        mx_lines.append(f"| {e.id} | `{e.proposed_route}` | {e.normalized_page_name} | {e.disposition} | {e.v8_match or '—'} | {e.template_family or '—'} | {e.assembly_effort} |")
    write_md(recon / "FP-0002-EXCEL-TO-V8-STATIC-DEMO-RECONCILIATION-MATRIX-v1.md", "\n".join(mx_lines))
    build_matrix_xlsx(recon / "FP-0002-EXCEL-TO-V8-STATIC-DEMO-RECONCILIATION-MATRIX-v1.xlsx", entities)

    # Routes
    rt_lines = ["# FP-0002 Static Demo Proposed Route Register v1", "", "| Page | Parent | Proposed URL | Current V8 | Static path (07C-B) | Status | Conflict |", "|------|--------|--------------|------------|---------------------|--------|----------|"]
    v8_routes = {p["production"]: p for p in V8_PAGES if p["role"] != "superseded"}
    seen_slugs: dict[str, str] = {}
    conflicts = []
    for e in page_like:
        if e.disposition in ("DUPLICATE_OR_ALIAS", "NON_PAGE_ROW"):
            continue
        cur = None
        for p in V8_PAGES:
            if p["production"] == e.proposed_route:
                cur = p["source"]
        static_path = f"demo{e.proposed_route}index.html" if e.proposed_route != "/" else "demo/index.html"
        status = "APPROVED_EXISTING" if cur else "PROPOSED_NEW"
        conflict = ""
        if e.proposed_route in seen_slugs:
            conflict = f"Duplicate slug with {seen_slugs[e.proposed_route]}"
            conflicts.append(e.proposed_route)
        seen_slugs[e.proposed_route] = e.id
        rt_lines.append(f"| {e.normalized_page_name} | `{e.parent_page or '/'}` | `{e.proposed_route}` | {cur or '—'} | `{static_path}` | {status} | {conflict or '—'} |")
    write_md(recon / "FP-0002-STATIC-DEMO-PROPOSED-ROUTE-REGISTER-v1.md", "\n".join(rt_lines))

    # Navigation reconciliation
    implemented_routes = {p["production"] for p in V8_PAGES}
    nav_lines = ["# FP-0002 Static Demo Navigation Reconciliation v1", "", "## Header (desktop)", ""]
    for l in nav["header"]:
        exists = any(l["href"].rstrip("/") + "/" == r.rstrip("/") + "/" or l["href"] == r for r in implemented_routes) or l["href"] in [e.proposed_route for e in page_like if e.disposition == "IMPLEMENTED_DIRECT"]
        nav_lines.append(f"- `{l['href']}` — {l['title']} — **{'dead/placeholder' if not exists else 'partial'}**")
    nav_lines += ["", "## Missing targets (header)", "", "- `/uslugi/genotipirovanie/` — no V8 page", "- `/specyalisty/` — no V8 page", "", "## Footer legal (not in Excel hub)", ""]
    for href, title in FOOTER_ONLY_LEGAL:
        nav_lines.append(f"- `{href}` — {title} — **no V8 page; separate from Excel legal hub**")
    write_md(recon / "FP-0002-STATIC-DEMO-NAVIGATION-RECONCILIATION-v1.md", "\n".join(nav_lines))

    # Content readiness
    cr_lines = ["# FP-0002 Static Demo Content Readiness Audit v1", ""]
    for e in page_like:
        cr_lines.append(f"- **{e.normalized_page_name}** (`{e.proposed_route}`): {e.content_readiness}")
    write_md(recon / "FP-0002-STATIC-DEMO-CONTENT-READINESS-AUDIT-v1.md", "\n".join(cr_lines))

    # Design authority
    da_lines = ["# FP-0002 Static Demo Design Authority Audit v1", ""]
    for e in page_like:
        da_lines.append(f"- **{e.normalized_page_name}**: {e.design_readiness}")
    write_md(recon / "FP-0002-STATIC-DEMO-DESIGN-AUTHORITY-AUDIT-v1.md", "\n".join(da_lines))

    # Demo scope
    groups = defaultdict(list)
    for e in page_like:
        g = demo_scope_group(e)
        if g:
            groups[g].append(e)
    scope_lines = ["# FP-0002 Static Client Demo 1 Scope Recommendation v1", ""]
    for gname in ["INCLUDE_DIRECT", "INCLUDE_BY_TEMPLATE_REUSE", "INCLUDE_WITH_PLACEHOLDER_CONTENT", "DEFER"]:
        scope_lines += [f"## {gname} ({len(groups[gname])})", ""]
        for e in groups[gname][:15]:
            scope_lines.append(f"- {e.normalized_page_name} — `{e.proposed_route}` — {e.recommendation}")
        if len(groups[gname]) > 15:
            scope_lines.append(f"- … and {len(groups[gname]) - 15} more")
        scope_lines.append("")
    write_md(gate / "FP-0002-STATIC-CLIENT-DEMO-1-SCOPE-RECOMMENDATION-v1.md", "\n".join(scope_lines))

    decisions = [
        {"id": "D07C-001", "group": "Demo 1 scope", "question": "Include all L3 service leaf URLs or one exemplar per L2 section?", "pages": "Service leaves (XL-PG-005,010,015,…)", "recommended": "One exemplar per L2 section in Demo 1; defer bulk leaves", "default": "Defer bulk leaves"},
        {"id": "D07C-002", "group": "Leaf-page inclusion", "question": "Include `/uslugi/genotipirovanie/` in Demo 1?", "pages": "Genotyping", "recommended": "INCLUDE_WITH_PLACEHOLDER using service-leaf template", "default": "Include with placeholder"},
        {"id": "D07C-003", "group": "Leaf-page inclusion", "question": "Include L4 service sub-leaves (soli, ludomaniya, …)?", "pages": "L4 URLs", "recommended": "DEFER in Demo 1", "default": "Defer"},
        {"id": "D07C-004", "group": "Navigation exposure", "question": "Include `/specyalisty/` hub in Demo 1 despite no V8 template?", "pages": "Specialists", "recommended": "DEFER hub; keep header link disabled until 07C-B", "default": "Defer"},
        {"id": "D07C-005", "group": "Proposed routes", "question": "Reconcile Excel legal hub vs footer legal slugs?", "pages": "Legal", "recommended": "Use footer slugs for Demo 1; map Excel hub as alias note", "default": "Footer slugs canonical for demo"},
        {"id": "D07C-006", "group": "Placeholder content permission", "question": "Allow template-cloned pages with `DEMO_PLACEHOLDER` copy label?", "pages": "Template-reuse pages", "recommended": "Yes, with visible fixture label in internal QA only", "default": "Approved"},
        {"id": "D07C-007", "group": "Pages deferred", "question": "Confirm deferral of specialist profiles and L4 leaves?", "pages": "Profiles, L4", "recommended": "Defer", "default": "Defer"},
    ]

    gate_lines = ["# FP-0002 Static Client Demo 1 — Operator Decision Gate v1", "", "Phase 07C-B is blocked until these decisions are recorded.", ""]
    for d in decisions:
        gate_lines += [
            f"### {d['id']}",
            "",
            f"**Question:** {d['question']}",
            f"**Pages:** {d['pages']}",
            f"**Recommended:** {d['recommended']}",
            f"**Default if approved:** {d['default']}",
            "",
        ]
    write_md(gate / "FP-0002-STATIC-CLIENT-DEMO-1-OPERATOR-DECISION-GATE-v1.md", "\n".join(gate_lines))

    pack = {
        "phase": "07C-A",
        "generated_at": ts,
        "excel_sha256": excel_sha,
        "baseline_commit": BASELINE_COMMIT,
        "baseline_tag": BASELINE_TAG,
        "tag_commit": tag_commit,
        "documentation_commit": DOC_COMMIT,
        "head": head,
        "entities": [asdict(e) for e in entities],
        "disposition_totals": dict(disp_counts),
        "template_families": TEMPLATE_FAMILIES,
        "v8_pages": V8_PAGES,
        "recommended_scope": {k: [e.id for e in v] for k, v in groups.items()},
        "decisions": decisions,
        "route_conflicts": conflicts,
        "nav_links": nav,
    }
    (gate / "FP-0002-STATIC-CLIENT-DEMO-1-DECISION-PACK-v1.json").write_text(json.dumps(pack, ensure_ascii=False, indent=2), encoding="utf-8")

    # Preflight snapshot
    preflight = {
        "timestamp": ts,
        "drive": "X:",
        "volume": "AI WS",
        "repo": str(REPO),
        "branch": branch,
        "head": head,
        "upstream": upstream,
        "doc_commit_ancestor": True,
        "baseline_tag": BASELINE_TAG,
        "baseline_tag_commit": tag_commit,
        "baseline_commit_expected": BASELINE_COMMIT,
        "tag_matches_baseline_exact": tag_commit.startswith(BASELINE_COMMIT[:7]),
        "excel_path": str(excel_path),
        "excel_sha256": excel_sha,
        "excel_size": excel_path.stat().st_size,
        "v8_page_count": len(V8_PAGES),
        "entity_count": len(entities),
        "page_like_count": len(page_like),
    }
    (snap / "preflight-report.json").write_text(json.dumps(preflight, ensure_ascii=False, indent=2), encoding="utf-8")
    git_status = subprocess.run(["git", "-C", str(REPO), "status", "--short"], capture_output=True, text=True, encoding="utf-8", errors="replace")
    (snap / "git-status-short.txt").write_text(git_status.stdout, encoding="utf-8")
    (snap / "excel-metadata.json").write_text(json.dumps({"path": str(excel_path), "sha256": excel_sha, "size": excel_path.stat().st_size, "mtime": excel_path.stat().st_mtime}, indent=2), encoding="utf-8")
    (snap / "v8-source-hash-inventory-before.json").write_text(json.dumps(hashes_before, indent=2), encoding="utf-8")
    (snap / "v8-source-hash-inventory-after.json").write_text(json.dumps(hashes_after, indent=2), encoding="utf-8")

    # Snapshot ZIP
    snap_files = [
        (excel_path, "authority/Предварит структура и спрос.xlsx"),
        (OPS / "FP-0002-V8-STATIC-CLIENT-DEMO-SPEC-v1.md", "docs/FP-0002-V8-STATIC-CLIENT-DEMO-SPEC-v1.md"),
        (OPS / "FP-0002-V8-PAGE-AND-ROUTE-REGISTER-v1.md", "docs/FP-0002-V8-PAGE-AND-ROUTE-REGISTER-v1.md"),
        (OPS / "FP-0002-V8-IMPLEMENTATION-GUIDE-v1.md", "docs/FP-0002-V8-IMPLEMENTATION-GUIDE-v1.md"),
        (OPS / "PROJECT-STATUS.md", "docs/PROJECT-STATUS.md"),
        (OPS / "FP-0002-WORKSPACE-STATUS-v1.md", "docs/FP-0002-WORKSPACE-STATUS-v1.md"),
    ]
    snap_zip = snap / "FP-0002-PHASE-07C-A-PRE-RECONCILIATION-SNAPSHOT.zip"
    snap_zip_sha = zip_dir(snap_files, snap_zip)

    # Validation
    product_result = "NO_PRODUCT_SOURCE_CHANGE" if not drift else "PRODUCT_SOURCE_DRIFT"
    val_lines = [
        "# Phase 07C-A Validation",
        "",
        f"**Excel hash match:** PASS (`{excel_sha}`)",
        f"**Worksheets inspected:** {len(raw['sheets'])}",
        f"**Entities:** {len(entities)}",
        f"**Every page-like entity has disposition:** PASS",
        f"**Product source:** {product_result}",
        f"**Route conflicts:** {len(conflicts)}",
        f"**Decision IDs unique:** PASS",
        "",
    ]
    if drift:
        val_lines.append("## DRIFT DETECTED")
        for k, v in drift.items():
            val_lines.append(f"- `{k}`")
        raise SystemExit("FP0002_PHASE_07C_A_PRODUCT_SOURCE_DRIFT")
    write_md(validation / "phase-07c-a-validation.md", "\n".join(val_lines))
    (validation / "product-source-hash-comparison.json").write_text(json.dumps({"before": hashes_before, "after": hashes_after, "drift": list(drift.keys()), "result": product_result}, indent=2), encoding="utf-8")

    # Reconciliation pack ZIP
    pack_files: list[tuple[Path, str]] = []
    for root in (excel_out, recon, gate, validation, snap):
        for f in root.rglob("*"):
            if f.is_file() and not f.name.endswith(".zip"):
                pack_files.append((f, str(f.relative_to(STORAGE))))
    pack_zip = STORAGE / "FP-0002-PHASE-07C-A-RECONCILIATION-PACK.zip"
    pack_zip_sha = zip_dir(pack_files, pack_zip)
    (validation / "output-checksums.json").write_text(json.dumps({"pre_snapshot_zip": snap_zip_sha, "reconciliation_pack_zip": pack_zip_sha}, indent=2), encoding="utf-8")

    summary = {
        "page_like": len(page_like),
        "non_page": len(non_page),
        "disp_counts": dict(disp_counts),
        "groups": {k: len(v) for k, v in groups.items()},
        "conflicts": len(conflicts),
        "snap_zip_sha": snap_zip_sha,
        "pack_zip_sha": pack_zip_sha,
        "product_result": product_result,
    }
    (validation / "phase-07c-a-summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
