#!/usr/bin/env python3
"""Read-only Excel intake for FP-0002 static demo pass opening. Does not modify workbook."""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

EXCEL_DIR = Path(
    r"C:\MARS Phenix\AI MARS\workspaces\website-factory-operations\FP-0002-SHPIGOVSKY\INCOMING\02_CONTENT"
)
OUT_DIR = Path(__file__).resolve().parent / "data"
HOME_TITLE = "Шпиговский дом — центр профилактики и лечения зависимостей"
PLACEHOLDER_MSG = "Раздел скоро будет опубликован"

MENU_TOP = {
    "/uslugi/",
    "/uslugi/genotipirovanie/",
    "/specyalisty/",
    "/o-centre/",
    "/otzyvy/",
    "/blog/",
    "/kontakty/",
}

MENU_INFORMATIONAL = {
    "/specyalisty/",
    "/o-centre/",
    "/otzyvy/",
    "/blog/",
    "/kontakty/",
    "/pravovaya-informaciya-pilzovatelyu/",
}


def find_excel() -> Path:
    matches = list(EXCEL_DIR.glob("*.xlsx"))
    if not matches:
        raise FileNotFoundError(f"No xlsx in {EXCEL_DIR}")
    if len(matches) > 1:
        # prefer structure+demand file by size/name pattern
        matches.sort(key=lambda p: (p.stat().st_size, p.name))
    return matches[0]


def sheet_audit(wb) -> list[dict]:
    audits = []
    for name in wb.sheetnames:
        ws = wb[name]
        hidden = "hidden" if ws.sheet_state != "visible" else "visible"
        merged = len(list(ws.merged_cells.ranges))
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        header_row = 1
        headers = [ws.cell(1, c).value for c in range(1, max_col + 1)]
        audits.append(
            {
                "sheet": name,
                "dimensions": f"{max_row}x{max_col}",
                "rows": max_row,
                "columns": max_col,
                "header_row": header_row,
                "headers": headers,
                "hidden": hidden,
                "merged_cells": merged,
                "verdict": "STRUCTURE" if name.strip().lower() == "структура" else "SUPPORTING",
            }
        )
    return audits


def normalize_url(raw: str | None) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    s = s.replace("https://shpigovsky.ru", "").replace("http://shpigovsky.ru", "")
    s = re.sub(r"/+", "/", s)
    if not s.startswith("/"):
        s = "/" + s
    if not s.endswith("/"):
        s = s + "/"
    return s


def slug_from_url(url: str) -> str:
    parts = [p for p in url.strip("/").split("/") if p]
    return parts[-1] if parts else "index"


def transliterate_demo(text: str) -> str:
    table = {
        "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e",
        "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
        "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
        "ф": "f", "х": "h", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "sch",
        "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
    }
    out = []
    for ch in text.lower():
        if ch in table:
            out.append(table[ch])
        elif ch.isalnum():
            out.append(ch)
        elif ch in (" ", "-", "_"):
            out.append("-")
    slug = re.sub(r"-+", "-", "".join(out)).strip("-")
    return slug or "page"


def classify_page(url: str, name: str, depth: int, children: int) -> tuple[str, str, str]:
    """Returns template, evidence, confidence."""
    if url == "/":
        return "HOME_PAGE_TEMPLATE", "EXPLICIT: единственная главная", "HIGH"
    if url == "/uslugi/":
        return "SERVICES_HUB_INTERNAL_PAGE", "EXPLICIT: каталог услуг L1", "HIGH"
    if url == "/uslugi/genotipirovanie/":
        return (
            "SERVICE_SUBDIVISION_INTERNAL_PAGE",
            "EXPLICIT: standalone service direction L2 under /uslugi/",
            "HIGH",
        )
    if url in MENU_INFORMATIONAL:
        return "PLACEHOLDER_PAGE", f"EXPLICIT: top menu informational {url}", "HIGH"
    if url.startswith("/blog/"):
        return "PLACEHOLDER_PAGE", "EXPLICIT: blog/article — нет шаблона", "HIGH"
    if url.startswith("/privacy-policy/") or url.startswith("/user-agreement/") or url.startswith("/consent-personal-data/") or url.startswith("/cookie-files-policy/"):
        return "PLACEHOLDER_PAGE", "INFERRED: legal footer links not in Excel", "MEDIUM"
    if url.startswith("/pravovaya-informaciya"):
        return "PLACEHOLDER_PAGE", "EXPLICIT: legal hub", "HIGH"
    if url.startswith("/o-centre/"):
        return "PLACEHOLDER_PAGE", "EXPLICIT: about sub-pages / hub", "HIGH"
    if url.startswith("/specyalisty/"):
        return "PLACEHOLDER_PAGE", "EXPLICIT: specialists hub/profiles", "HIGH"
    if children > 0 and depth <= 3:
        return "SERVICE_SUBDIVISION_INTERNAL_PAGE", f"EXPLICIT: parent with {children} children", "HIGH"
    if url.startswith("/uslugi/"):
        return "SERVICE_LEAF_INTERNAL_PAGE", "EXPLICIT: terminal service URL", "HIGH"
    return "PLACEHOLDER_PAGE", "INFERRED: non-service page", "MEDIUM"


def parse_structure(ws) -> list[dict]:
    rows = []
    for r in range(2, ws.max_row + 1):
        url_raw = ws.cell(r, 1).value
        levels = [ws.cell(r, c).value for c in range(2, 6)]
        url = normalize_url(url_raw if url_raw else None)
        name = None
        depth = 0
        for i, lv in enumerate(levels, start=1):
            if lv and str(lv).strip():
                name = str(lv).strip()
                depth = i
        if not name and not url:
            continue
        if name and name.lower() == "название":
            note = "PLACEHOLDER_SLOT"
        elif name and re.match(r"специалист\s*\d+", name.lower()):
            note = "PLACEHOLDER_SLOT"
        else:
            note = ""
        if not url and name:
            # reserved rows without URL in Excel
            note = note or "NO_URL_IN_EXCEL"
        rows.append(
            {
                "source_row": f"Структура!{r}",
                "raw_url": str(url_raw).strip() if url_raw else None,
                "url": url,
                "raw_name": name,
                "hierarchy_level": depth,
                "levels": [str(x).strip() if x else None for x in levels],
                "source_notes": note,
                "confidence": "LOW" if note else ("HIGH" if url else "MEDIUM"),
            }
        )
    return rows


def build_hierarchy(rows: list[dict]) -> list[dict]:
    # assign parent by URL prefix / level columns
    by_url = {r["url"]: r for r in rows if r.get("url")}
    for r in rows:
        url = r.get("url")
        if not url or url == "/":
            r["parent_raw"] = None
            r["parent_resolved"] = None
            continue
        parent_url = None
        parts = url.strip("/").split("/")
        for i in range(len(parts) - 1, 0, -1):
            cand = "/" + "/".join(parts[:i]) + "/"
            if cand in by_url:
                parent_url = cand
                break
        if not parent_url and url != "/":
            if len(parts) == 1:
                parent_url = "/"
            elif parts[0] == "uslugi" and len(parts) == 2:
                parent_url = "/uslugi/"
        r["parent_raw"] = parent_url
        r["parent_resolved"] = parent_url
    # children count
    children_map: dict[str, int] = {}
    for r in rows:
        p = r.get("parent_resolved")
        if p:
            children_map[p] = children_map.get(p, 0) + 1
    for r in rows:
        r["children_count"] = children_map.get(r.get("url"), 0)
    return rows


def resolve_url_collisions(rows: list[dict]) -> list[dict]:
    """Assign unique demo URLs where Excel repeats the same slug."""
    seen: dict[str, int] = {}
    for r in rows:
        url = r.get("url")
        if not url:
            continue
        if url not in seen:
            seen[url] = 0
            continue
        seen[url] += 1
        suffix = seen[url]
        base = url.strip("/").split("/")
        base[-1] = f"{base[-1]}-{suffix}"
        r["url"] = "/" + "/".join(base) + "/"
        r["slug_source"] = "DEMO_GENERATED_SLUG"
        r["source_notes"] = (r.get("source_notes") or "") + "; EXCEL_DUPLICATE_URL"
        r["confidence"] = "MEDIUM"

    # rows without URL — infer under parent from level columns
    slot_counters: dict[str, int] = {}
    for r in rows:
        if r.get("url") or not r.get("raw_name"):
            continue
        parent_url = infer_parent_from_levels(r)
        r["parent_resolved"] = parent_url
        key = parent_url or "root"
        slot_counters[key] = slot_counters.get(key, 0) + 1
        slug = transliterate_demo(r["raw_name"]) + f"-slot-{slot_counters[key]:02d}"
        if parent_url:
            r["url"] = parent_url.rstrip("/") + f"/{slug}/"
        else:
            r["url"] = f"/{slug}/"
        r["slug_source"] = "DEMO_GENERATED_SLUG"
        r["source_notes"] = (r.get("source_notes") or "NO_URL_IN_EXCEL") + "; INFERRED_URL"
        r["confidence"] = "LOW"
        r["template"] = "PLACEHOLDER_PAGE"
    return rows


def infer_parent_from_levels(r: dict) -> str | None:
    name = (r.get("raw_name") or "").lower()
    depth = r.get("hierarchy_level") or 0
    if "специалист" in name:
        return "/specyalisty/"
    if depth == 4:
        # zavisimosti reserved L4 slots (rows 16-17)
        return "/uslugi/zavisimosti/lechenie-narkoticheskoy-zavisimosti/"
    if depth == 3 and r.get("source_notes", "").startswith("PLACEHOLDER"):
        return "/uslugi/psihicheskoe-zdorovie/"
    return None


def assign_ids(rows: list[dict]) -> list[dict]:
    def sort_key(r):
        url = r.get("url") or ""
        depth = r.get("hierarchy_level") or 99
        menu_order = [
            "/",
            "/uslugi/",
            "/uslugi/zavisimosti/",
            "/uslugi/psihicheskoe-zdorovie/",
            "/uslugi/rasstroystva-pischevogo-povedeniya/",
            "/uslugi/genotipirovanie/",
        ]
        try:
            mo = menu_order.index(url)
        except ValueError:
            mo = 1000
        return (mo, depth, url)

    ordered = sorted([r for r in rows if r.get("url")], key=sort_key)
    # manual priority pass per task spec
    priority_urls = [
        "/",
        "/uslugi/",
        "/uslugi/genotipirovanie/",
        "/specyalisty/",
        "/o-centre/",
        "/otzyvy/",
        "/blog/",
        "/kontakty/",
        "/pravovaya-informaciya-pilzovatelyu/",
    ]
    rest = [r for r in ordered if r["url"] not in priority_urls]
    pri = [by_url for u in priority_urls for by_url in [next((x for x in ordered if x["url"] == u), None)] if by_url]
    rest_sorted = sorted(rest, key=lambda r: (r.get("hierarchy_level", 99), r.get("url", "")))
    final_order = pri + rest_sorted
    for i, r in enumerate(final_order, start=1):
        r["page_id"] = f"FP0002-DEMO-PG-{i:03d}"
    return final_order


def output_path(url: str) -> str:
    if url == "/":
        return "dist/index.html"
    path = url.strip("/")
    return f"dist/{path}/index.html"


def unique_display_name(r: dict, page_id_map: dict) -> str:
    name = r.get("raw_name") or "Страница"
    if name.lower() != "название":
        return name
    parent = r.get("parent_resolved")
    parent_name = page_id_map.get(parent, {}).get("raw_name") if parent else None
    slot = (r.get("url") or "").rstrip("/").split("/")[-1]
    if parent_name:
        return f"Зарезервированная страница — {parent_name} ({slot})"
    return f"Зарезервированная страница ({r.get('source_row', slot)})"


def title_for(name: str, url: str, page_id: str = "") -> str:
    if url == "/":
        return HOME_TITLE
    return f"{name} — Шпиговский Дом"


def breadcrumbs(page_id_map: dict, url: str) -> list[dict]:
    if url == "/":
        return []
    trail = []
    current = url
    seen = set()
    while current and current not in seen:
        seen.add(current)
        p = page_id_map.get(current)
        if p:
            trail.insert(0, {"name": p["raw_name"], "url": current, "page_id": p["page_id"]})
        parent = None
        if current in page_id_map:
            parent = page_id_map[current].get("parent_resolved")
        current = parent
    return trail


def add_footer_legal_pages(pages: list[dict], page_id_map: dict) -> list[dict]:
    legal = [
        ("/privacy-policy/", "Политика конфиденциальности"),
        ("/user-agreement/", "Пользовательское соглашение"),
        ("/consent-personal-data/", "Согласие на обработку персональных данных"),
        ("/cookie-files-policy/", "Политика Cookie-файлов"),
    ]
    extra = []
    for url, name in legal:
        if url in page_id_map:
            continue
        extra.append(
            {
                "source_row": "INFERRED:footer.html",
                "raw_url": url,
                "url": url,
                "raw_name": name,
                "normalized_name": name,
                "hierarchy_level": 1,
                "parent_raw": "/",
                "parent_resolved": "/",
                "source_notes": "INFERRED: footer legal links not in Excel",
                "confidence": "MEDIUM",
                "children_count": 0,
                "slug_source": "DEMO_GENERATED_SLUG",
            }
        )
    return pages + extra


def main():
    xlsx = find_excel()
    wb = load_workbook(str(xlsx), read_only=False, data_only=True, keep_links=True)
    audits = sheet_audit(wb)
    ws = wb["Структура"]
    raw_rows = parse_structure(ws)
    rows = build_hierarchy(raw_rows)
    rows = resolve_url_collisions(rows)
    rows = build_hierarchy(rows)
    rows = assign_ids(rows)
    page_id_map = {r["url"]: r for r in rows}

    # footer legal (not in excel)
    rows = assign_ids(add_footer_legal_pages(rows, page_id_map))
    page_id_map = {r["url"]: r for r in rows}

    pages = []
    for r in rows:
        url = r["url"]
        display_name = unique_display_name(r, page_id_map)
        name = display_name
        slug = slug_from_url(url)
        slug_source = r.get("slug_source") or ("EXPLICIT" if r.get("raw_url") else "DEMO_GENERATED_SLUG")
        tmpl, evidence, conf = classify_page(url, r["raw_name"], r["hierarchy_level"], r["children_count"])
        if r.get("source_notes", "").startswith("PLACEHOLDER") or "NO_URL" in (r.get("source_notes") or ""):
            tmpl = "PLACEHOLDER_PAGE"
            conf = "LOW"
        bc = breadcrumbs(page_id_map, url)
        pages.append(
            {
                **r,
                "normalized_name": display_name,
                "slug": slug,
                "slug_source": slug_source,
                "canonical_url": url,
                "output_path": output_path(url),
                "template": tmpl,
                "template_evidence": evidence,
                "template_confidence": conf,
                "title": title_for(display_name, url, r.get("page_id", "")),
                "h1": display_name,
                "breadcrumbs": bc,
                "menu_presence": url in MENU_TOP or url == "/",
                "placeholder_message": PLACEHOLDER_MSG if tmpl == "PLACEHOLDER_PAGE" else None,
            }
        )

    # validations
    titles = [p["title"] for p in pages]
    h1s = [p["h1"] for p in pages]
    urls = [p["canonical_url"] for p in pages]

  # navigation draft from runtime header/footer
    nav = []
    header_links = [
        ("desktop_header", "Лечение и профилактика", "/uslugi/"),
        ("desktop_header", "Генотипирование", "/uslugi/genotipirovanie/"),
        ("desktop_header", "Специалисты", "/specyalisty/"),
        ("desktop_header", "О центре", "/o-centre/"),
        ("desktop_header", "Отзывы", "/otzyvy/"),
        ("desktop_header", "Статьи", "/blog/"),
        ("desktop_header", "Контакты", "/kontakty/"),
        ("mobile_header", "Лечение и профилактика", "/uslugi/"),
        ("mobile_header", "Генотипирование", "/uslugi/genotipirovanie/"),
        ("mobile_header", "Специалисты", "/specyalisty/"),
        ("mobile_header", "О центре", "/o-centre/"),
        ("mobile_header", "Отзывы", "/otzyvy/"),
        ("mobile_header", "Статьи", "/blog/"),
        ("mobile_header", "Контакты", "/kontakty/"),
        ("logo", "Шпиговский дом", "/"),
        ("footer_services", "Зависимости и пристрастия", "/uslugi/zavisimosti/"),
        ("footer_services", "Психическое здоровье", "/uslugi/psihicheskoe-zdorovie/"),
        ("footer_services", "Расстройства пищевого поведения", "/uslugi/rasstroystva-pischevogo-povedeniya/"),
        ("footer_services", "Генотипирование", "/uslugi/genotipirovanie/"),
        ("footer_about", "О нас", "/o-centre/o-nas/"),
        ("footer_about", "Программа лечения", "/o-centre/programma-lecheniya/"),
        ("footer_about", "Галерея о доме", "/o-centre/galereya-o-dome/"),
        ("footer_about", "Специалистам", "/o-centre/specialistam/"),
        ("footer_about", "Родственникам", "/o-centre/rodstvennikam/"),
        ("footer_legal", "Политика конфиденциальности", "/privacy-policy/"),
        ("footer_legal", "Пользовательское соглашение", "/user-agreement/"),
        ("footer_legal", "Согласие на обработку персональных данных", "/consent-personal-data/"),
        ("footer_legal", "Политика Cookie-файлов", "/cookie-files-policy/"),
    ]
    for surface, label, target_url in header_links:
        target = page_id_map.get(target_url)
        nav.append(
            {
                "surface": surface,
                "current_label": label,
                "target_page_id": target["page_id"] if target else None,
                "target_url": target_url,
                "source_of_mapping": "EXPLICIT: runtime header/footer.html",
                "confidence": "HIGH" if target else "UNRESOLVED_LINK_TARGET",
            }
        )

    stats = {
        "excel_path": str(xlsx),
        "excel_size": xlsx.stat().st_size,
        "excel_modified": datetime.fromtimestamp(xlsx.stat().st_mtime).isoformat(),
        "sheet_audits": audits,
        "total_source_rows": ws.max_row,
        "page_rows": len(pages),
        "placeholder_slots": sum(1 for p in pages if p.get("source_notes") == "PLACEHOLDER_SLOT"),
        "duplicate_titles": len(titles) - len(set(titles)),
        "duplicate_h1": len(h1s) - len(set(h1s)),
        "duplicate_urls": len(urls) - len(set(urls)),
        "max_depth": max((p["hierarchy_level"] for p in pages), default=0),
        "template_counts": {},
    }
    for p in pages:
        stats["template_counts"][p["template"]] = stats["template_counts"].get(p["template"], 0) + 1

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "meta": stats,
        "source_normalization": [
            {k: v for k, v in p.items() if k in (
                "source_row", "raw_name", "normalized_name", "hierarchy_level",
                "parent_raw", "parent_resolved", "raw_url", "page_role",
                "menu_presence", "source_notes", "confidence", "url"
            ) or k.startswith("level")}
            for p in pages
        ],
        "pages": pages,
        "navigation": nav,
    }
    (OUT_DIR / "demo-page-registry.draft.json").write_text(
        json.dumps({"meta": stats, "pages": pages}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (OUT_DIR / "demo-navigation-registry.draft.json").write_text(
        json.dumps({"meta": {"excel": str(xlsx)}, "links": nav}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (OUT_DIR / "_parse-result.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
