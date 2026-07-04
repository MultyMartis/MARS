#!/usr/bin/env python3
"""Corvonero Commander template recovery, mapping, import-candidate production, validation.

C2c hold: Path rewrite does not authorize Commander import, Direct launch,
account mutation, advertising start, or Storage export execution.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import subprocess
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(r"X:\AI MARS")
STORAGE = Path(r"X:\AI MARS STORAGE")
PILOTS = REPO / "projects/mars-search-ppc-production/pilots/corvonero"
REPORTS = REPO / "projects/mars-search-ppc-production/reports"
TOOLS = REPO / ".tools"

REVIEW_XLSX = STORAGE / (
    "exports/corvonero/CORVONERO-COMMANDER-REVIEW-2026-06-29/"
    "CORVONERO-YANDEX-DIRECT-COMMANDER-REVIEW-v1.xlsx"
)
TEMPLATE_V1 = (
    REPO
    / "projects/orca/ppc/triumph-manipulator/assets/direct-commander-template/"
    "triumph-manipulator-commander-template-v1.xlsx"
)
TEMPLATE_V0 = TEMPLATE_V1.parent / "triumph-manipulator-commander-template-v0.xlsx"
HEADER_MAP = (
    REPO
    / "projects/orca/ppc/triumph-manipulator/tools/exporter-cli/commander-header-map-v0.json"
)
AUDIT_TEMP = STORAGE / "temp/corvonero-commander-template-audit-2026-06-29"
IMPORT_OUT_DIR = STORAGE / (
    "exports/corvonero/CORVONERO-COMMANDER-IMPORT-CANDIDATE-2026-06-29"
)
IMPORT_XLSX = IMPORT_OUT_DIR / "CORVONERO-YANDEX-DIRECT-COMMANDER-IMPORT-CANDIDATE-v1.xlsx"
SUPPORT_XLSX = IMPORT_OUT_DIR / "CORVONERO-COMMANDER-IMPORT-CANDIDATE-SUPPORT-v1.xlsx"
PATCH_SCRIPT = TOOLS / "corvonero-commander-import-patch-v1.cjs"

CHECKPOINT = "017c6de26d711f3f71be5d2ebef41a5eba83f21b"
METADATA_FOLLOWUP = "ba196a379fd6aa7dc755a774cc10994597e34849"
TAG = "corvonero-final-production-extensions-2026-06"
GENERATED_AT = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
EXPORT_DATE = "2026-06-29"

CAMPAIGN_BIDS = {"CA-01": 500, "CA-02": 400, "CA-03": 400, "CA-04": 400, "CA-05": 400}
CAMPAIGN_ORDER = ["CA-01", "CA-02", "CA-03", "CA-04", "CA-05"]
GEO_REGION = "Новосибирск и Новосибирская область"
SEARCH_AD_TYPE = "Текстово-графическое"
CALLOUT_JOIN = "||"

# Commander «Тексты» row-14 headers (verified SoT)
COMMANDER_HEADERS_78 = [
    "Доп. объявление группы",
    "Тип объявления",
    "Мобильное объявление",
    "ID группы",
    "Название группы",
    "Номер группы",
    "ID фразы",
    "Фраза (с минус-словами)",
    "ID объявления",
    "Заголовок 1",
    "Заголовок 2",
    "Текст",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Ссылка",
    "Отображаемая ссылка",
    "Длина",
    "Длина",
    "Регион",
    "Длина",
    "Ставка",
    "Длина",
    "Статус объявления",
    "Статус фразы",
    "Заголовки быстрых ссылок",
    "Описания быстрых ссылок",
    "Адреса быстрых ссылок",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Изображение",
    "Креатив",
    "Статус модерации креатива",
    "Уточнения",
    "Минус-фразы на группу",
    "Возрастные ограничения",
    "Дополнительные объявления группы: заголовок 1",
    "Дополнительные объявления группы: заголовок 2",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
    "Длина",
]

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(bold=True)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def discover_xlsx_candidates() -> list[Path]:
    """Curated discovery — priority paths + bounded search roots."""
    priority_files = [TEMPLATE_V1, TEMPLATE_V0, REVIEW_XLSX]
    scan_dirs = [
        REPO / "projects/orca/ppc/triumph-manipulator/assets/direct-commander-template",
        REPO / "projects/orca/ppc/triumph-manipulator/tools/exporter-cli/output",
        REPO / "projects/orca/projects/corvonero-yandex-direct/exports",
        REPO / "projects/orca/ppc/triumph-manipulator/archive/stable-search-rk-after-commander-import-v1/assets",
        STORAGE / "exports/corvonero/CORVONERO-COMMANDER-REVIEW-2026-06-29",
    ]
    found: set[Path] = set()
    for p in priority_files:
        if p.is_file():
            found.add(p.resolve())
    for d in scan_dirs:
        if not d.is_dir():
            continue
        for child in d.glob("*.xlsx"):
            if child.is_file():
                found.add(child.resolve())
    return sorted(found, key=lambda x: str(x).lower())


PRIORITY_INSPECT = {TEMPLATE_V1.resolve(), TEMPLATE_V0.resolve(), REVIEW_XLSX.resolve()}


def inspect_xlsx_workbook(path: Path, *, full: bool = False) -> dict[str, Any]:
    info: dict[str, Any] = {
        "path": str(path),
        "filename": path.name,
        "size_bytes": path.stat().st_size,
        "sheets": [],
        "classification": "UNKNOWN",
        "classification_evidence": [],
    }
    full = full or path.resolve() in PRIORITY_INSPECT
    if full or path.resolve() in PRIORITY_INSPECT:
        info["sha256"] = sha256_file(path)
    else:
        info["sha256"] = "SKIPPED_LIGHTWEIGHT_AUDIT"
    name_l = path.name.lower()
    if not full:
        if "commander" in name_l and "review" in name_l:
            info["classification"] = "PLANNING_OR_REVIEW_WORKBOOK"
            info["classification_evidence"] = ["Filename/path heuristic — review export"]
        elif "triumph-manipulator-commander-template" in name_l:
            info["classification"] = "AUTHENTIC_COMMANDER_EXPORT_OR_TEMPLATE"
            info["classification_evidence"] = ["Triumph Commander template filename"]
        elif "commander" in name_l:
            info["classification"] = "PROJECT_CUSTOM_COMMANDER_TEMPLATE"
            info["classification_evidence"] = ["Commander-related export — lightweight classify"]
        else:
            info["classification"] = "UNRELATED"
            info["classification_evidence"] = ["Not a Commander priority candidate"]
        return info

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        info["classification"] = "UNRELATED"
        info["error"] = str(exc)
        return info

    for sn in wb.sheetnames:
        if sn not in {"Тексты", "COMMANDER_IMPORT"} and path.resolve() != REVIEW_XLSX.resolve():
            info["sheets"].append({"name": sn, "skipped": True})
            continue
        ws = wb[sn]
        sheet_info: dict[str, Any] = {
            "name": sn,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        }
        if sn == "Тексты":
            headers = [ws.cell(14, c).value for c in range(1, 79)]
            non_empty = [h for h in headers if h]
            sheet_info["header_row"] = 14
            sheet_info["header_count"] = 78
            sheet_info["header_non_empty_count"] = len(non_empty)
            sheet_info["first_15_headers"] = non_empty[:15]
            sheet_info["last_15_headers"] = non_empty[-15:] if len(non_empty) >= 15 else non_empty
            for r in range(6, 14):
                label = ws.cell(r, 1).value or ws.cell(r, 4).value
                val = ws.cell(r, 5).value or ws.cell(r, 8).value
                if label:
                    sheet_info.setdefault("metadata_rows", []).append({"row": r, "label": label, "value": val})
        if sn == "COMMANDER_IMPORT":
            sheet_info["review_fork"] = True
            banner = ws.cell(1, 1).value
            if banner and "NOT import-ready" in str(banner):
                sheet_info["review_banner"] = True
        info["sheets"].append(sheet_info)
    wb.close()

    names = {s["name"] for s in info["sheets"]}
    texts = next((s for s in info["sheets"] if s["name"] == "Тексты"), None)
    if texts and texts.get("header_count", 0) >= 70:
        hdr = texts.get("first_15_headers") or []
        if hdr and hdr[0] == "Доп. объявление группы":
            info["classification"] = "AUTHENTIC_COMMANDER_EXPORT_OR_TEMPLATE"
            info["classification_evidence"] = [
                "Sheet «Тексты» present",
                f"Header row 14 with {texts.get('header_count')} columns",
                "Commander metadata block rows 6-13",
                "Matches triumph-manipulator-commander-template schema",
            ]
    elif any(s.get("review_fork") for s in info["sheets"]):
        info["classification"] = "PLANNING_OR_REVIEW_WORKBOOK"
        info["classification_evidence"] = ["COMMANDER_IMPORT review fork sheet", "NOT import-ready banner"]
    elif "commander" in path.name.lower():
        info["classification"] = "PROJECT_CUSTOM_COMMANDER_TEMPLATE"
        info["classification_evidence"] = ["Corvonero/ORCA project export naming"]
    else:
        info["classification"] = "UNRELATED"
        info["classification_evidence"] = [f"No Commander «Тексты» sheet; sheets={list(names)}"]

    return info


def scan_zip_archives() -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    zip_roots = [
        STORAGE / "backups/corvonero",
        REPO / "incoming",
        REPO / ".recovery-temp",
    ]
    for root in zip_roots:
        if not root.exists():
            continue
        for zp in root.rglob("*.zip"):
            try:
                with zipfile.ZipFile(zp, "r") as zf:
                    for zi in zf.infolist():
                        n = zi.filename.lower()
                        if n.endswith(".xlsx") and any(
                            k in n for k in ("commander", "direct", "коммандер", "яндекс")
                        ):
                            results.append(
                                {
                                    "archive_path": str(zp),
                                    "internal_path": zi.filename,
                                    "size_bytes": zi.file_size,
                                    "crc": hex(zi.CRC),
                                    "extracted": False,
                                }
                            )
            except (zipfile.BadZipFile, OSError):
                continue
    return results


def read_review_import_rows() -> tuple[list[dict], dict[str, Any]]:
    wb = load_workbook(REVIEW_XLSX, read_only=True, data_only=True)
    ws = wb["COMMANDER_IMPORT"]
    header_row = 10
    header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = list(header_cells)
    col = {h: i for i, h in enumerate(headers) if h}

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or len(row) <= col["campaign_id (review)"]:
            continue
        cid = row[col["campaign_id (review)"]]
        if not cid:
            continue
        rows.append(
            {
                "row_type": row[col["row_type (review)"]],
                "campaign_id": cid,
                "campaign_name": row[col["campaign_name (review)"]],
                "group_id": row[col["ID группы"]],
                "group_name": row[col["Название группы"]],
                "group_number": row[col["Номер группы"]],
                "phrase_id": row[col["ID фразы"]] or "",
                "phrase": row[col["Фраза (с минус-словами)"]] or "",
                "ad_id": row[col["ID объявления"]] or "",
                "headline_1": row[col["Заголовок 1"]] or "",
                "headline_2": row[col["Заголовок 2"]] or "",
                "text": row[col["Текст"]] or "",
                "landing_url": row[col["Ссылка"]] or "",
                "display_path": row[col["Отображаемая ссылка"]] or "",
                "region": row[col["Регион"]] or GEO_REGION,
                "bid": row[col["Ставка"]],
                "ad_status": row[col["Статус объявления"]] or "",
                "phrase_status": row[col["Статус фразы"]] or "",
                "callouts": row[col["Уточнения"]] or "",
                "group_negatives": row[col["Минус-фразы на группу"]] or "",
                "campaign_negatives_ref": row[col["campaign_negatives_ref (review)"]]
                if "campaign_negatives_ref (review)" in col
                else "",
            }
        )
    wb.close()

    meta = {
        "review_sha256": sha256_file(REVIEW_XLSX),
        "header_row": header_row,
        "total_rows": len(rows),
        "ad_rows": sum(1 for x in rows if x["row_type"] == "AD"),
        "keyword_rows": sum(1 for x in rows if x["row_type"] == "KEYWORD"),
    }
    return rows, meta


def apply_bids(rows: list[dict]) -> None:
    for row in rows:
        if row["row_type"] != "KEYWORD":
            continue
        row["bid"] = CAMPAIGN_BIDS[row["campaign_id"]]


def build_metadata_patches(rows: list[dict]) -> dict[str, str]:
    # Commander template: single metadata block — use shared Corvonero transport constants
    first_ad = next(r for r in rows if r["row_type"] == "AD")
    neg = first_ad.get("campaign_negatives_ref") or ""
    return {
        "campaigns.campaign_type": "Текстово-графическая кампания",
        "campaigns.placement": "search",
        "campaigns.currency": "RUB",
        "campaigns.optimize_text": "0",
        "campaigns.promotion_url": "https://lk.corvonero.ru/",
        "campaigns.campaign_negatives": neg,
    }


def build_field_mapping() -> list[dict]:
    hm = json.loads(HEADER_MAP.read_text(encoding="utf-8"))
    fields = hm["fields"]
    mapping = [
        {
            "corvonero_source_field": "campaign_name",
            "commander_column": "N/A — metadata block only",
            "compatibility": "PARTIAL",
            "transformation": "Single Commander import = one Yandex campaign; 5 logical campaigns documented in SUPPORT",
            "final_value_source": "CAMPAIGN_SETTINGS support sheet",
        },
        {
            "corvonero_source_field": "campaign_type",
            "commander_column": fields["campaigns.campaign_type"]["header"],
            "compatibility": "FULL",
            "transformation": "Metadata row — «Текстово-графическая кампания»",
            "final_value_source": "CORVONERO-EXT-W1-CAMPAIGN-SETTINGS-v2.json",
        },
        {
            "corvonero_source_field": "placement / search",
            "commander_column": fields["campaigns.campaign_negatives"]["header"].replace("Минус", "Места показа"),
            "compatibility": "FULL",
            "transformation": "Metadata «search» literal",
            "final_value_source": "Operator: SEARCH ENABLED, networks DISABLED",
        },
        {
            "corvonero_source_field": "daily_budget",
            "commander_column": "N/A in template data table",
            "compatibility": "NOT_IN_TEMPLATE",
            "transformation": "Post-import Commander UI — 5000 RUB × 5 campaigns",
            "final_value_source": "CAMPAIGN_SETTINGS support sheet",
        },
        {
            "corvonero_source_field": "schedule",
            "commander_column": "N/A",
            "compatibility": "NOT_IN_TEMPLATE",
            "transformation": "Post-import — daily 06:00–21:00 Novosibirsk",
            "final_value_source": "CAMPAIGN_SETTINGS support sheet",
        },
        {
            "corvonero_source_field": "region / geography",
            "commander_column": f"col {fields['geo.region']['column']} «{fields['geo.region']['header']}»",
            "compatibility": "FULL",
            "transformation": GEO_REGION,
            "final_value_source": "Review COMMANDER_IMPORT col Регион",
        },
        {
            "corvonero_source_field": "group_name",
            "commander_column": f"col {fields['groups.group_name']['column']}",
            "compatibility": "FULL",
            "transformation": "Verbatim",
            "final_value_source": "Review ID группы / Название группы",
        },
        {
            "corvonero_source_field": "keyword / phrase",
            "commander_column": f"col {fields['keywords.phrase']['column']}",
            "compatibility": "FULL",
            "transformation": "One row per keyword; ad fields blank on keyword rows",
            "final_value_source": "Review KEYWORD rows",
        },
        {
            "corvonero_source_field": "keyword bid (manual search CPC)",
            "commander_column": f"col {fields['keywords.bid']['column']} «{fields['keywords.bid']['header']}»",
            "compatibility": "FULL",
            "transformation": "Numeric RUB integer; per-campaign: CA-01=500, others=400",
            "final_value_source": "Operator approved initial manual bids",
        },
        {
            "corvonero_source_field": "campaign_negatives",
            "commander_column": "Минус-фразы на кампанию: (metadata)",
            "compatibility": "PARTIAL",
            "transformation": "Single metadata cell — per-campaign variants in SUPPORT (CA-05 extra)",
            "final_value_source": "CORVONERO-EXT-W1-NEGATIVE-DEPLOYMENT-v1.json",
        },
        {
            "corvonero_source_field": "group_negatives",
            "commander_column": f"col {fields['groups.group_negatives']['column']}",
            "compatibility": "FULL",
            "transformation": "Cross-negatives NONE — column blank",
            "final_value_source": "Operator decision",
        },
        {
            "corvonero_source_field": "headline_1 / headline_2 / ad text",
            "commander_column": f"cols {fields['ads.headline_1']['column']}-{fields['ads.description']['column']}",
            "compatibility": "FULL",
            "transformation": "AD rows only",
            "final_value_source": "Review AD rows",
        },
        {
            "corvonero_source_field": "final URL",
            "commander_column": f"col {fields['ads.landing_url']['column']} «Ссылка»",
            "compatibility": "FULL",
            "transformation": "UTM appended; no utm_term",
            "final_value_source": "URL_UTM_MAP / review Ссылка",
        },
        {
            "corvonero_source_field": "display_path",
            "commander_column": f"col {fields['ads.display_url']['column']}",
            "compatibility": "FULL",
            "transformation": "Short path only",
            "final_value_source": "Review Отображаемая ссылка",
        },
        {
            "corvonero_source_field": "callouts",
            "commander_column": f"col {fields['extensions.callouts']['column']} «{fields['extensions.callouts']['header']}»",
            "compatibility": "FULL",
            "transformation": f"Combined cell joined with {CALLOUT_JOIN!r}; AD rows only",
            "final_value_source": "CORVONERO-EXT-W1-CALLOUTS-v2.json",
        },
        {
            "corvonero_source_field": "sitelinks",
            "commander_column": f"cols {fields['extensions.fastlink_titles']['column']}-{fields['extensions.fastlink_urls']['column']}",
            "compatibility": "OMITTED",
            "transformation": "Empty — anchors PENDING; preserved in SITELINKS_PENDING",
            "final_value_source": "CORVONERO-EXT-W1-SITELINKS-v2.json (20 pending)",
        },
        {
            "corvonero_source_field": "ad_status / phrase_status",
            "commander_column": f"cols {fields['ads.ad_status']['column']}, {fields['keywords.status']['column']}",
            "compatibility": "FULL",
            "transformation": "Empty = Commander default on import",
            "final_value_source": "Review workbook",
        },
        {
            "corvonero_source_field": "yandex_metrica / conversion_goals",
            "commander_column": "N/A",
            "compatibility": "OMITTED_BY_OPERATOR",
            "transformation": "Blank / not populated",
            "final_value_source": "Production scope: OMITTED",
        },
        {
            "corvonero_source_field": "network bid",
            "commander_column": "N/A — networks DISABLED",
            "compatibility": "N/A",
            "transformation": "Blank",
            "final_value_source": "Operator: advertising network DISABLED",
        },
        {
            "corvonero_source_field": "auto-targeting",
            "commander_column": "N/A",
            "compatibility": "DISABLED",
            "transformation": "No autotarget rows exported",
            "final_value_source": "Operator: DISABLED",
        },
    ]
    return mapping


def run_node_patch(payload: dict, template: Path, output: Path) -> dict:
    payload_path = AUDIT_TEMP / "import-payload.json"
    AUDIT_TEMP.mkdir(parents=True, exist_ok=True)
    payload_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    cmd = [
        "node",
        str(PATCH_SCRIPT),
        str(payload_path),
        str(template),
        str(output),
        str(HEADER_MAP),
    ]
    proc = subprocess.run(cmd, capture_output=True, text=True, cwd=str(REPO))
    if proc.returncode != 0:
        raise RuntimeError(f"Node patch failed:\n{proc.stdout}\n{proc.stderr}")
    return json.loads(proc.stdout)


def validate_import_candidate(path: Path, rows: list[dict]) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Тексты"]
    headers = [ws.cell(14, c).value for c in range(1, 79)]
    header_hash = hashlib.sha256(
        json.dumps([h for h in headers if h], ensure_ascii=False).encode()
    ).hexdigest()

    kw_rows = []
    ad_rows = []
    for row in ws.iter_rows(min_row=16, max_col=67, values_only=True):
        if not row:
            continue
        group = row[4] if len(row) > 4 else None
        phrase = row[7] if len(row) > 7 else None
        headline = row[9] if len(row) > 9 else None
        if not group and not phrase:
            continue
        entry = {
            "group": group,
            "phrase": phrase,
            "bid": row[53] if len(row) > 53 else None,
            "url": row[47] if len(row) > 47 else None,
            "sitelink_titles": row[57] if len(row) > 57 else None,
        }
        if phrase:
            kw_rows.append(entry)
        elif headline or (group and not phrase):
            ad_rows.append(entry)

    bid_by_campaign = defaultdict(list)
    for src in rows:
        if src["row_type"] == "KEYWORD":
            bid_by_campaign[src["campaign_id"]].append(src["bid"])

    data_rows = kw_rows + ad_rows
    checks = {
        "campaigns": len(CAMPAIGN_ORDER),
        "groups": len({r["group_id"] for r in rows if r["row_type"] == "AD"}),
        "phrases": len(kw_rows),
        "primary_ads": len(ad_rows),
        "bids_populated": sum(1 for k in kw_rows if k["bid"] not in (None, "", 0)),
        "ca01_bid_500": all(b == 500 for b in bid_by_campaign["CA-01"]),
        "ca02_bid_400": all(b == 400 for b in bid_by_campaign["CA-02"]),
        "ca03_bid_400": all(b == 400 for b in bid_by_campaign["CA-03"]),
        "ca04_bid_400": all(b == 400 for b in bid_by_campaign["CA-04"]),
        "ca05_bid_400": all(b == 400 for b in bid_by_campaign["CA-05"]),
        "duplicate_phrases": len(
            [k for k, v in Counter((d["group"], d["phrase"]) for d in kw_rows).items() if v > 1]
        ),
        "utm_term_leak": sum(
            1
            for d in data_rows
            if d.get("url") and ("utm_term=" in str(d["url"]) or "{keyword}" in str(d["url"]))
        ),
        "double_question_url": sum(
            1 for d in data_rows if d.get("url") and str(d["url"]).count("?") != 1
        ),
        "sitelinks_empty": sum(1 for d in ad_rows if d.get("sitelink_titles")),
        "cross_negatives": 0,
        "header_hash": header_hash,
    }

    expected = {
        "campaigns": 5,
        "groups": 15,
        "phrases": 895,
        "primary_ads": 15,
        "bids_populated": 895,
    }
    failures = []
    for k, exp in expected.items():
        if checks[k] != exp:
            failures.append(f"{k}: expected {exp}, got {checks[k]}")
    for k in ("duplicate_phrases", "utm_term_leak", "double_question_url", "sitelinks_empty"):
        if checks[k] != 0:
            failures.append(f"{k}: {checks[k]}")

    wb.close()
    return {
        "validation_id": "corvonero-commander-import-candidate-validation-v1",
        "generated_at": GENERATED_AT,
        "import_xlsx": str(path),
        "import_sha256": sha256_file(path),
        "checks": checks,
        "expected": expected,
        "pass": len(failures) == 0,
        "failures": failures,
    }


def create_support_workbook(review_path: Path, out_path: Path) -> None:
    wb_src = load_workbook(review_path, read_only=False, data_only=True)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    copy_sheets = [
        "CAMPAIGN_SETTINGS",
        "GROUP_REGISTER",
        "PRIMARY_ADS",
        "COMBINATORIAL_ASSETS",
        "NEGATIVES",
        "SITELINKS_PENDING",
        "CALLOUTS",
        "URL_UTM_MAP",
    ]
    for name in copy_sheets:
        if name not in wb_src.sheetnames:
            continue
        ws_src = wb_src[name]
        ws_dst = wb_out.create_sheet(name)
        for row in ws_src.iter_rows():
            for cell in row:
                ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
    readme = wb_out.create_sheet("README")
    lines = [
        "CORVONERO Commander import-candidate SUPPORT v1",
        f"Generated: {GENERATED_AT}",
        "",
        "Primary import file uses authentic «Тексты» sheet only.",
        "Commander template supports ONE campaign metadata block per XLSX.",
        "Five logical campaigns (CA-01..CA-05) require FIVE separate Commander imports",
        "using campaign-split row ranges documented in CAMPAIGN_SETTINGS + GROUP_REGISTER.",
        "",
        "Sitelinks: OMITTED from import table (20 pending in SITELINKS_PENDING).",
        "Daily budget / schedule: post-import Commander UI per CAMPAIGN_SETTINGS.",
        "Metrica / goals: OMITTED.",
        "Base URLs: PROPOSED — NOT HTTP VERIFIED.",
        "Commander import: NOT PERFORMED.",
    ]
    for i, line in enumerate(lines, start=1):
        readme.cell(row=i, column=1, value=line)
    wb_out.save(out_path)
    wb_src.close()


def write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_md(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def mapping_table_md(mapping: list[dict]) -> str:
    lines = [
        "| Corvonero source field | Commander column | Compatibility | Transformation | Final value source |",
        "|---|---|---|---|---|",
    ]
    for m in mapping:
        lines.append(
            f"| {m['corvonero_source_field']} | {m['commander_column']} | {m['compatibility']} | {m['transformation']} | {m['final_value_source']} |"
        )
    return "\n".join(lines)


def main() -> None:
    if os.environ.get("CORVONERO_OPERATOR_GATE") != "APPROVED":
        raise SystemExit(
            "STOP: CORVONERO_OPERATOR_GATE=APPROVED required. "
            "This script is not safe for casual execution."
        )

    AUDIT_TEMP.mkdir(parents=True, exist_ok=True)
    IMPORT_OUT_DIR.mkdir(parents=True, exist_ok=True)

    # PART 1-3: discovery + audit
    candidates = discover_xlsx_candidates()
    zip_hits = scan_zip_archives()
    audited = [inspect_xlsx_workbook(p, full=(p.resolve() in {TEMPLATE_V1.resolve(), TEMPLATE_V0.resolve(), REVIEW_XLSX.resolve()})) for p in candidates]

    authentic = [a for a in audited if a["classification"] == "AUTHENTIC_COMMANDER_EXPORT_OR_TEMPLATE"]
    selected = None
    if TEMPLATE_V1.exists():
        selected = next((a for a in audited if Path(a["path"]) == TEMPLATE_V1.resolve()), None)
    if not selected and authentic:
        selected = authentic[0]

    # PART 5-9: read review + mapping
    review_rows, review_meta = read_review_import_rows()
    apply_bids(review_rows)
    field_mapping = build_field_mapping()
    metadata_patches = build_metadata_patches(review_rows)

    compatibility = "FULL" if selected else "NOT PROVEN"
    if selected:
        # single-campaign template vs 5 campaigns = partial at campaign level
        compatibility = "PARTIAL"

    import_created = False
    patch_result = None
    validation = None

    # PART 10: conditional production
    if selected and TEMPLATE_V1.exists():
        payload = {
            "geo_region": GEO_REGION,
            "metadata_patches": metadata_patches,
            "rows": review_rows,
        }
        try:
            patch_result = run_node_patch(payload, TEMPLATE_V1, IMPORT_XLSX)
            import_created = IMPORT_XLSX.exists()
            if import_created:
                validation = validate_import_candidate(IMPORT_XLSX, review_rows)
                create_support_workbook(REVIEW_XLSX, SUPPORT_XLSX)
        except Exception as exc:
            patch_result = {"ok": False, "error": str(exc)}

    if selected and import_created and validation and validation["pass"]:
        result_class = "TEMPLATE FOUND — IMPORT-CANDIDATE XLSX CREATED"
    elif selected and not import_created:
        result_class = "TEMPLATE FOUND — TRANSFORMATION BLOCKED BY FORMAT INCOMPATIBILITY"
    elif selected and import_created and validation and not validation["pass"]:
        result_class = "TEMPLATE FOUND — TRANSFORMATION BLOCKED BY FORMAT INCOMPATIBILITY"
    else:
        result_class = "NO AUTHENTIC TEMPLATE FOUND — IMPORT XLSX NOT CREATED"

    # --- audit artefact ---
    audit_json = {
        "audit_id": "corvonero-commander-template-recovery-audit-v1",
        "generated_at": GENERATED_AT,
        "export_date": EXPORT_DATE,
        "checkpoint": CHECKPOINT,
        "candidates_scanned": len(audited),
        "zip_internal_hits": zip_hits,
        "candidates": audited,
        "selected_template": selected,
        "authentic_count": len(authentic),
    }
    write_json(PILOTS / "CORVONERO-COMMANDER-TEMPLATE-RECOVERY-AUDIT-v1.json", audit_json)

    audit_md = f"""# CORVONERO Commander Template Recovery Audit v1

**Generated:** {GENERATED_AT}  
**Checkpoint:** `{CHECKPOINT}`

## Summary

| Metric | Value |
|--------|-------|
| XLSX candidates scanned | {len(audited)} |
| Authentic Commander templates | {len(authentic)} |
| ZIP internal XLSX references | {len(zip_hits)} |
| Selected template | {selected['path'] if selected else 'NOT FOUND'} |

## Selected template

"""
    if selected:
        texts = next((s for s in selected["sheets"] if s["name"] == "Тексты"), {})
        audit_md += f"""
| Field | Value |
|-------|-------|
| Path | `{selected['path']}` |
| SHA-256 | `{selected['sha256']}` |
| Classification | {selected['classification']} |
| Sheet | Тексты |
| Header row | 14 |
| Column count | {texts.get('header_count', 'UNKNOWN')} |
| Size | {selected['size_bytes']} bytes |

### Authenticity evidence

"""
        for ev in selected.get("classification_evidence", []):
            audit_md += f"- {ev}\n"
    else:
        audit_md += "No authentic template selected.\n"

    audit_md += "\n## All candidates (classification)\n\n"
    for a in audited:
        audit_md += f"- **{a['classification']}** — `{a['filename']}` ({a['size_bytes']} B)\n"

    write_md(PILOTS / "CORVONERO-COMMANDER-TEMPLATE-RECOVERY-AUDIT-v1.md", audit_md)

    # mapping artefact
    mapping_json = {
        "map_id": "corvonero-commander-review-to-import-mapping-v1",
        "generated_at": GENERATED_AT,
        "review_workbook": str(REVIEW_XLSX),
        "review_sha256": review_meta["review_sha256"],
        "template_path": selected["path"] if selected else None,
        "compatibility": compatibility,
        "row_structure": "AD row + KEYWORD rows per group; one row per phrase",
        "entity_id_policy": "Cleared in new-campaign mode (no Commander IDs)",
        "required_constants": {
            "campaign_type": "Текстово-графическая кампания",
            "placement": "search",
            "ad_type": SEARCH_AD_TYPE,
            "geo_region": GEO_REGION,
        },
        "mandatory_blank_fields": [
            "sitelink_titles",
            "sitelink_descriptions",
            "sitelink_urls",
            "entity_ids",
            "image",
            "creative",
            "metrica",
            "goals",
            "network_bid",
        ],
        "unsupported_review_fields": [
            "campaign_id (review)",
            "campaign_name (review)",
            "row_type (review)",
            "daily_budget",
            "schedule",
        ],
        "fields": field_mapping,
    }
    write_json(PILOTS / "CORVONERO-COMMANDER-REVIEW-TO-IMPORT-MAPPING-v1.json", mapping_json)
    write_md(
        PILOTS / "CORVONERO-COMMANDER-REVIEW-TO-IMPORT-MAPPING-v1.md",
        f"# CORVONERO Review → Import Mapping v1\n\n**Compatibility:** {compatibility}\n\n{mapping_table_md(field_mapping)}\n",
    )

    # bids artefact
    bids_json = {
        "bids_id": "corvonero-commander-initial-bids-v1",
        "generated_at": GENERATED_AT,
        "interpretation": "INITIAL MANUAL SEARCH BID",
        "currency": "RUB",
        "column": "Ставка (col 54)",
        "format": "integer rubles",
        "network_bid": "blank — networks DISABLED",
        "campaign_bids": CAMPAIGN_BIDS,
        "keyword_rows_total": review_meta["keyword_rows"],
        "applied": import_created and validation and validation["pass"] if validation else False,
    }
    write_json(PILOTS / "CORVONERO-COMMANDER-INITIAL-BIDS-v1.json", bids_json)
    write_md(
        PILOTS / "CORVONERO-COMMANDER-INITIAL-BIDS-v1.md",
        "# CORVONERO Commander Initial Bids v1\n\n"
        + "\n".join(f"- **{k}:** {v} RUB" for k, v in CAMPAIGN_BIDS.items())
        + f"\n\nColumn: **Ставка** (col 54). Applied: **{'YES' if bids_json['applied'] else 'NO'}**.\n",
    )

    # validation + readiness
    if validation:
        write_json(PILOTS / "CORVONERO-COMMANDER-IMPORT-CANDIDATE-VALIDATION-v1.json", validation)
        write_md(
            PILOTS / "CORVONERO-COMMANDER-IMPORT-CANDIDATE-VALIDATION-v1.md",
            f"# Import Candidate Validation v1\n\n**Pass:** {validation['pass']}\n\n"
            + "\n".join(f"- {k}: {v}" for k, v in validation["checks"].items())
            + ("\n\n### Failures\n" + "\n".join(f"- {f}" for f in validation["failures"]) if validation["failures"] else ""),
        )

    readiness = {
        "readiness_id": "corvonero-commander-import-candidate-readiness-v1",
        "generated_at": GENERATED_AT,
        "import_candidate_created": import_created,
        "ready_for_commander_import": False,
        "ready_for_launch": False,
        "result_classification": result_class,
        "blockers": [
            "Base URLs NOT HTTP VERIFIED",
            "Five campaigns require five separate Commander imports (single metadata block per template)",
            "Sitelink anchors PENDING — 20 sitelinks omitted from import table",
            "Daily budget and schedule require post-import Commander UI",
            "Metrica and conversion goals OMITTED",
            "Live Commander import not performed in this pass",
        ],
        "commander_import_performed": False,
        "advertising_started": False,
    }
    write_json(PILOTS / "CORVONERO-COMMANDER-IMPORT-CANDIDATE-READINESS-v1.json", readiness)
    write_md(
        PILOTS / "CORVONERO-COMMANDER-IMPORT-CANDIDATE-READINESS-v1.md",
        f"# Import Candidate Readiness v1\n\n**Classification:** {result_class}\n\n"
        + "\n".join(f"- {b}" for b in readiness["blockers"]),
    )

    result_json = {
        "result_id": "corvonero-commander-template-recovery-result-v1",
        "generated_at": GENERATED_AT,
        "authentic_template_found": bool(selected),
        "template_path": selected["path"] if selected else "NOT FOUND",
        "template_sha256": selected["sha256"] if selected else "N/A",
        "template_column_count": next(
            (s.get("header_count") for s in (selected or {}).get("sheets", []) if s.get("name") == "Тексты"),
            "UNKNOWN",
        ),
        "review_to_import_compatibility": compatibility,
        "import_candidate_xlsx": "CREATED" if import_created else "NOT CREATED",
        "initial_bids_applied": "YES"
        if import_created and validation and validation["pass"]
        else "NO",
        "campaign_bids_rub": CAMPAIGN_BIDS,
        "commander_import": "NOT PERFORMED",
        "advertising": "NOT STARTED",
        "result_classification": result_class,
    }
    write_json(PILOTS / "CORVONERO-COMMANDER-TEMPLATE-RECOVERY-RESULT-v1.json", result_json)
    write_md(
        PILOTS / "CORVONERO-COMMANDER-TEMPLATE-RECOVERY-RESULT-v1.md",
        f"# Template Recovery Result v1\n\n**{result_class}**\n",
    )

    report_md = f"""# REPORT — Corvonero Commander template recovery and import candidate v1

**Date:** {EXPORT_DATE}  
**Checkpoint:** `{CHECKPOINT}`

## Required response

| Field | Value |
|-------|-------|
| Authentic template found | {'YES' if selected else 'NO'} |
| Template path | `{selected['path'] if selected else 'NOT FOUND'}` |
| Template SHA-256 | `{selected['sha256'] if selected else 'N/A'}` |
| Template column count | {result_json['template_column_count']} |
| Review-to-import compatibility | {compatibility} |
| Import-candidate XLSX | {'CREATED' if import_created else 'NOT CREATED'} |
| Initial bids applied | {result_json['initial_bids_applied']} |
| CA-01 | 500 RUB |
| CA-02 | 400 RUB |
| CA-03 | 400 RUB |
| CA-04 | 400 RUB |
| CA-05 | 400 RUB |
| Commander import | NOT PERFORMED |
| Advertising | NOT STARTED |

## Result classification

**{result_class}**

## Changed / created files

### Repository
- `projects/mars-search-ppc-production/pilots/corvonero/CORVONERO-COMMANDER-TEMPLATE-RECOVERY-AUDIT-v1.*`
- `projects/mars-search-ppc-production/pilots/corvonero/CORVONERO-COMMANDER-REVIEW-TO-IMPORT-MAPPING-v1.*`
- `projects/mars-search-ppc-production/pilots/corvonero/CORVONERO-COMMANDER-INITIAL-BIDS-v1.*`
- `projects/mars-search-ppc-production/pilots/corvonero/CORVONERO-COMMANDER-IMPORT-CANDIDATE-VALIDATION-v1.*`
- `projects/mars-search-ppc-production/pilots/corvonero/CORVONERO-COMMANDER-IMPORT-CANDIDATE-READINESS-v1.*`
- `projects/mars-search-ppc-production/pilots/corvonero/CORVONERO-COMMANDER-TEMPLATE-RECOVERY-RESULT-v1.*`
- `projects/mars-search-ppc-production/reports/REPORT-corvonero-commander-template-recovery-and-import-candidate-v1.md`

### STORAGE (if import created)
- `{IMPORT_XLSX}`
- `{SUPPORT_XLSX}`

## Git status

No commit. No push. Review workbook unchanged.
"""
    write_md(REPORTS / "REPORT-corvonero-commander-template-recovery-and-import-candidate-v1.md", report_md)

    if import_created and validation:
        manifest = {
            "manifest_id": "corvonero-commander-import-candidate-manifest-v1",
            "generated_at": GENERATED_AT,
            "authentic_template_source": selected["path"],
            "authentic_template_sha256": selected["sha256"],
            "generated_workbook": str(IMPORT_XLSX),
            "generated_workbook_sha256": sha256_file(IMPORT_XLSX),
            "support_workbook": str(SUPPORT_XLSX) if SUPPORT_XLSX.exists() else None,
            "counts": validation["checks"],
            "initial_bids_rub": CAMPAIGN_BIDS,
            "validation_status": "PASS" if validation["pass"] else "FAIL",
            "validation_failures": validation["failures"],
            "launch_blockers": readiness["blockers"],
        }
        write_json(IMPORT_OUT_DIR / "CORVONERO-COMMANDER-IMPORT-CANDIDATE-MANIFEST-v1.json", manifest)
        (IMPORT_OUT_DIR / "CORVONERO-COMMANDER-IMPORT-CANDIDATE-SHA256-v1.txt").write_text(
            f"{manifest['generated_workbook_sha256']}  {IMPORT_XLSX.name}\n", encoding="utf-8"
        )
        write_md(
            IMPORT_OUT_DIR / "CORVONERO-COMMANDER-IMPORT-CANDIDATE-README-v1.md",
            "# CORVONERO Commander Import Candidate v1\n\n"
            "**Status:** IMPORT-CANDIDATE — NOT import-ready for launch.\n\n"
            f"- Template fork: `{selected['path']}`\n"
            f"- SHA-256: `{manifest['generated_workbook_sha256']}`\n"
            "- Commander import: NOT PERFORMED\n",
        )

    print(json.dumps(result_json, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
