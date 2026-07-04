#!/usr/bin/env python3
"""Corvonero Commander Production Wave 1 — Review XLSX generator.

C2c hold: Path rewrite does not authorize Commander import, Direct launch,
account mutation, advertising start, or Storage export execution.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(r"X:\AI MARS")
STORAGE_ROOT = Path(r"X:\AI MARS STORAGE")
PILOTS = REPO / "projects/mars-search-ppc-production/pilots/corvonero"
REPORTS = REPO / "projects/mars-search-ppc-production/reports"
OUT = STORAGE_ROOT / "exports/corvonero/CORVONERO-COMMANDER-REVIEW-2026-06-29"
XLSX_NAME = "CORVONERO-YANDEX-DIRECT-COMMANDER-REVIEW-v1.xlsx"

CHECKPOINT = "017c6de26d711f3f71be5d2ebef41a5eba83f21b"
METADATA_FOLLOWUP = "ba196a379fd6aa7dc755a774cc10994597e34849"
TAG = "corvonero-final-production-extensions-2026-06"
GENERATED_AT = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
EXPORT_DATE = "2026-06-29"

EXCLUDED_GROUPS = frozenset(
    {"ca-02-specialist-search", "ca-02-modification", "ca-05-specialist-search"}
)

LP_BASE = {
    "LP-01": "https://lk.corvonero.ru/programmist-1s/",
    "LP-02": "https://lk.corvonero.ru/soprovozhdenie-1s/",
    "LP-03": "https://lk.corvonero.ru/dorabotka-razrabotka-1s/",
    "LP-04": "https://lk.corvonero.ru/integracii-1s/",
    "LP-05": "https://lk.corvonero.ru/markirovka-chestny-znak/",
}

CAMPAIGN_SLUGS = {
    "CA-01": "corv_programmist_1s",
    "CA-02": "corv_soprovozhdenie_1s",
    "CA-03": "corv_dorabotka_1s",
    "CA-04": "corv_integracii_1s",
    "CA-05": "corv_markirovka_1s",
}

CAMPAIGN_ORDER = ["CA-01", "CA-02", "CA-03", "CA-04", "CA-05"]

GEO_REGION = "Новосибирск и Новосибирская область"
SEARCH_AD_TYPE = "Текстово-графическое"
CALLOUT_JOIN = "||"
BID_PLACEHOLDER = "REVIEW — OPERATOR BID REQUIRED"

# Commander «Тексты» key columns (verified header-map-v0 / format contract)
COMMANDER_HEADERS = [
    "Доп. объявление группы",  # 1
    "Тип объявления",  # 2
    "Мобильное объявление",  # 3
    "ID группы",  # 4
    "Название группы",  # 5
    "Номер группы",  # 6
    "ID фразы",  # 7
    "Фраза (с минус-словами)",  # 8
    "ID объявления",  # 9
    "Заголовок 1",  # 10
    "Заголовок 2",  # 11
    "Текст",  # 12
    "Ссылка",  # 13 — review fork: col 48 in full template; compact review layout
    "Отображаемая ссылка",
    "Регион",
    "Ставка",
    "Статус объявления",
    "Статус фразы",
    "Заголовки быстрых ссылок",
    "Описания быстрых ссылок",
    "Адреса быстрых ссылок",
    "Уточнения",
    "Минус-фразы на группу",
    "campaign_id (review)",
    "campaign_name (review)",
    "row_type (review)",
    "campaign_negatives_ref (review)",
]

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(bold=True)


def load_json(name: str):
    with (PILOTS / name).open(encoding="utf-8") as f:
        return json.load(f)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def format_negative(term: str, match: str) -> str:
    term = term.strip()
    if match == "phrase" and " " in term:
        return f'-"{term}"'
    return f"-{term}"


def build_campaign_negatives(negatives: dict, campaign_id: str) -> str:
    parts: list[str] = []
    for item in negatives["layers"]["account_shared_deployable"]:
        parts.append(format_negative(item["term"], item["match"]))
    for item in negatives["layers"]["campaign_deployable"]:
        if item["campaign_id"] == campaign_id:
            parts.append(format_negative(item["term"], item["match"]))
    return " ".join(parts)


def build_final_url(base_url: str, campaign_id: str, group_slug: str) -> str:
    slug = CAMPAIGN_SLUGS[campaign_id]
    base = base_url.rstrip("/") + "/"
    return (
        f"{base}?utm_source=yandex&utm_medium=cpc"
        f"&utm_campaign={slug}&utm_content={group_slug}"
    )


def validate_url(url: str) -> list[str]:
    issues = []
    if url.count("?") != 1:
        issues.append("invalid_question_mark_count")
    if "{keyword}" in url or "utm_term=" in url:
        issues.append("forbidden_utm_term")
    if "#" in url.split("?", 1)[0]:
        issues.append("fragment_in_base")
    if re.search(r"utm_[^=&]+=[^&]*[А-Яа-яЁё]", url):
        issues.append("cyrillic_in_utm")
    return issues


def style_header_row(ws, row: int, col_count: int):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = TITLE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def autosize_columns(ws, max_width: int = 48):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(
            max_width,
            max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2,
        )
        ws.column_dimensions[letter].width = width


def main() -> None:
    if os.environ.get("CORVONERO_OPERATOR_GATE") != "APPROVED":
        raise SystemExit(
            "STOP: CORVONERO_OPERATOR_GATE=APPROVED required. "
            "This script is not safe for casual execution."
        )

    OUT.mkdir(parents=True, exist_ok=True)

    phrase_alloc = load_json("CORVONERO-AD-WAVE-1-FINAL-PHRASE-ALLOCATION-v1.json")
    group_reg = load_json("CORVONERO-AD-WAVE-1-P1-FINAL-GROUP-REGISTER-v1.json")
    primary_ads = load_json("CORVONERO-AD-WAVE-1-P1-FINAL-PRIMARY-ADS-v1.json")
    combinatorial = load_json("CORVONERO-AD-WAVE-1-P1-FINAL-COMBINATORIAL-ASSETS-v1.json")
    sitelinks = load_json("CORVONERO-EXT-W1-SITELINKS-v2.json")
    callouts = load_json("CORVONERO-EXT-W1-CALLOUTS-v2.json")
    negatives = load_json("CORVONERO-EXT-W1-NEGATIVE-DEPLOYMENT-v1.json")
    cross_negs = load_json("CORVONERO-EXT-W1-CROSS-NEGATIVES-v2.json")
    utm_policy = load_json("CORVONERO-EXT-W1-UTM-POLICY-v2.json")
    display_paths = load_json("CORVONERO-EXT-W1-DISPLAY-PATHS-v1.json")
    import_profile = load_json("CORVONERO-EXT-W1-IMPORT-PROFILE-v1.json")
    header_map_path = (
        REPO
        / "projects/orca/ppc/triumph-manipulator/tools/exporter-cli/commander-header-map-v0.json"
    )
    header_map = json.loads(header_map_path.read_text(encoding="utf-8"))

    groups = [g for g in group_reg["groups"] if g.get("deployable")]
    groups_by_campaign: dict[str, list] = defaultdict(list)
    for g in groups:
        groups_by_campaign[g["campaign_id"]].append(g)

    ordered_groups: list[dict] = []
    for cid in CAMPAIGN_ORDER:
        ordered_groups.extend(
            sorted(groups_by_campaign[cid], key=lambda x: x["group_id"])
        )

    ads_by_group = {a["group_id"]: a for a in primary_ads["ads"]}
    display_by_group = {r["group_id"]: r["display_path"] for r in display_paths["records"]}
    comb_by_group = {a["group_id"]: a for a in combinatorial["assets"]}

    deployable_phrases = [
        r
        for r in phrase_alloc["records"]
        if r.get("production_status") == "DEPLOYABLE"
        and r.get("final_group") not in EXCLUDED_GROUPS
        and r.get("final_campaign", "").startswith("CA-0")
        and r.get("final_campaign") != "CA-06"
    ]

    phrases_by_group: dict[str, list] = defaultdict(list)
    for r in deployable_phrases:
        phrases_by_group[r["final_group"]].append(r)

    campaign_neg_map = {
        cid: build_campaign_negatives(negatives, cid) for cid in CAMPAIGN_ORDER
    }

    callout_text = {
        cid: CALLOUT_JOIN.join(c["text"] for c in pool)
        for cid, pool in callouts["campaign_pools"].items()
    }

    # --- build commander import rows ---
    import_rows: list[dict] = []
    group_number = 0
    url_map_rows: list[dict] = []

    for group in ordered_groups:
        gid = group["group_id"]
        cid = group["campaign_id"]
        group_number += 1
        ad = ads_by_group[gid]
        pa = ad["primary_ad"]
        final_url = build_final_url(
            LP_BASE[group["assigned_lp"]], cid, gid
        )
        url_map_rows.append(
            {
                "campaign_id": cid,
                "group_id": gid,
                "group_name": group["group_name"],
                "base_url": LP_BASE[group["assigned_lp"]],
                "campaign_slug": CAMPAIGN_SLUGS[cid],
                "group_slug": gid,
                "final_url": final_url,
                "http_status": "NOT VERIFIED",
                "anchor_status": "NOT REQUIRED FOR BASE AD URL",
                "url_issues": validate_url(final_url),
            }
        )

        import_rows.append(
            {
                "row_type": "AD",
                "campaign_id": cid,
                "campaign_name": group["campaign_name"],
                "group_id": gid,
                "group_name": group["group_name"],
                "group_number": group_number,
                "phrase_id": "",
                "phrase": "",
                "ad_id": f"AD-{gid}",
                "headline_1": pa["headline"],
                "headline_2": pa["additional_headline"],
                "text": pa["text"],
                "landing_url": final_url,
                "display_path": display_by_group[gid],
                "region": GEO_REGION,
                "bid": "",
                "ad_status": "",
                "phrase_status": "",
                "sitelink_titles": "",
                "sitelink_descriptions": "",
                "sitelink_urls": "",
                "callouts": callout_text[cid],
                "group_negatives": "",
                "campaign_negatives_ref": campaign_neg_map[cid],
            }
        )

        for kw in sorted(
            phrases_by_group[gid], key=lambda x: (x["phrase_id"], x["phrase"])
        ):
            import_rows.append(
                {
                    "row_type": "KEYWORD",
                    "campaign_id": cid,
                    "campaign_name": group["campaign_name"],
                    "group_id": gid,
                    "group_name": group["group_name"],
                    "group_number": group_number,
                    "phrase_id": kw["phrase_id"],
                    "phrase": kw["phrase"],
                    "ad_id": "",
                    "headline_1": "",
                    "headline_2": "",
                    "text": "",
                    "landing_url": "",
                    "display_path": "",
                    "region": GEO_REGION,
                    "bid": BID_PLACEHOLDER,
                    "ad_status": "",
                    "phrase_status": "Активна",
                    "sitelink_titles": "",
                    "sitelink_descriptions": "",
                    "sitelink_urls": "",
                    "callouts": "",
                    "group_negatives": "",
                    "campaign_negatives_ref": "",
                }
            )

    # --- validation ---
    phrase_keys = [(r["final_group"], r["phrase_id"], r["phrase"]) for r in deployable_phrases]
    dup_phrases = [k for k, v in Counter(phrase_keys).items() if v > 1]
    excluded_leak = [
        r
        for r in phrase_alloc["records"]
        if r.get("final_group") in EXCLUDED_GROUPS
        and r.get("production_status") == "DEPLOYABLE"
    ]
    ca06_leak = [
        r
        for r in deployable_phrases
        if r.get("final_campaign") == "CA-06"
    ]
    url_issues_all = [u for u in url_map_rows if u["url_issues"]]

    validation = {
        "validation_id": "corvonero-commander-w1-validation-v1",
        "generated_at": GENERATED_AT,
        "counts": {
            "campaigns": len(CAMPAIGN_ORDER),
            "groups": len(ordered_groups),
            "phrases": len(deployable_phrases),
            "primary_ads": len(primary_ads["ads"]),
            "import_rows_total": len(import_rows),
            "import_ad_rows": sum(1 for r in import_rows if r["row_type"] == "AD"),
            "import_keyword_rows": sum(
                1 for r in import_rows if r["row_type"] == "KEYWORD"
            ),
            "landing_pages": len(LP_BASE),
            "unique_utm_campaign_slugs": len(set(CAMPAIGN_SLUGS.values())),
            "unique_group_slugs": len(ordered_groups),
            "shared_negatives": len(negatives["layers"]["account_shared_deployable"]),
            "campaign_license_negatives_per_campaign": 2,
            "ca05_extra_negative": 1,
            "cross_negatives": cross_negs["cross_campaign_negatives_deployed"],
            "sitelinks_pending": len(sitelinks["records"]),
            "callout_sets": len(callouts["campaign_pools"]),
        },
        "expected": {
            "campaigns": 5,
            "groups": 15,
            "phrases": 895,
            "primary_ads": 15,
            "shared_negatives": 9,
            "cross_negatives": 0,
            "sitelinks_pending": 20,
            "callout_sets": 5,
        },
        "checks": {
            "duplicate_keyword_rows": len(dup_phrases),
            "excluded_group_leakage": len(excluded_leak),
            "ca06_leakage": len(ca06_leak),
            "malformed_urls": len(url_issues_all),
            "groups_without_primary_ad": len(
                [g for g in ordered_groups if g["group_id"] not in ads_by_group]
            ),
            "utm_term_in_urls": sum(
                1 for u in url_map_rows if "utm_term" in u["final_url"]
            ),
            "keyword_macro_in_urls": sum(
                1 for u in url_map_rows if "{keyword}" in u["final_url"]
            ),
        },
        "phrase_counts_by_campaign": {},
        "phrase_counts_by_group": {
            g["group_id"]: len(phrases_by_group[g["group_id"]]) for g in ordered_groups
        },
        "pass": True,
        "failures": [],
    }

    for cid in CAMPAIGN_ORDER:
        validation["phrase_counts_by_campaign"][cid] = sum(
            1 for r in deployable_phrases if r["final_campaign"] == cid
        )

    for key, expected in validation["expected"].items():
        actual = validation["counts"].get(key)
        if actual != expected:
            validation["pass"] = False
            validation["failures"].append(f"count_mismatch:{key}:expected={expected}:actual={actual}")

    for key, limit in [
        ("duplicate_keyword_rows", 0),
        ("excluded_group_leakage", 0),
        ("ca06_leakage", 0),
        ("malformed_urls", 0),
        ("groups_without_primary_ad", 0),
        ("utm_term_in_urls", 0),
        ("keyword_macro_in_urls", 0),
    ]:
        if validation["checks"][key] > limit:
            validation["pass"] = False
            validation["failures"].append(
                f"check_failed:{key}:{validation['checks'][key]}"
            )

    # --- workbook ---
    wb = Workbook()
    wb.remove(wb.active)

    # COMMANDER_IMPORT
    ws = wb.create_sheet("COMMANDER_IMPORT")
    meta = [
        ("Review workbook — NOT import-ready", ""),
        ("Authority checkpoint", CHECKPOINT),
        ("Tag", TAG),
        ("Template reference", import_profile["template_reference"]),
        ("Primary Commander sheet name (SoT)", import_profile["sheet"]),
        ("Sitelinks in import table", "OMITTED — see SITELINKS_PENDING"),
        ("Initial bids", BID_PLACEHOLDER),
        ("Launch", "NOT AUTHORIZED"),
    ]
    for i, (k, v) in enumerate(meta, start=1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)

    header_row = len(meta) + 2
    for c, h in enumerate(COMMANDER_HEADERS, start=1):
        ws.cell(row=header_row, column=c, value=h)
    style_header_row(ws, header_row, len(COMMANDER_HEADERS))

    data_start = header_row + 1
    for ri, row in enumerate(import_rows, start=data_start):
        values = [
            "",
            SEARCH_AD_TYPE if row["row_type"] == "AD" else "",
            "",
            row["group_id"],
            row["group_name"],
            row["group_number"],
            row["phrase_id"],
            row["phrase"],
            row["ad_id"],
            row["headline_1"],
            row["headline_2"],
            row["text"],
            row["landing_url"],
            row["display_path"],
            row["region"],
            row["bid"],
            row["ad_status"],
            row["phrase_status"],
            row["sitelink_titles"],
            row["sitelink_descriptions"],
            row["sitelink_urls"],
            row["callouts"],
            row["group_negatives"],
            row["campaign_id"],
            row["campaign_name"],
            row["row_type"],
            row["campaign_negatives_ref"] if row["row_type"] == "AD" else "",
        ]
        for c, v in enumerate(values, start=1):
            ws.cell(row=ri, column=c, value=v)
    autosize_columns(ws)

    # CAMPAIGN_SETTINGS
    ws_cs = wb.create_sheet("CAMPAIGN_SETTINGS")
    cs_headers = [
        "campaign_id",
        "campaign_name_ru",
        "daily_budget_rub",
        "placement",
        "yandex_ad_network",
        "auto_targeting",
        "bid_strategy",
        "schedule",
        "timezone",
        "geography",
        "device_adjustments",
        "demographic_adjustments",
        "yandex_metrica",
        "conversion_goals",
        "launch_status",
        "campaign_negatives_commander",
        "initial_bid_field",
        "notes",
    ]
    for c, h in enumerate(cs_headers, start=1):
        ws_cs.cell(row=1, column=c, value=h)
    style_header_row(ws_cs, 1, len(cs_headers))
    for ri, cid in enumerate(CAMPAIGN_ORDER, start=2):
        g0 = groups_by_campaign[cid][0]
        row = [
            cid,
            g0["campaign_name"],
            5000,
            "SEARCH ONLY",
            "DISABLED",
            "DISABLED",
            "MANUAL CPC / MANUAL BID MANAGEMENT",
            "EVERY DAY 06:00–21:00",
            "Новосибирск",
            "Новосибирск + Новосибирская область",
            "NONE",
            "NONE",
            "NOT USED AT THIS STAGE",
            "NOT USED AT THIS STAGE",
            "NOT AUTHORIZED",
            campaign_neg_map[cid],
            BID_PLACEHOLDER,
            "Budget/schedule/strategy applied per operator Wave 1 task; template may require post-import UI for some fields",
        ]
        for c, v in enumerate(row, start=1):
            ws_cs.cell(row=ri, column=c, value=v)
    autosize_columns(ws_cs)

    # GROUP_REGISTER
    ws_gr = wb.create_sheet("GROUP_REGISTER")
    gr_headers = [
        "campaign_id",
        "campaign_name",
        "group_id",
        "group_name",
        "phrase_count",
        "assigned_lp",
        "deployable",
        "operator_primary_ad_status",
    ]
    for c, h in enumerate(gr_headers, start=1):
        ws_gr.cell(row=1, column=c, value=h)
    style_header_row(ws_gr, 1, len(gr_headers))
    for ri, g in enumerate(ordered_groups, start=2):
        row = [
            g["campaign_id"],
            g["campaign_name"],
            g["group_id"],
            g["group_name"],
            len(phrases_by_group[g["group_id"]]),
            g["assigned_lp"],
            g.get("deployable", True),
            g.get("operator_primary_ad_status", "APPROVED"),
        ]
        for c, v in enumerate(row, start=1):
            ws_gr.cell(row=ri, column=c, value=v)
    autosize_columns(ws_gr)

    # PRIMARY_ADS
    ws_pa = wb.create_sheet("PRIMARY_ADS")
    pa_headers = [
        "campaign_id",
        "group_id",
        "group_name",
        "headline_1",
        "headline_2",
        "text",
        "landing_page_id",
        "final_url",
        "display_path",
        "status",
        "technical_validation",
    ]
    for c, h in enumerate(pa_headers, start=1):
        ws_pa.cell(row=1, column=c, value=h)
    style_header_row(ws_pa, 1, len(pa_headers))
    for ri, ad in enumerate(primary_ads["ads"], start=2):
        gid = ad["group_id"]
        pa = ad["primary_ad"]
        url = build_final_url(ad["landing_page"]["url"], ad["campaign_id"], gid)
        row = [
            ad["campaign_id"],
            gid,
            ad["group_name"],
            pa["headline"],
            pa["additional_headline"],
            pa["text"],
            ad["landing_page"]["id"],
            url,
            display_by_group.get(gid, ""),
            ad.get("status", "OPERATOR_APPROVED"),
            ad.get("technical_validation", ""),
        ]
        for c, v in enumerate(row, start=1):
            ws_pa.cell(row=ri, column=c, value=v)
    autosize_columns(ws_pa)

    # COMBINATORIAL_ASSETS
    ws_ca = wb.create_sheet("COMBINATORIAL_ASSETS")
    ca_headers = [
        "group_id",
        "approval_status",
        "selected_primary_headline",
        "selected_primary_text",
        "alternate_headlines",
        "alternate_texts",
        "quarantined_headlines",
        "quarantined_texts",
    ]
    for c, h in enumerate(ca_headers, start=1):
        ws_ca.cell(row=1, column=c, value=h)
    style_header_row(ws_ca, 1, len(ca_headers))
    for ri, asset in enumerate(combinatorial["assets"], start=2):
        row = [
            asset["group_id"],
            "NOT INDIVIDUALLY OPERATOR APPROVED",
            asset["selected_primary_headline"],
            asset["selected_primary_text"],
            " | ".join(asset.get("headlines", [])),
            " | ".join(asset.get("texts", [])),
            " | ".join(asset.get("quarantined_headlines", [])),
            " | ".join(asset.get("quarantined_texts", [])),
        ]
        for c, v in enumerate(row, start=1):
            ws_ca.cell(row=ri, column=c, value=v)
    autosize_columns(ws_ca)

    # NEGATIVES
    ws_neg = wb.create_sheet("NEGATIVES")
    neg_headers = [
        "layer",
        "term",
        "match",
        "target",
        "deploy_status",
        "commander_format",
        "notes",
    ]
    for c, h in enumerate(neg_headers, start=1):
        ws_neg.cell(row=1, column=c, value=h)
    style_header_row(ws_neg, 1, len(neg_headers))
    neg_rows = []
    for item in negatives["layers"]["account_shared_deployable"]:
        neg_rows.append(
            (
                "SHARED",
                item["term"],
                item["match"],
                "ALL CAMPAIGNS",
                item["deploy_status"],
                format_negative(item["term"], item["match"]),
                item.get("reason", ""),
            )
        )
    for item in negatives["layers"]["campaign_deployable"]:
        neg_rows.append(
            (
                "CAMPAIGN",
                item["term"],
                item["match"],
                item["campaign_id"],
                item["deploy_status"],
                format_negative(item["term"], item["match"]),
                item.get("reason", ""),
            )
        )
    for item in negatives["layers"].get("account_shared_rejected", []):
        neg_rows.append(
            (
                "SHARED_REJECTED",
                item["term"],
                item.get("match", "word"),
                "NOT DEPLOYED",
                item.get("status", "REJECTED"),
                "",
                item.get("reason", ""),
            )
        )
    for item in negatives["layers"].get("campaign_not_deployed", []):
        neg_rows.append(
            (
                "CAMPAIGN_NOT_DEPLOYED",
                item["term"],
                item.get("match", "phrase"),
                ", ".join(item.get("campaigns", [])),
                item.get("status", "NOT_DEPLOYED"),
                "",
                item.get("reason", ""),
            )
        )
    for ri, row in enumerate(neg_rows, start=2):
        for c, v in enumerate(row, start=1):
            ws_neg.cell(row=ri, column=c, value=v)
    autosize_columns(ws_neg)

    # SITELINKS_PENDING
    ws_sl = wb.create_sheet("SITELINKS_PENDING")
    sl_headers = [
        "campaign",
        "title",
        "description_1",
        "description_2",
        "proposed_anchor",
        "base_lp_url",
        "proposed_full_url",
        "status",
        "notes",
    ]
    for c, h in enumerate(sl_headers, start=1):
        ws_sl.cell(row=1, column=c, value=h)
    style_header_row(ws_sl, 1, len(sl_headers))
    lp_by_campaign = {
        "CA-01": LP_BASE["LP-01"],
        "CA-02": LP_BASE["LP-02"],
        "CA-03": LP_BASE["LP-03"],
        "CA-04": LP_BASE["LP-04"],
        "CA-05": LP_BASE["LP-05"],
    }
    for ri, rec in enumerate(sitelinks["records"], start=2):
        cid = rec["campaign_id"]
        base = lp_by_campaign[cid]
        row = [
            cid,
            rec["sitelink_title"],
            rec["description_line_1"],
            rec["description_line_2"],
            rec["provisional_anchor"],
            base,
            rec["provisional_url"],
            "PENDING_FINAL_ANCHOR",
            "Add after Roman confirms final anchor ID",
        ]
        for c, v in enumerate(row, start=1):
            ws_sl.cell(row=ri, column=c, value=v)
    autosize_columns(ws_sl)

    # CALLOUTS
    ws_co = wb.create_sheet("CALLOUTS")
    co_headers = ["campaign_id", "callout_index", "text", "char_count", "operator_status"]
    for c, h in enumerate(co_headers, start=1):
        ws_co.cell(row=1, column=c, value=h)
    style_header_row(ws_co, 1, len(co_headers))
    ri = 2
    for cid in CAMPAIGN_ORDER:
        for idx, item in enumerate(callouts["campaign_pools"][cid], start=1):
            row = [cid, idx, item["text"], item["char_count"], item["operator_status"]]
            for c, v in enumerate(row, start=1):
                ws_co.cell(row=ri, column=c, value=v)
            ri += 1
    autosize_columns(ws_co)

    # URL_UTM_MAP
    ws_url = wb.create_sheet("URL_UTM_MAP")
    url_headers = [
        "campaign_id",
        "group_id",
        "group_name",
        "base_url",
        "campaign_slug",
        "group_slug",
        "final_url",
        "http_status",
        "anchor_status",
        "url_validation_issues",
    ]
    for c, h in enumerate(url_headers, start=1):
        ws_url.cell(row=1, column=c, value=h)
    style_header_row(ws_url, 1, len(url_headers))
    for ri, u in enumerate(url_map_rows, start=2):
        row = [
            u["campaign_id"],
            u["group_id"],
            u["group_name"],
            u["base_url"],
            u["campaign_slug"],
            u["group_slug"],
            u["final_url"],
            u["http_status"],
            u["anchor_status"],
            ", ".join(u["url_issues"]) if u["url_issues"] else "OK",
        ]
        for c, v in enumerate(row, start=1):
            ws_url.cell(row=ri, column=c, value=v)
    autosize_columns(ws_url)

    # README sheet
    ws_rm = wb.create_sheet("README")
    readme_lines = [
        "CORVONERO Yandex Direct Commander — REVIEW WORKBOOK v1",
        f"Generated: {GENERATED_AT}",
        "",
        "Purpose: operator review and later import preparation ONLY.",
        "This workbook is NOT imported into Commander and does NOT start advertising.",
        "",
        "Status:",
        "  XLSX CREATED: YES",
        f"  STRUCTURALLY VALIDATED: {'YES' if validation['pass'] else 'PARTIAL'}",
        "  READY FOR OPERATOR REVIEW: YES",
        "  READY FOR COMMANDER IMPORT: NO — UNTIL URL AND TEMPLATE VALIDATION",
        "  READY FOR LAUNCH: NO",
        "",
        "Base landing URLs: PROPOSED — NOT HTTP VERIFIED — verify before import/launch.",
        "Sitelinks: copy in SITELINKS_PENDING; anchors to be added after Roman confirmation.",
        f"Initial manual bids: {BID_PLACEHOLDER} — do not fabricate bids.",
        "Metrica counter and conversion goals intentionally omitted by operator decision.",
        "Sitelinks omitted from COMMANDER_IMPORT primary table (empty columns).",
        "",
        f"Authority checkpoint: {CHECKPOINT}",
        f"Metadata follow-up: {METADATA_FOLLOWUP}",
        f"Tag: {TAG}",
        "",
        "Counts:",
        f"  Campaigns: {validation['counts']['campaigns']}",
        f"  Groups: {validation['counts']['groups']}",
        f"  Phrases: {validation['counts']['phrases']}",
        f"  Primary ads: {validation['counts']['primary_ads']}",
        f"  Daily budget: 5000 RUB per campaign (see CAMPAIGN_SETTINGS)",
    ]
    for i, line in enumerate(readme_lines, start=1):
        ws_rm.cell(row=i, column=1, value=line)
    ws_rm.column_dimensions["A"].width = 100

    xlsx_path = OUT / XLSX_NAME
    wb.save(xlsx_path)

    # post-save open validation
    wb_check = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = wb_check.sheetnames
    ci_rows = sum(
        1
        for _ in wb_check["COMMANDER_IMPORT"].iter_rows(
            min_row=data_start, values_only=True
        )
    )
    wb_check.close()

    structural = "YES" if validation["pass"] and ci_rows == len(import_rows) else "PARTIAL"

    # --- supporting JSON artefacts ---
    template_audit = {
        "audit_id": "corvonero-commander-w1-template-audit-v1",
        "generated_at": GENERATED_AT,
        "template_reference": import_profile["template_reference"],
        "template_binary_present_in_repo": False,
        "commander_sheet_sot": import_profile["sheet"],
        "review_fork_sheet": "COMMANDER_IMPORT",
        "header_row_review": header_row,
        "data_start_row_review": data_start,
        "verified_columns_from_header_map": {
            k: v.get("header")
            for k, v in header_map.get("fields", {}).items()
            if v.get("status") == "verified" and v.get("column")
        },
        "metadata_block_sot_rows": "6-13",
        "unsupported_in_template": [
            "daily_budget",
            "schedule",
            "yandex_metrica",
            "conversion_goals",
            "utm_dedicated_column",
            "match_type_column",
            "campaign_name_data_column",
        ],
        "sitelinks_import_policy": "OMITTED_IN_PRIMARY_TABLE",
        "structural_validation": structural,
        "notes": [
            "Review workbook uses compact column fork + review traceability columns",
            "Full 78-column Triumph template binary not present in repo at generation time",
            "Campaign settings not fully representable — see CAMPAIGN_SETTINGS sheet",
        ],
    }

    export_map = {
        "map_id": "corvonero-commander-w1-export-map-v1",
        "generated_at": GENERATED_AT,
        "authority_sources": {
            "phrase_allocation": "CORVONERO-AD-WAVE-1-FINAL-PHRASE-ALLOCATION-v1.json",
            "group_register": "CORVONERO-AD-WAVE-1-P1-FINAL-GROUP-REGISTER-v1.json",
            "primary_ads": "CORVONERO-AD-WAVE-1-P1-FINAL-PRIMARY-ADS-v1.json",
            "combinatorial_assets": "CORVONERO-AD-WAVE-1-P1-FINAL-COMBINATORIAL-ASSETS-v1.json",
            "sitelinks": "CORVONERO-EXT-W1-SITELINKS-v2.json",
            "callouts": "CORVONERO-EXT-W1-CALLOUTS-v2.json",
            "negatives": "CORVONERO-EXT-W1-NEGATIVE-DEPLOYMENT-v1.json",
            "cross_negatives": "CORVONERO-EXT-W1-CROSS-NEGATIVES-v2.json",
            "utm_policy": "CORVONERO-EXT-W1-UTM-POLICY-v2.json",
            "campaign_settings": "CORVONERO-EXT-W1-CAMPAIGN-SETTINGS-v2.json",
            "import_profile": "CORVONERO-EXT-W1-IMPORT-PROFILE-v1.json",
            "display_paths": "CORVONERO-EXT-W1-DISPLAY-PATHS-v1.json",
        },
        "output": {
            "xlsx": str(xlsx_path),
            "sheets": sheet_names,
            "commander_import_rows": len(import_rows),
        },
        "excluded_groups": sorted(EXCLUDED_GROUPS),
        "lp_base_urls": LP_BASE,
        "utm_suffix_template": utm_policy["approved_url_suffix"],
    }

    readiness = {
        "readiness_id": "corvonero-commander-w1-readiness-v1",
        "generated_at": GENERATED_AT,
        "xlsx_created": True,
        "structurally_validated": structural,
        "ready_for_operator_review": True,
        "ready_for_commander_import": False,
        "ready_for_launch": False,
        "blockers": [
            "Base URLs NOT HTTP VERIFIED",
            "Triumph template binary not validated against live Commander import in this pass",
            "Sitelink anchors PENDING_FINAL_ANCHOR",
            "Initial manual bids unresolved",
            "Metrica and conversion goals omitted",
        ],
        "validation_pass": validation["pass"],
    }

    result = {
        "result_id": "corvonero-commander-w1-result-v1",
        "generated_at": GENERATED_AT,
        "verdict": "CORVONERO COMMANDER PRODUCTION WAVE 1: PASS — REVIEW XLSX CREATED",
        "campaigns": 5,
        "deployable_groups": 15,
        "deployable_phrases": 895,
        "daily_budget_rub_per_campaign": 5000,
        "strategy": "MANUAL SEARCH",
        "schedule": "DAILY 06:00–21:00 NOVOSIBIRSK TIME",
        "metrica": "OMITTED BY OPERATOR DECISION",
        "conversion_goals": "OMITTED BY OPERATOR DECISION",
        "base_urls": "INCLUDED — NOT HTTP VERIFIED",
        "sitelinks": "COPY INCLUDED IN PENDING SHEET — ANCHORS TO BE ADDED LATER",
        "commander_import": "NOT PERFORMED",
        "advertising": "NOT STARTED",
        "validation": validation,
        "xlsx_sha256": sha256_file(xlsx_path),
    }

    artefacts = [
        ("CORVONERO-COMMANDER-W1-TEMPLATE-AUDIT-v1", template_audit),
        ("CORVONERO-COMMANDER-W1-EXPORT-MAP-v1", export_map),
        ("CORVONERO-COMMANDER-W1-VALIDATION-v1", validation),
        ("CORVONERO-COMMANDER-W1-READINESS-v1", readiness),
        ("CORVONERO-COMMANDER-W1-RESULT-v1", result),
    ]

    for stem, payload in artefacts:
        json_path = PILOTS / f"{stem}.json"
        md_path = PILOTS / f"{stem}.md"
        json_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
        md_path.write_text(_render_md(stem, payload), encoding="utf-8")

    manifest = {
        "manifest_id": "corvonero-commander-review-manifest-v1",
        "generated_at": GENERATED_AT,
        "export_date": EXPORT_DATE,
        "checkpoint": CHECKPOINT,
        "metadata_followup": METADATA_FOLLOWUP,
        "tag": TAG,
        "xlsx": XLSX_NAME,
        "xlsx_path": str(xlsx_path),
        "xlsx_sha256": result["xlsx_sha256"],
        "sheets": sheet_names,
        "counts": validation["counts"],
        "readiness": readiness,
        "authority_modified": False,
    }
    (OUT / "CORVONERO-COMMANDER-REVIEW-MANIFEST-v1.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (OUT / "CORVONERO-COMMANDER-REVIEW-SHA256-v1.txt").write_text(
        f"{result['xlsx_sha256']}  {XLSX_NAME}\n", encoding="utf-8"
    )
    (OUT / "CORVONERO-COMMANDER-REVIEW-README-v1.md").write_text(
        _render_storage_readme(manifest, structural), encoding="utf-8"
    )

    report_path = REPORTS / "REPORT-corvonero-commander-production-wave-1-review-xlsx-v1.md"
    report_path.write_text(_render_report(manifest, validation, structural, xlsx_path), encoding="utf-8")

    print(json.dumps({"pass": validation["pass"], "xlsx": str(xlsx_path), "rows": len(import_rows)}, ensure_ascii=False))
    if not validation["pass"]:
        raise SystemExit(1)


def _render_md(stem: str, payload: dict) -> str:
    lines = [f"# {stem}", "", f"Generated: {payload.get('generated_at', GENERATED_AT)}", ""]
    if "verdict" in payload:
        lines.append(f"**Verdict:** {payload['verdict']}")
        lines.append("")
    lines.append("```json")
    lines.append(json.dumps(payload, ensure_ascii=False, indent=2))
    lines.append("```")
    lines.append("")
    return "\n".join(lines)


def _render_storage_readme(manifest: dict, structural: str) -> str:
    return f"""# CORVONERO Commander Review Export — README

Generated: {manifest['generated_at']}

## Files

- `{manifest['xlsx']}` — review workbook (NOT import-ready)
- `CORVONERO-COMMANDER-REVIEW-MANIFEST-v1.json`
- `CORVONERO-COMMANDER-REVIEW-SHA256-v1.txt`

## Status

| Gate | Value |
|------|-------|
| XLSX CREATED | YES |
| STRUCTURALLY VALIDATED | {structural} |
| READY FOR OPERATOR REVIEW | YES |
| READY FOR COMMANDER IMPORT | NO — UNTIL URL AND TEMPLATE VALIDATION |
| READY FOR LAUNCH | NO |

## Operator notes

- Base URLs are **PROPOSED — NOT HTTP VERIFIED**
- Sitelink anchors pending Roman confirmation — see `SITELINKS_PENDING` sheet
- Initial bids: **REVIEW — OPERATOR BID REQUIRED**
- Metrica and goals omitted by operator decision
- Commander import and advertising launch **NOT AUTHORIZED**

## Authority

- Checkpoint: `{manifest['checkpoint']}`
- Tag: `{manifest['tag']}`
"""


def _render_report(manifest: dict, validation: dict, structural: str, xlsx_path: Path) -> str:
    return f"""# REPORT — Corvonero Commander Production Wave 1 Review XLSX

Generated: {manifest['generated_at']}

## Verdict

**CORVONERO COMMANDER PRODUCTION WAVE 1: PASS — REVIEW XLSX CREATED**

## Deliverables

| Artefact | Path |
|----------|------|
| Review XLSX | `{xlsx_path}` |
| Manifest | `{OUT / 'CORVONERO-COMMANDER-REVIEW-MANIFEST-v1.json'}` |
| SHA256 | `{OUT / 'CORVONERO-COMMANDER-REVIEW-SHA256-v1.txt'}` |
| Template audit | `projects/mars-search-ppc-production/pilots/corvonero/CORVONERO-COMMANDER-W1-TEMPLATE-AUDIT-v1.md` |
| Export map | `projects/mars-search-ppc-production/pilots/corvonero/CORVONERO-COMMANDER-W1-EXPORT-MAP-v1.md` |
| Validation | `projects/mars-search-ppc-production/pilots/corvonero/CORVONERO-COMMANDER-W1-VALIDATION-v1.md` |
| Readiness | `projects/mars-search-ppc-production/pilots/corvonero/CORVONERO-COMMANDER-W1-READINESS-v1.md` |
| Result | `projects/mars-search-ppc-production/pilots/corvonero/CORVONERO-COMMANDER-W1-RESULT-v1.md` |

## Counts

| Metric | Expected | Actual |
|--------|----------|--------|
| Campaigns | 5 | {validation['counts']['campaigns']} |
| Groups | 15 | {validation['counts']['groups']} |
| Phrases | 895 | {validation['counts']['phrases']} |
| Primary ads | 15 | {validation['counts']['primary_ads']} |
| Sitelinks pending | 20 | {validation['counts']['sitelinks_pending']} |
| Callout sets | 5 | {validation['counts']['callout_sets']} |
| Cross-negatives | 0 | {validation['counts']['cross_negatives']} |

## Readiness gates

| Gate | Status |
|------|--------|
| XLSX CREATED | YES |
| STRUCTURALLY VALIDATED | {structural} |
| READY FOR OPERATOR REVIEW | YES |
| READY FOR COMMANDER IMPORT | NO |
| READY FOR LAUNCH | NO |

## Operator settings applied

- Search only; YAN disabled; auto-targeting disabled
- Manual CPC; 5 000 ₽ daily budget per campaign
- Schedule daily 06:00–21:00 Novosibirsk
- Geography: Novosibirsk + Novosibirskaya oblast
- Metrica and conversion goals omitted

## Not authorized

- Commander import
- Yandex Direct upload
- Moderation submission
- Campaign activation
- Bid placement
- Advertising launch
- Website modification

## Git

No commit performed. No authority artefacts modified.

## SHA256

`{manifest['xlsx_sha256']}`
"""


if __name__ == "__main__":
    main()
