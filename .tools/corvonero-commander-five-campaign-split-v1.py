#!/usr/bin/env python3
"""Corvonero Commander Wave 1 — split combined import candidate into five campaign workbooks."""

from __future__ import annotations

import hashlib
import json
import re
import subprocess
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

REPO = Path(r"C:\MARS Phenix\AI MARS")
STORAGE = Path(r"C:\MARS Phenix\AI MARS STORAGE")
PILOTS = REPO / "projects/mars-search-ppc-production/pilots/corvonero"
REPORTS = REPO / "projects/mars-search-ppc-production/reports"
TOOLS = REPO / ".tools"

REVIEW_XLSX = STORAGE / (
    "exports/corvonero/CORVONERO-COMMANDER-REVIEW-2026-06-29/"
    "CORVONERO-YANDEX-DIRECT-COMMANDER-REVIEW-v1.xlsx"
)
COMBINED_CANDIDATE = STORAGE / (
    "exports/corvonero/CORVONERO-COMMANDER-IMPORT-CANDIDATE-2026-06-29/"
    "CORVONERO-YANDEX-DIRECT-COMMANDER-IMPORT-CANDIDATE-v1.xlsx"
)
TEMPLATE_V1 = (
    REPO
    / "projects/orca/ppc/triumph-manipulator/assets/direct-commander-template/"
    "triumph-manipulator-commander-template-v1.xlsx"
)
HEADER_MAP = (
    REPO
    / "projects/orca/ppc/triumph-manipulator/tools/exporter-cli/commander-header-map-v0.json"
)
PATCH_SCRIPT = TOOLS / "corvonero-commander-import-patch-v1.cjs"
AUDIT_TEMP = STORAGE / "temp/corvonero-commander-five-campaign-split-2026-06-29"

OUT_DIR = STORAGE / (
    "exports/corvonero/CORVONERO-COMMANDER-5-CAMPAIGN-IMPORT-CANDIDATES-2026-06-29"
)

EXPECTED_TEMPLATE_SHA = "1112793a888ac2e0762317fa0bf728a116e36a143fc72fa0f5fe729c56c3f1fa"
GENERATED_AT = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
EXPORT_DATE = "2026-06-29"
GEO_REGION = "Новосибирск и Новосибирская область"
CALLOUT_JOIN = "||"

CAMPAIGN_ORDER = ["CA-01", "CA-02", "CA-03", "CA-04", "CA-05"]

CAMPAIGN_SPEC: dict[str, dict[str, Any]] = {
    "CA-01": {
        "filename": "CORVONERO-CA-01-PROGRAMMIST-1S-COMMANDER-IMPORT-v1.xlsx",
        "title": "CA-01 — Программист / специалист 1С",
        "campaign_name": "Программист / специалист 1С",
        "bid": 500,
        "landing_base": "https://lk.corvonero.ru/programmist-1s/",
        "utm_slug": "corv_programmist_1s",
        "groups": 3,
        "keywords": 404,
        "ads": 3,
    },
    "CA-02": {
        "filename": "CORVONERO-CA-02-SOPROVOZHDENIE-1S-COMMANDER-IMPORT-v1.xlsx",
        "title": "CA-02 — Сопровождение и обслуживание 1С",
        "campaign_name": "Сопровождение и обслуживание 1С",
        "bid": 400,
        "landing_base": "https://lk.corvonero.ru/soprovozhdenie-1s/",
        "utm_slug": "corv_soprovozhdenie_1s",
        "groups": 4,
        "keywords": 153,
        "ads": 4,
    },
    "CA-03": {
        "filename": "CORVONERO-CA-03-DORABOTKA-1S-COMMANDER-IMPORT-v1.xlsx",
        "title": "CA-03 — Доработка и разработка 1С",
        "campaign_name": "Доработка и разработка 1С",
        "bid": 400,
        "landing_base": "https://lk.corvonero.ru/dorabotka-razrabotka-1s/",
        "utm_slug": "corv_dorabotka_1s",
        "groups": 3,
        "keywords": 69,
        "ads": 3,
    },
    "CA-04": {
        "filename": "CORVONERO-CA-04-INTEGRACII-1S-COMMANDER-IMPORT-v1.xlsx",
        "title": "CA-04 — Интеграции 1С",
        "campaign_name": "Интеграции 1С",
        "bid": 400,
        "landing_base": "https://lk.corvonero.ru/integracii-1s/",
        "utm_slug": "corv_integracii_1s",
        "groups": 1,
        "keywords": 48,
        "ads": 1,
    },
    "CA-05": {
        "filename": "CORVONERO-CA-05-MARKIROVKA-1S-COMMANDER-IMPORT-v1.xlsx",
        "title": "CA-05 — Маркировка / Честный знак",
        "campaign_name": "Маркировка / Честный знак",
        "bid": 400,
        "landing_base": "https://lk.corvonero.ru/markirovka-chestny-znak/",
        "utm_slug": "corv_markirovka_1s",
        "groups": 4,
        "keywords": 221,
        "ads": 4,
    },
}

SHARED_NEGATIVES = [
    "вакансия",
    "работа программистом",
    "резюме",
    "сертификация",
    "кряк",
    "зарплата",
    "стань программистом",
    "становится программистом",
    "скачать",
]
CAMPAIGN_NEGATIVES_ALL = ["купить 1с", "лицензия 1с"]
CA05_EXTRA_NEGATIVE = "заказать коды маркировки"
FORBIDDEN_NEGATIVES = [
    "обучение",
    "курс",
    "курсы",
    "как сделать самому",
    "инструкция",
    "трактир",
    "erp",
]

TITLE_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def load_json_pilot(name: str) -> dict:
    return json.loads((PILOTS / name).read_text(encoding="utf-8"))


def format_negative(term: str, match: str = "word") -> str:
    term = term.strip()
    if match == "phrase" and " " in term:
        return f'-"{term}"'
    return f"-{term}"


def build_campaign_negatives(campaign_id: str) -> str:
    neg_data = load_json_pilot("CORVONERO-EXT-W1-NEGATIVE-DEPLOYMENT-v1.json")
    parts: list[str] = []
    for item in neg_data["layers"]["account_shared_deployable"]:
        parts.append(format_negative(item["term"], item["match"]))
    for item in neg_data["layers"]["campaign_deployable"]:
        if item["campaign_id"] == campaign_id:
            parts.append(format_negative(item["term"], item["match"]))
    return " ".join(parts)


def read_review_import_rows() -> list[dict]:
    wb = load_workbook(REVIEW_XLSX, read_only=True, data_only=True)
    ws = wb["COMMANDER_IMPORT"]
    header_row = 10
    header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = list(header_cells)
    col = {h: i for i, h in enumerate(headers) if h}

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or len(row) <= col["campaign_id (review)"]:
            continue
        cid = row[col["campaign_id (review)"]]
        if not cid:
            continue
        rows.append(
            {
                "row_type": row[col["row_type (review)"]],
                "campaign_id": cid,
                "campaign_name": row[col["campaign_name (review)"]],
                "group_id": row[col["ID группы"]],
                "group_name": row[col["Название группы"]],
                "group_number": row[col["Номер группы"]],
                "phrase_id": row[col["ID фразы"]] or "",
                "phrase": row[col["Фраза (с минус-словами)"]] or "",
                "ad_id": row[col["ID объявления"]] or "",
                "headline_1": row[col["Заголовок 1"]] or "",
                "headline_2": row[col["Заголовок 2"]] or "",
                "text": row[col["Текст"]] or "",
                "landing_url": row[col["Ссылка"]] or "",
                "display_path": row[col["Отображаемая ссылка"]] or "",
                "region": row[col["Регион"]] or GEO_REGION,
                "bid": row[col["Ставка"]],
                "ad_status": row[col["Статус объявления"]] or "",
                "phrase_status": row[col["Статус фразы"]] or "",
                "callouts": row[col["Уточнения"]] or "",
                "group_negatives": row[col["Минус-фразы на группу"]] or "",
                "campaign_negatives_ref": row[col["campaign_negatives_ref (review)"]]
                if "campaign_negatives_ref (review)" in col
                else "",
            }
        )
    wb.close()
    return rows


def split_campaign_rows(all_rows: list[dict], campaign_id: str) -> list[dict]:
    campaign_rows = [r for r in all_rows if r["campaign_id"] == campaign_id]
    group_order: list[str] = []
    for r in campaign_rows:
        if r["row_type"] == "AD" and r["group_id"] not in group_order:
            group_order.append(r["group_id"])
    group_num_map = {gid: i + 1 for i, gid in enumerate(group_order)}

    bid = CAMPAIGN_SPEC[campaign_id]["bid"]
    out: list[dict] = []
    for r in campaign_rows:
        row = dict(r)
        row["group_number"] = group_num_map.get(r["group_id"], r["group_number"])
        if row["row_type"] == "KEYWORD":
            row["bid"] = bid
        out.append(row)
    return out


def build_metadata_patches(campaign_id: str, campaign_rows: list[dict]) -> dict[str, str]:
    spec = CAMPAIGN_SPEC[campaign_id]
    return {
        "campaigns.campaign_type": "Текстово-графическая кампания",
        "campaigns.placement": "search",
        "campaigns.currency": "RUB",
        "campaigns.optimize_text": "0",
        "campaigns.promotion_url": spec["landing_base"],
        "campaigns.campaign_negatives": build_campaign_negatives(campaign_id),
    }


def run_node_patch(payload: dict, template: Path, output: Path) -> dict:
    payload_path = AUDIT_TEMP / f"payload-{output.stem}.json"
    AUDIT_TEMP.mkdir(parents=True, exist_ok=True)
    payload_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    cmd = [
        "node",
        str(PATCH_SCRIPT),
        str(payload_path),
        str(template),
        str(output),
        str(HEADER_MAP),
    ]
    proc = subprocess.run(cmd, capture_output=True, text=True, cwd=str(REPO))
    if proc.returncode != 0:
        raise RuntimeError(f"Node patch failed for {output.name}:\n{proc.stdout}\n{proc.stderr}")
    return json.loads(proc.stdout)


def read_workbook_data(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Тексты"]
    headers = [ws.cell(14, c).value for c in range(1, 79)]
    header_hash = hashlib.sha256(
        json.dumps([h for h in headers if h], ensure_ascii=False).encode()
    ).hexdigest()

    metadata: dict[str, str] = {}
    for r in range(6, 14):
        label = ws.cell(r, 1).value or ws.cell(r, 4).value
        val = ws.cell(r, 5).value or ws.cell(r, 8).value
        if label:
            metadata[str(label).strip()] = str(val or "").strip()

    kw_rows: list[dict] = []
    ad_rows: list[dict] = []
    for row in ws.iter_rows(min_row=16, max_col=78, values_only=True):
        if not row:
            continue
        group = row[4] if len(row) > 4 else None
        phrase = row[7] if len(row) > 7 else None
        headline = row[9] if len(row) > 9 else None
        if not group and not phrase and not headline:
            continue
        entry = {
            "group": group,
            "phrase": phrase,
            "bid": row[53] if len(row) > 53 else None,
            "url": row[47] if len(row) > 47 else None,
            "callouts": row[66] if len(row) > 66 else None,
            "sitelink_titles": row[57] if len(row) > 57 else None,
            "headline": headline,
        }
        if phrase:
            kw_rows.append(entry)
        elif headline or (group and not phrase):
            ad_rows.append(entry)

    wb.close()
    return {
        "headers": headers,
        "header_hash": header_hash,
        "metadata": metadata,
        "kw_rows": kw_rows,
        "ad_rows": ad_rows,
    }


def validate_workbook(
    path: Path,
    campaign_id: str,
    source_rows: list[dict],
) -> dict[str, Any]:
    spec = CAMPAIGN_SPEC[campaign_id]
    data = read_workbook_data(path)
    failures: list[str] = []

    if len(data["headers"]) != 78:
        failures.append(f"column_count: expected 78, got {len(data['headers'])}")
    if data["headers"][0] != "Доп. объявление группы":
        failures.append("header_row_14_drift")

    groups = len({r["group_id"] for r in source_rows if r["row_type"] == "AD"})
    kws = len(data["kw_rows"])
    ads = len(data["ad_rows"])

    for field, expected in (
        ("groups", spec["groups"]),
        ("keywords", spec["keywords"]),
        ("ads", spec["ads"]),
    ):
        actual = groups if field == "groups" else kws if field == "keywords" else ads
        if actual != expected:
            failures.append(f"{field}: expected {expected}, got {actual}")

    bid = spec["bid"]

    def bid_value(raw: Any) -> int | None:
        if raw in (None, ""):
            return None
        try:
            return int(float(str(raw).replace(",", ".")))
        except (TypeError, ValueError):
            return None

    bad_bids = [k for k in data["kw_rows"] if bid_value(k["bid"]) != bid]
    if bad_bids:
        failures.append(f"bids: expected all {bid}, mismatches {len(bad_bids)}")

    dup = [k for k, v in Counter((d["group"], d["phrase"]) for d in data["kw_rows"]).items() if v > 1]
    if dup:
        failures.append(f"duplicate_phrases: {len(dup)}")

    all_data = data["kw_rows"] + data["ad_rows"]
    if any(d.get("url") and ("utm_term=" in str(d["url"]) or "{keyword}" in str(d["url"])) for d in all_data):
        failures.append("utm_term_or_keyword_macro_leak")
    if any(d.get("url") and str(d["url"]).count("?") != 1 for d in all_data if d.get("url")):
        failures.append("malformed_url_question_mark")
    if any(d.get("sitelink_titles") for d in data["ad_rows"]):
        failures.append("sitelinks_present")

    neg_text = data["metadata"].get("Минус-фразы на кампанию:", "")
    for term in FORBIDDEN_NEGATIVES:
        if term in neg_text.lower():
            failures.append(f"forbidden_negative_leak: {term}")
    if campaign_id != "CA-05" and CA05_EXTRA_NEGATIVE in neg_text:
        failures.append("ca05_only_negative_in_non_ca05")
    if campaign_id == "CA-05" and CA05_EXTRA_NEGATIVE not in neg_text:
        failures.append("ca05_extra_negative_missing")

    callouts_data = load_json_pilot("CORVONERO-EXT-W1-CALLOUTS-v2.json")
    expected_callouts = CALLOUT_JOIN.join(
        c["text"] for c in callouts_data["campaign_pools"][campaign_id]
    )
    ad_callouts = [d["callouts"] for d in data["ad_rows"] if d.get("callouts")]
    if not ad_callouts or not all(c == expected_callouts for c in ad_callouts):
        failures.append("callouts_mismatch")

    if campaign_id == "CA-05":
        if any(
            d.get("callouts") and "Поддержка рабочей базы" in str(d["callouts"])
            for d in data["ad_rows"]
        ):
            failures.append("ca05_wrong_callout_support_base")
        if not any(
            d.get("callouts") and "Настройка маркировки" in str(d["callouts"])
            for d in data["ad_rows"]
        ):
            failures.append("ca05_markirovka_callout_missing")

    try:
        with zipfile.ZipFile(path, "r") as zf:
            zf.getinfo("xl/worksheets/sheet1.xml")
    except Exception as exc:
        failures.append(f"zip_xml_error: {exc}")

    return {
        "campaign_id": campaign_id,
        "filename": path.name,
        "sha256": sha256_file(path),
        "checks": {
            "groups": groups,
            "keywords": kws,
            "ads": ads,
            "bid": bid,
            "bids_populated": sum(
                1 for k in data["kw_rows"] if bid_value(k["bid"]) is not None
            ),
            "duplicate_phrases": len(dup),
            "header_hash": data["header_hash"],
        },
        "metadata": data["metadata"],
        "pass": len(failures) == 0,
        "failures": failures,
    }


def cross_validate(workbook_results: list[dict], all_phrases: list[tuple]) -> dict[str, Any]:
    failures: list[str] = []
    combined_groups = sum(r["checks"]["groups"] for r in workbook_results)
    combined_kw = sum(r["checks"]["keywords"] for r in workbook_results)
    combined_ads = sum(r["checks"]["ads"] for r in workbook_results)

    if len(workbook_results) != 5:
        failures.append(f"workbook_count: expected 5, got {len(workbook_results)}")
    if combined_groups != 15:
        failures.append(f"combined_groups: expected 15, got {combined_groups}")
    if combined_kw != 895:
        failures.append(f"combined_keywords: expected 895, got {combined_kw}")
    if combined_ads != 15:
        failures.append(f"combined_ads: expected 15, got {combined_ads}")

    dup_cross = [k for k, v in Counter(all_phrases).items() if v > 1]
    if dup_cross:
        failures.append(f"duplicate_phrases_across_workbooks: {len(dup_cross)}")

    return {
        "workbook_count": len(workbook_results),
        "unique_campaign_names": 5,
        "combined_groups": combined_groups,
        "combined_phrases": combined_kw,
        "combined_primary_ads": combined_ads,
        "duplicate_phrase_rows_across_workbooks": len(dup_cross),
        "pass": len(failures) == 0,
        "failures": failures,
    }


def style_header_row(ws, row: int, col_count: int) -> None:
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = TITLE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def create_support_workbook(out_path: Path, workbook_results: list[dict]) -> None:
    wb_src = load_workbook(REVIEW_XLSX, read_only=False, data_only=True)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for name in [
        "CAMPAIGN_SETTINGS",
        "GROUP_REGISTER",
        "PRIMARY_ADS",
        "COMBINATORIAL_ASSETS",
        "NEGATIVES",
        "SITELINKS_PENDING",
        "CALLOUTS",
        "URL_UTM_MAP",
    ]:
        if name not in wb_src.sheetnames:
            continue
        ws_src = wb_src[name]
        ws_dst = wb_out.create_sheet(name)
        for row in ws_src.iter_rows():
            for cell in row:
                ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)

    ws_map = wb_out.create_sheet("WORKBOOK_MAP")
    map_headers = [
        "import_order",
        "campaign_id",
        "campaign_title",
        "workbook_filename",
        "sha256",
        "groups",
        "keywords",
        "primary_ads",
        "initial_bid_rub",
        "daily_budget_rub_post_import",
        "landing_base_url",
        "utm_campaign_slug",
        "validation",
    ]
    for c, h in enumerate(map_headers, start=1):
        ws_map.cell(row=1, column=c, value=h)
    style_header_row(ws_map, 1, len(map_headers))

    import_sequence = [
        ("CA-01", "Программист 1С"),
        ("CA-02", "Сопровождение 1С"),
        ("CA-03", "Доработка 1С"),
        ("CA-04", "Интеграции 1С"),
        ("CA-05", "Маркировка 1С"),
    ]
    for i, (cid, short) in enumerate(import_sequence, start=1):
        spec = CAMPAIGN_SPEC[cid]
        result = next(r for r in workbook_results if r["campaign_id"] == cid)
        ws_map.append(
            [
                i,
                cid,
                spec["title"],
                spec["filename"],
                result["sha256"],
                spec["groups"],
                spec["keywords"],
                spec["ads"],
                spec["bid"],
                5000,
                spec["landing_base"],
                spec["utm_slug"],
                "PASS" if result["pass"] else "FAIL",
            ]
        )

    ws_post = wb_out.create_sheet("POST_IMPORT_SETTINGS")
    post_headers = [
        "campaign_id",
        "daily_budget_rub",
        "schedule",
        "timezone",
        "networks",
        "auto_targeting",
        "metrica",
        "goals",
        "activation",
    ]
    for c, h in enumerate(post_headers, start=1):
        ws_post.cell(row=1, column=c, value=h)
    style_header_row(ws_post, 1, len(post_headers))
    for cid in CAMPAIGN_ORDER:
        ws_post.append(
            [
                cid,
                5000,
                "Every day 06:00–21:00",
                "Novosibirsk",
                "DISABLED",
                "DISABLED",
                "Omitted",
                "Omitted",
                "Forbidden until operator approval",
            ]
        )

    ws_check = wb_out.create_sheet("LAUNCH_CHECKLIST")
    checklist = [
        "CORVONERO Commander 5-campaign support v1",
        f"Generated: {GENERATED_AT}",
        "",
        "Import sequence:",
        "1. CA-01 — Программист 1С",
        "2. CA-02 — Сопровождение 1С",
        "3. CA-03 — Доработка 1С",
        "4. CA-04 — Интеграции 1С",
        "5. CA-05 — Маркировка 1С",
        "",
        "Sitelinks: NOT in import workbooks — PENDING_FINAL_ANCHOR (20 in SITELINKS_PENDING)",
        "Base URLs: INCLUDED — NOT HTTP VERIFIED",
        "Daily budget / schedule: post-import Commander UI — 5000 RUB, daily 06:00–21:00 Novosibirsk",
        "Metrica / goals: OMITTED",
        "Commander import: NOT AUTHORIZED in this task",
        "Campaign activation: FORBIDDEN until operator approval",
        "Cross-campaign negatives: NONE",
    ]
    for i, line in enumerate(checklist, start=1):
        ws_check.cell(row=i, column=1, value=line)

    wb_out.save(out_path)
    wb_src.close()


def write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_md(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    AUDIT_TEMP.mkdir(parents=True, exist_ok=True)

    if not TEMPLATE_V1.exists():
        raise FileNotFoundError(f"Authentic template missing: {TEMPLATE_V1}")

    template_sha = sha256_file(TEMPLATE_V1)
    template_verified = template_sha == EXPECTED_TEMPLATE_SHA

    all_rows = read_review_import_rows()
    workbook_results: list[dict] = []
    all_phrases: list[tuple] = []
    patch_results: dict[str, Any] = {}

    for cid in CAMPAIGN_ORDER:
        spec = CAMPAIGN_SPEC[cid]
        out_path = OUT_DIR / spec["filename"]
        campaign_rows = split_campaign_rows(all_rows, cid)
        payload = {
            "geo_region": GEO_REGION,
            "metadata_patches": build_metadata_patches(cid, campaign_rows),
            "rows": campaign_rows,
        }
        patch_results[cid] = run_node_patch(payload, TEMPLATE_V1, out_path)
        result = validate_workbook(out_path, cid, campaign_rows)
        workbook_results.append(result)
        for r in campaign_rows:
            if r["row_type"] == "KEYWORD":
                all_phrases.append((r["group_id"], r["phrase"]))

    cross = cross_validate(workbook_results, all_phrases)
    all_pass = template_verified and all(r["pass"] for r in workbook_results) and cross["pass"]

    support_path = OUT_DIR / "CORVONERO-COMMANDER-5-CAMPAIGN-SUPPORT-v1.xlsx"
    create_support_workbook(support_path, workbook_results)

    manifest_entries = []
    sha_lines = []
    for r in workbook_results:
        cid = r["campaign_id"]
        spec = CAMPAIGN_SPEC[cid]
        manifest_entries.append(
            {
                "filename": r["filename"],
                "sha256": r["sha256"],
                "campaign_id": cid,
                "campaign_title": spec["title"],
                "group_count": spec["groups"],
                "keyword_count": spec["keywords"],
                "ad_count": spec["ads"],
                "bid_rub": spec["bid"],
                "template_source": str(TEMPLATE_V1),
                "template_sha256": template_sha,
                "template_verified": template_verified,
                "validation_result": "PASS" if r["pass"] else "FAIL",
                "validation_failures": r["failures"],
            }
        )
        sha_lines.append(f"{r['sha256']}  {r['filename']}")

    support_sha = sha256_file(support_path)
    sha_lines.append(f"{support_sha}  {support_path.name}")

    manifest = {
        "manifest_id": "corvonero-commander-5-campaign-manifest-v1",
        "generated_at": GENERATED_AT,
        "export_date": EXPORT_DATE,
        "authentic_template_source": str(TEMPLATE_V1),
        "authentic_template_sha256": template_sha,
        "authentic_template_sha256_expected": EXPECTED_TEMPLATE_SHA,
        "authentic_template_verified": template_verified,
        "combined_candidate_evidence": str(COMBINED_CANDIDATE),
        "workbooks": manifest_entries,
        "support_workbook": {
            "filename": support_path.name,
            "sha256": support_sha,
        },
        "cross_validation": cross,
        "totals": {
            "workbooks": 5,
            "groups": 15,
            "keywords": 895,
            "primary_ads": 15,
        },
        "verdict": "PASS — FIVE IMPORT-CANDIDATE XLSX FILES CREATED" if all_pass else "FAIL",
        "commander_import_performed": False,
        "advertising_started": False,
    }
    write_json(OUT_DIR / "CORVONERO-COMMANDER-5-CAMPAIGN-MANIFEST-v1.json", manifest)
    (OUT_DIR / "CORVONERO-COMMANDER-5-CAMPAIGN-SHA256-v1.txt").write_text(
        "\n".join(sha_lines) + "\n", encoding="utf-8"
    )

    readme = f"""# CORVONERO Commander Five-Campaign Import Candidates v1

**Generated:** {GENERATED_AT}  
**Status:** IMPORT-CANDIDATE — five independent Commander workbooks; **NOT** import-ready for launch.

## Authentic template

- Path: `{TEMPLATE_V1}`
- SHA-256: `{template_sha}`
- Verified: **{'YES' if template_verified else 'NO'}**

## Workbooks

| File | Campaign | Groups | Keywords | Ads | Bid |
|------|----------|--------|----------|-----|-----|
"""
    for cid in CAMPAIGN_ORDER:
        s = CAMPAIGN_SPEC[cid]
        r = next(x for x in workbook_results if x["campaign_id"] == cid)
        readme += (
            f"| `{s['filename']}` | {s['title']} | {s['groups']} | {s['keywords']} | "
            f"{s['ads']} | {s['bid']} RUB |\n"
        )
    readme += """
## Post-import (Commander UI)

- Daily budget: **5 000 RUB** each campaign
- Schedule: **Every day 06:00–21:00** (Novosibirsk)
- Networks: **DISABLED**
- Auto-targeting: **DISABLED**
- Metrica / goals: **Omitted**

## Blockers

- Base URLs: **NOT HTTP VERIFIED**
- Sitelinks: **NOT INCLUDED** — PENDING_FINAL_ANCHOR
- Commander import: **NOT PERFORMED**
- Advertising: **NOT STARTED**
"""
    write_md(OUT_DIR / "CORVONERO-COMMANDER-5-CAMPAIGN-README-v1.md", readme)

    split_map = {
        "map_id": "corvonero-commander-5-campaign-split-map-v1",
        "generated_at": GENERATED_AT,
        "source_evidence": {
            "review_workbook": str(REVIEW_XLSX),
            "combined_import_candidate": str(COMBINED_CANDIDATE),
            "note": "Combined candidate used as transformation evidence only; per-campaign metadata not copied from CA-01 block",
        },
        "template": {
            "path": str(TEMPLATE_V1),
            "sha256": template_sha,
            "verified": template_verified,
            "sheet": "Тексты",
            "header_row": 14,
            "columns": 78,
        },
        "campaigns": [
            {
                "campaign_id": cid,
                **{k: v for k, v in CAMPAIGN_SPEC[cid].items() if k != "filename"},
                "output_filename": CAMPAIGN_SPEC[cid]["filename"],
                "output_sha256": next(r["sha256"] for r in workbook_results if r["campaign_id"] == cid),
            }
            for cid in CAMPAIGN_ORDER
        ],
        "import_sequence": [
            {"order": i + 1, "campaign_id": cid, "label": label}
            for i, (cid, label) in enumerate(
                [
                    ("CA-01", "Программист 1С"),
                    ("CA-02", "Сопровождение 1С"),
                    ("CA-03", "Доработка 1С"),
                    ("CA-04", "Интеграции 1С"),
                    ("CA-05", "Маркировка 1С"),
                ]
            )
        ],
    }
    write_json(PILOTS / "CORVONERO-COMMANDER-5-CAMPAIGN-SPLIT-MAP-v1.json", split_map)
    write_md(
        PILOTS / "CORVONERO-COMMANDER-5-CAMPAIGN-SPLIT-MAP-v1.md",
        "# CORVONERO Commander 5-Campaign Split Map v1\n\n"
        f"**Generated:** {GENERATED_AT}\n\n"
        + "\n".join(
            f"- **{CAMPAIGN_SPEC[c]['title']}** → `{CAMPAIGN_SPEC[c]['filename']}` "
            f"({CAMPAIGN_SPEC[c]['groups']} groups, {CAMPAIGN_SPEC[c]['keywords']} kw, bid {CAMPAIGN_SPEC[c]['bid']} RUB)"
            for c in CAMPAIGN_ORDER
        ),
    )

    validation_json = {
        "validation_id": "corvonero-commander-5-campaign-validation-v1",
        "generated_at": GENERATED_AT,
        "template_sha256_verified": template_verified,
        "per_workbook": workbook_results,
        "cross_file": cross,
        "pass": all_pass,
    }
    write_json(PILOTS / "CORVONERO-COMMANDER-5-CAMPAIGN-VALIDATION-v1.json", validation_json)
    val_md = "# CORVONERO Commander 5-Campaign Validation v1\n\n"
    val_md += f"**Overall pass:** {all_pass}\n\n## Per workbook\n\n"
    for r in workbook_results:
        val_md += f"### {r['campaign_id']} — {r['filename']}\n\n"
        val_md += f"- Pass: **{r['pass']}**\n"
        for k, v in r["checks"].items():
            val_md += f"- {k}: {v}\n"
        if r["failures"]:
            val_md += "\nFailures:\n" + "\n".join(f"- {f}" for f in r["failures"]) + "\n"
        val_md += "\n"
    val_md += "## Cross-file\n\n" + "\n".join(f"- {k}: {v}" for k, v in cross.items() if k != "failures")
    if cross.get("failures"):
        val_md += "\n\nFailures:\n" + "\n".join(f"- {f}" for f in cross["failures"])
    write_md(PILOTS / "CORVONERO-COMMANDER-5-CAMPAIGN-VALIDATION-v1.md", val_md)

    readiness = {
        "readiness_id": "corvonero-commander-5-campaign-readiness-v1",
        "generated_at": GENERATED_AT,
        "five_workbooks_created": True,
        "ready_for_commander_import": False,
        "ready_for_launch": False,
        "verdict": "PASS — FIVE IMPORT-CANDIDATE XLSX FILES CREATED" if all_pass else "FAIL",
        "blockers": [
            "Base URLs NOT HTTP VERIFIED",
            "Sitelink anchors PENDING — 20 sitelinks omitted from import workbooks",
            "Daily budget and schedule require post-import Commander UI",
            "Metrica and conversion goals OMITTED",
            "Live Commander import not performed in this pass",
            "Campaign activation forbidden until operator approval",
        ],
        "commander_import_performed": False,
        "advertising_started": False,
    }
    write_json(PILOTS / "CORVONERO-COMMANDER-5-CAMPAIGN-READINESS-v1.json", readiness)
    write_md(
        PILOTS / "CORVONERO-COMMANDER-5-CAMPAIGN-READINESS-v1.md",
        f"# CORVONERO Commander 5-Campaign Readiness v1\n\n**Verdict:** {readiness['verdict']}\n\n"
        + "\n".join(f"- {b}" for b in readiness["blockers"]),
    )

    result = {
        "result_id": "corvonero-commander-5-campaign-result-v1",
        "generated_at": GENERATED_AT,
        "verdict": readiness["verdict"],
        "authentic_template_verified": template_verified,
        "workbook_count": 5,
        "groups": 15,
        "keywords": 895,
        "primary_ads": 15,
        "campaign_bids_rub": {c: CAMPAIGN_SPEC[c]["bid"] for c in CAMPAIGN_ORDER},
        "daily_budgets": "DOCUMENTED FOR POST-IMPORT — 5000 RUB EACH",
        "schedule": "DOCUMENTED FOR POST-IMPORT — DAILY 06:00–21:00",
        "sitelinks": "NOT INCLUDED — PENDING FINAL ANCHORS",
        "base_urls": "INCLUDED — NOT HTTP VERIFIED",
        "commander_import": "NOT PERFORMED",
        "advertising": "NOT STARTED",
        "all_validation_pass": all_pass,
        "output_directory": str(OUT_DIR),
    }
    write_json(PILOTS / "CORVONERO-COMMANDER-5-CAMPAIGN-RESULT-v1.json", result)

    report = f"""# REPORT — Corvonero Commander five-campaign import candidates v1

**Date:** {EXPORT_DATE}  
**Task:** CORVONERO COMMANDER WAVE 1 — SPLIT INTO FIVE CAMPAIGN IMPORT CANDIDATES

## Verdict

**{readiness['verdict']}**

| Field | Value |
|-------|-------|
| Authentic template verified | {'YES' if template_verified else 'NO'} |
| Template SHA-256 | `{template_sha}` |
| Campaign workbooks | 5 |
| Groups | 15 |
| Keywords | 895 |
| Primary ads | 15 |
| CA-01 bid | 500 RUB |
| CA-02 bid | 400 RUB |
| CA-03 bid | 400 RUB |
| CA-04 bid | 400 RUB |
| CA-05 bid | 400 RUB |
| Daily budgets | DOCUMENTED FOR POST-IMPORT — 5 000 RUB EACH |
| Schedule | DOCUMENTED FOR POST-IMPORT — DAILY 06:00–21:00 |
| Sitelinks | NOT INCLUDED — PENDING FINAL ANCHORS |
| Base URLs | INCLUDED — NOT HTTP VERIFIED |
| Commander import | NOT PERFORMED |
| Advertising | NOT STARTED |

## Output directory

`{OUT_DIR}`

## Files created

### STORAGE
"""
    for cid in CAMPAIGN_ORDER:
        report += f"- `{CAMPAIGN_SPEC[cid]['filename']}`\n"
    report += f"""- `CORVONERO-COMMANDER-5-CAMPAIGN-SUPPORT-v1.xlsx`
- `CORVONERO-COMMANDER-5-CAMPAIGN-MANIFEST-v1.json`
- `CORVONERO-COMMANDER-5-CAMPAIGN-SHA256-v1.txt`
- `CORVONERO-COMMANDER-5-CAMPAIGN-README-v1.md`

### Repository
- `projects/mars-search-ppc-production/pilots/corvonero/CORVONERO-COMMANDER-5-CAMPAIGN-SPLIT-MAP-v1.*`
- `projects/mars-search-ppc-production/pilots/corvonero/CORVONERO-COMMANDER-5-CAMPAIGN-VALIDATION-v1.*`
- `projects/mars-search-ppc-production/pilots/corvonero/CORVONERO-COMMANDER-5-CAMPAIGN-READINESS-v1.*`
- `projects/mars-search-ppc-production/pilots/corvonero/CORVONERO-COMMANDER-5-CAMPAIGN-RESULT-v1.*`
- `projects/mars-search-ppc-production/reports/REPORT-corvonero-commander-five-campaign-import-candidates-v1.md`

## Git status

No commit. No push. Authentic template unchanged.
"""
    write_md(PILOTS / "CORVONERO-COMMANDER-5-CAMPAIGN-RESULT-v1.md", f"# Result v1\n\n**{readiness['verdict']}**\n")
    write_md(REPORTS / "REPORT-corvonero-commander-five-campaign-import-candidates-v1.md", report)

    print(json.dumps(result, ensure_ascii=False, indent=2))
    if not all_pass:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
