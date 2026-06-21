#!/usr/bin/env python3
"""Generate BZPM MI Presentation Pack Excel workbooks from authority markdown sources."""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = Path(__file__).resolve().parent.parent
OUT = Path(__file__).resolve().parent
REGISTRY_MD = BASE / "BZPM-COMPETITOR-REGISTRY-v2.md"
MASTER_MD = BASE / "BZPM-MARKET-INTELLIGENCE-MASTER-REPORT-v1.md"
INSIGHTS_MD = BASE / "BZPM-OPERATOR-INSIGHTS-v1.md"
GEN_DATE = date.today().isoformat()

# --- styling helpers ---
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=11, color="333333")
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT_FILL = PatternFill("solid", fgColor="F2F7FB")
ZEBRA_FILL = PatternFill("solid", fgColor="FAFAFA")


def parse_md_table(text: str, start_marker: str, end_marker: str | None = None) -> list[list[str]]:
    """Extract markdown table rows between markers."""
    idx = text.find(start_marker)
    if idx == -1:
        return []
    chunk = text[idx:]
    if end_marker:
        end = chunk.find(end_marker, len(start_marker))
        if end != -1:
            chunk = chunk[:end]
    rows: list[list[str]] = []
    for line in chunk.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            if rows and not line.startswith("|"):
                break
            continue
        if re.match(r"^\|\s*[-:]+\s*\|", line):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        rows.append(cells)
    return rows


def normalize_status(raw: str) -> str:
    s = raw.strip()
    if s == "Approved":
        return "Approved"
    if "Strong Expansion" in s:
        return "Strong Expansion"
    if "Possible Expansion" in s:
        return "Possible Expansion"
    if s == "Deferred":
        return "Deferred"
    if s == "Excluded":
        return "Excluded"
    return s


def load_registry() -> list[dict]:
    text = REGISTRY_MD.read_text(encoding="utf-8")
    marker = "## TASK 3 — BZPM Competitor Registry v2"
    rows = parse_md_table(text, marker, "## TASK 5")
    header = rows[0]
    entities = []
    for row in rows[1:]:
        if len(row) < 9:
            continue
        entities.append(
            {
                "ID": row[0],
                "Company": row[1],
                "Website": row[2],
                "Tier": row[3],
                "Geography": row[4],
                "Coverage Zone": row[5],
                "Type": row[6],
                "Source Waves": row[7],
                "Status": normalize_status(row[8]),
            }
        )
    return entities


def style_header_row(ws, row: int, ncol: int):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def write_data_rows(ws, start_row: int, data: list[list], headers: list[str]):
    for ri, row in enumerate(data, start=start_row):
        fill = ALT_FILL if ri % 2 == 0 else ZEBRA_FILL
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if ri > start_row:
                cell.fill = fill


def add_table(ws, ref: str, name: str):
    tab = Table(displayName=name[:255], ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def autofit_columns(ws, min_w=10, max_w=55):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, min_w), max_w)


def freeze_and_filter(ws, cell: str = "A2"):
    ws.freeze_panes = cell
    ws.auto_filter.ref = ws.dimensions


def add_bar_chart(ws, title: str, data_col: int, cat_col: int, start: int, end: int, anchor: str):
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.y_axis.title = "Count"
    chart.x_axis.title = ""
    chart.style = 10
    chart.width = 16
    chart.height = 10
    data = Reference(ws, min_col=data_col, min_row=start, max_row=end)
    cats = Reference(ws, min_col=cat_col, min_row=start + 1, max_row=end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def add_pie_chart(ws, title: str, data_col: int, cat_col: int, start: int, end: int, anchor: str):
    chart = PieChart()
    chart.title = title
    chart.style = 10
    chart.width = 14
    chart.height = 10
    data = Reference(ws, min_col=data_col, min_row=start, max_row=end)
    cats = Reference(ws, min_col=cat_col, min_row=start + 1, max_row=end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def validate_sources(entities: list[dict]) -> dict:
    expected = {
        "total": 126,
        "Approved": 46,
        "Strong Expansion": 21,
        "Possible Expansion": 22,
        "Deferred": 26,
        "Excluded": 11,
    }
    counts = Counter(e["Status"] for e in entities)
    result = {
        "entity_count": len(entities),
        "status_counts": dict(counts),
        "expected": expected,
        "mismatches": [],
    }
    if len(entities) != expected["total"]:
        result["mismatches"].append(
            f"Total entities: parsed={len(entities)} expected={expected['total']}"
        )
    for key in ("Approved", "Strong Expansion", "Possible Expansion", "Deferred", "Excluded"):
        got = counts.get(key, 0)
        exp = expected[key]
        if got != exp:
            result["mismatches"].append(f"{key}: parsed={got} expected={exp}")
    return result


# --- workbook builders ---

def build_dashboard(entities: list[dict], validation: dict):
    wb = Workbook()
    approved = [e for e in entities if e["Status"] == "Approved"]

    # 01_Overview
    ws = wb.active
    ws.title = "01_Overview"
    ws["A1"] = "BZPM Market Intelligence — Executive Dashboard"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Generation date"
    ws["B3"] = GEN_DATE
    ws["A4"] = "Program"
    ws["B4"] = "BZPM Market Intelligence"
    ws["A5"] = "Stage"
    ws["B5"] = "Presentation & Packaging Layer (W3X + W3Y Approved)"
    metrics = [
        ("Total canonical entities", validation["entity_count"]),
        ("Approved registry", validation["status_counts"].get("Approved", 0)),
        ("Strong expansion candidates", validation["status_counts"].get("Strong Expansion", 0)),
        ("Possible expansion candidates", validation["status_counts"].get("Possible Expansion", 0)),
        ("Deferred", validation["status_counts"].get("Deferred", 0)),
        ("Excluded", validation["status_counts"].get("Excluded", 0)),
    ]
    ws["A7"] = "Key metrics"
    ws["A7"].font = SUBTITLE_FONT
    metric_rows = [[m, v] for m, v in metrics]
    ws.append(["Metric", "Value"])
    style_header_row(ws, 9, 2)
    for m, v in metrics:
        ws.append([m, v])
    write_data_rows(ws, 9, metric_rows, ["Metric", "Value"])
    add_table(ws, "A9:B14", "OverviewMetrics")
    ws["A16"] = "Waves completed"
    ws["A16"].font = SUBTITLE_FONT
    waves = [
        ("W1", "Market Mapping", "Approved"),
        ("W2", "Competitor Discovery", "Approved"),
        ("W2.5", "Competitor Prioritization", "Approved"),
        ("W3", "Competitor Registry", "Approved"),
        ("W3R", "Regional Reinforcement", "Approved"),
        ("W3S", "SERP Visibility Expansion", "Approved"),
        ("W3X", "Registry Consolidation", "Approved"),
        ("W3Y", "Operator Insight Capture", "Approved"),
    ]
    ws.append(["Wave", "Name", "Status"])
    style_header_row(ws, 18, 3)
    for w in waves:
        ws.append(list(w))
    write_data_rows(ws, 18, [list(w) for w in waves], ["Wave", "Name", "Status"])
    add_table(ws, "A18:C25", "WaveStatus")
    autofit_columns(ws)

    # 02_Geography
    ws2 = wb.create_sheet("02_Geography")
    ws2["A1"] = "Geography Overview"
    ws2["A1"].font = TITLE_FONT
    geo_counter: Counter = Counter()
    for e in entities:
        g = e["Geography"].split("(")[0].strip()
        if g.startswith("Russia"):
            geo_counter["Russia"] += 1
        elif "Kazakhstan" in g:
            geo_counter["Kazakhstan"] += 1
        elif "Belarus" in g:
            geo_counter["Belarus"] += 1
        elif "International" in g:
            geo_counter["International"] += 1
        else:
            geo_counter[g or "Unknown"] += 1
    ws2.append(["Geography", "Entity Count"])
    style_header_row(ws2, 2, 2)
    chart_start = 2
    for g, c in sorted(geo_counter.items(), key=lambda x: -x[1]):
        ws2.append([g, c])
    write_data_rows(ws2, 2, [[g, c] for g, c in sorted(geo_counter.items(), key=lambda x: -x[1])], ["Geography", "Entity Count"])
    add_table(ws2, f"A2:B{1 + len(geo_counter)}", "GeoTable")
    add_bar_chart(ws2, "Entities by Geography", 2, 1, chart_start, 1 + len(geo_counter), "D2")
    autofit_columns(ws2)

    # 03_Tiers
    ws3 = wb.create_sheet("03_Tiers")
    ws3["A1"] = "Tier Distribution (Approved Registry)"
    ws3["A1"].font = TITLE_FONT
    tier_labels = {
        "A": "A — Direct Competitors / OEM",
        "B": "B — Federal Leaders",
        "C": "C — Strategic Regional Coverage",
        "D": "D — Aggregators",
        "E": "E — Industry Marketplaces",
        "F": "F — International References",
    }
    tier_counter = Counter(e["Tier"] for e in approved if e["Tier"] in tier_labels)
    ws3.append(["Tier", "Label", "Count"])
    style_header_row(ws3, 2, 3)
    tier_rows = []
    for t in "ABCDEF":
        tier_rows.append([t, tier_labels[t], tier_counter.get(t, 0)])
    for r in tier_rows:
        ws3.append(r)
    write_data_rows(ws3, 2, tier_rows, ["Tier", "Label", "Count"])
    add_table(ws3, "A2:C7", "TierTable")
    add_bar_chart(ws3, "Entities by Tier (Approved)", 3, 2, 2, 7, "E2")
    autofit_columns(ws3)

    # 04_SERP
    ws4 = wb.create_sheet("04_SERP")
    ws4["A1"] = "SERP Visibility Leaders (W3S)"
    ws4["A1"].font = TITLE_FONT
    serp = [
        ("Trapeza", "https://www.trapeza.ru/", 8, "Approved"),
        ("КЛЕН", "https://www.klenmarket.ru/", 7, "Approved"),
        ("РЕФРО", "https://www.refro.ru/", 5, "Approved"),
        ("Практика (Pectopah)", "https://www.pectopah.ru/", 5, "Strong Expansion"),
        ("Завод Проммаш", "https://prommash.com/", 5, "Strong Expansion"),
        ("Kobor", "https://kobor.ru/", 5, "Strong Expansion"),
        ("Abat", "https://abat.ru/", 4, "Approved"),
        ("Restoll", "https://restoll.ru/", 4, "Approved"),
        ("Finist", "https://f-inox.ru/", 4, "Approved"),
        ("Ресторан Комплект", "https://r-komplekt.ru/", 4, "Strong Expansion"),
        ("КАМИК", "https://kamik-group.ru/", 4, "Strong Expansion"),
        ("Завод МетаКон", "https://zavod-metakon.ru/", 4, "Strong Expansion"),
        ("НОТИС", "https://www.notis.ru/", 4, "Strong Expansion"),
        ("Энтерo", "https://entero.ru/", 3, "Approved"),
        ("Комплекс Трейд", "https://kompleks-trade.ru/", 3, "Strong Expansion"),
    ]
    ws4.append(["Company", "Website", "Appearances", "Registry Status"])
    style_header_row(ws4, 2, 4)
    for row in serp:
        ws4.append(list(row))
    write_data_rows(ws4, 2, [list(r) for r in serp], ["Company", "Website", "Appearances", "Registry Status"])
    add_table(ws4, f"A2:D{1 + len(serp)}", "SerpTable")
    add_bar_chart(ws4, "SERP Visibility Leaders", 3, 1, 2, 1 + len(serp), "F2")
    autofit_columns(ws4)

    # 05_Program_Status
    ws5 = wb.create_sheet("05_Program_Status")
    ws5["A1"] = "Program Status Summary"
    ws5["A1"].font = TITLE_FONT
    status_groups = [
        ("Approved", validation["status_counts"].get("Approved", 0)),
        ("Strong Expansion", validation["status_counts"].get("Strong Expansion", 0)),
        ("Possible Expansion", validation["status_counts"].get("Possible Expansion", 0)),
        ("Deferred", validation["status_counts"].get("Deferred", 0)),
        ("Excluded", validation["status_counts"].get("Excluded", 0)),
    ]
    ws5.append(["Status Group", "Count"])
    style_header_row(ws5, 2, 2)
    for sg in status_groups:
        ws5.append(list(sg))
    write_data_rows(ws5, 2, [list(sg) for sg in status_groups], ["Status Group", "Count"])
    add_table(ws5, "A2:B6", "StatusTable")
    add_pie_chart(ws5, "Approved vs Expansion vs Deferred", 2, 1, 2, 6, "D2")

    # Regional coverage
    ws5["A10"] = "Regional Coverage (multi-value, all entities)"
    ws5["A10"].font = SUBTITLE_FONT
    zone_counter: Counter = Counter()
    for e in entities:
        for z in re.split(r"[;,]", e["Coverage Zone"]):
            z = z.strip()
            if z and z != "—":
                zone_counter[z] += 1
    ws5.append(["Coverage Zone", "Entity Count"])
    style_header_row(ws5, 11, 2)
    zone_rows = sorted(zone_counter.items(), key=lambda x: -x[1])
    for z, c in zone_rows:
        ws5.append([z, c])
    write_data_rows(ws5, 11, [[z, c] for z, c in zone_rows], ["Coverage Zone", "Entity Count"])
    end_row = 11 + len(zone_rows)
    add_table(ws5, f"A11:B{end_row}", "ZoneTable")
    add_bar_chart(ws5, "Regional Coverage", 2, 1, 11, end_row, "D10")
    autofit_columns(ws5)

    wb.save(OUT / "BZPM-MI-DASHBOARD.xlsx")
    return 5  # chart count in dashboard


def build_competitor_registry(entities: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Full_Registry"
    headers = ["ID", "Company", "Website", "Tier", "Geography", "Coverage Zone", "Type", "Source Waves", "Status"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    rows = [
        [
            e["ID"],
            e["Company"],
            e["Website"],
            e["Tier"],
            e["Geography"],
            e["Coverage Zone"],
            e["Type"],
            e["Source Waves"],
            e["Status"],
        ]
        for e in entities
    ]
    for r in rows:
        ws.append(r)
    write_data_rows(ws, 1, rows, headers)
    add_table(ws, f"A1:I{1 + len(rows)}", "CompetitorRegistry")
    freeze_and_filter(ws, "A2")
    autofit_columns(ws)
    wb.save(OUT / "BZPM-COMPETITOR-REGISTRY.xlsx")


def apply_status_conditional(ws, status_col: int, start: int, end: int):
    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    orange = PatternFill("solid", fgColor="FCE4D6")
    gray = PatternFill("solid", fgColor="D9D9D9")
    red = PatternFill("solid", fgColor="FFC7CE")
    letter = get_column_letter(status_col)
    rng = f"{letter}{start}:{letter}{end}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Approved"'], fill=green))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Strong Expansion"'], fill=yellow))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Possible Expansion"'], fill=orange))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Deferred"'], fill=gray))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Excluded"'], fill=red))


def build_core_research_set(entities: list[dict]):
    wb = Workbook()
    approved = [e for e in entities if e["Status"] == "Approved"]
    strong = [e for e in entities if e["Status"] == "Strong Expansion"]
    benchmark = [
        {"Company": "УЗНМ", "Website": "https://zavod-uznm.ru/", "ID": "COMP-BZPM-007", "Why": "Tier A OEM peer; simplified filter pattern flagged"},
        {"Company": "КЛЕН", "Website": "https://www.klenmarket.ru/", "ID": "COMP-BZPM-012", "Why": "Federal leader; view switcher pattern; high W3S visibility (7)"},
        {"Company": "ГК Юниторг", "Website": "https://www.unitorg.ru/", "ID": "CAN-EXP-005", "Why": "Ural regional candidate; dual-column card layout flagged"},
        {"Company": "Trapeza (Деловая Русь)", "Website": "https://www.trapeza.ru/", "ID": "COMP-BZPM-011", "Why": "Federal anchor; information density; highest W3S visibility (8)"},
        {"Company": "Kobor", "Website": "https://kobor.ru/", "ID": "CAN-EXP-003", "Why": "Repeated operator return; Strong expansion; multi-region subsites"},
        {"Company": "Комплекс Трейд", "Website": "https://kompleks-trade.ru/", "ID": "CAN-EXP-004", "Why": "Repeated operator return; Strong expansion; multi-city presence"},
    ]
    serp = [
        ("Trapeza", "https://www.trapeza.ru/", 8, "COMP-BZPM-011", "Approved"),
        ("КЛЕН", "https://www.klenmarket.ru/", 7, "COMP-BZPM-012", "Approved"),
        ("РЕФРО", "https://www.refro.ru/", 5, "COMP-BZPM-016", "Approved"),
        ("Практика (Pectopah)", "https://www.pectopah.ru/", 5, "CAN-EXP-014", "Strong Expansion"),
        ("Завод Проммаш", "https://prommash.com/", 5, "CAN-EXP-015", "Strong Expansion"),
        ("Kobor", "https://kobor.ru/", 5, "CAN-EXP-003", "Strong Expansion"),
        ("Abat", "https://abat.ru/", 4, "COMP-BZPM-001", "Approved"),
        ("Restoll", "https://restoll.ru/", 4, "COMP-BZPM-018", "Approved"),
        ("Finist", "https://f-inox.ru/", 4, "COMP-BZPM-003", "Approved"),
        ("Ресторан Комплект", "https://r-komplekt.ru/", 4, "CAN-EXP-016", "Strong Expansion"),
        ("КАМИК", "https://kamik-group.ru/", 4, "CAN-EXP-017", "Strong Expansion"),
        ("Завод МетаКон", "https://zavod-metakon.ru/", 4, "CAN-EXP-018", "Strong Expansion"),
        ("НОТИС", "https://www.notis.ru/", 4, "CAN-EXP-019", "Strong Expansion"),
    ]

    def write_sheet(ws, title, headers, data_rows, status_col=None):
        ws["A1"] = title
        ws["A1"].font = TITLE_FONT
        ws.append(headers)
        style_header_row(ws, 2, len(headers))
        for r in data_rows:
            ws.append(r)
        write_data_rows(ws, 2, data_rows, headers)
        end = 1 + len(data_rows)
        add_table(ws, f"A2:{get_column_letter(len(headers))}{1 + len(data_rows)}", ws.title.replace(" ", "")[:20])
        if status_col:
            apply_status_conditional(ws, status_col, 3, 1 + len(data_rows))
        freeze_and_filter(ws, "A3")
        autofit_columns(ws)

    ws1 = wb.active
    ws1.title = "01_Approved_Registry"
    h1 = ["ID", "Company", "Website", "Tier", "Geography", "Coverage Zone", "Type", "Status"]
    d1 = [[e["ID"], e["Company"], e["Website"], e["Tier"], e["Geography"], e["Coverage Zone"], e["Type"], e["Status"]] for e in approved]
    write_sheet(ws1, "Approved Registry (46 entities)", h1, d1, status_col=8)

    ws2 = wb.create_sheet("02_Strong_Expansion")
    h2 = ["ID", "Company", "Website", "Tier", "Geography", "Coverage Zone", "Type", "Status"]
    d2 = [[e["ID"], e["Company"], e["Website"], e["Tier"], e["Geography"], e["Coverage Zone"], e["Type"], e["Status"]] for e in strong]
    write_sheet(ws2, "Strong Expansion Candidates (21 entities)", h2, d2, status_col=8)

    ws3 = wb.create_sheet("03_Native_Benchmark_Group")
    h3 = ["Company", "Website", "Registry ID", "Why Operator Attention"]
    d3 = [[b["Company"], b["Website"], b["ID"], b["Why"]] for b in benchmark]
    write_sheet(ws3, "Native Benchmark Group (W3Y)", h3, d3)

    ws4 = wb.create_sheet("04_SERP_Leaders")
    h4 = ["Company", "Website", "Appearances", "Registry ID", "Status"]
    d4 = [list(r) for r in serp]
    write_sheet(ws4, "SERP Visibility Leaders", h4, d4, status_col=5)

    wb.save(OUT / "BZPM-CORE-RESEARCH-SET.xlsx")


def build_operator_insights():
    text = INSIGHTS_MD.read_text(encoding="utf-8")
    wb = Workbook()

    # Highlights
    ws1 = wb.active
    ws1.title = "01_Highlights"
    ws1["A1"] = "Operator Highlight Registry (W3Y)"
    ws1["A1"].font = TITLE_FONT
    h1 = ["Company", "URL", "Registry ID", "Observation Type", "Operator Comment Summary"]
    rows1 = parse_md_table(text, "## Operator Highlight Registry", "## Observed Pattern Registry")
    if rows1:
        rows1 = rows1[1:]
    ws1.append(h1)
    style_header_row(ws1, 2, len(h1))
    for r in rows1:
        ws1.append(r[:5])
    write_data_rows(ws1, 2, [r[:5] for r in rows1], h1)
    add_table(ws1, f"A2:E{1 + len(rows1)}", "Highlights")
    freeze_and_filter(ws1, "A3")
    autofit_columns(ws1)

    # Patterns - flatten all pattern categories
    ws2 = wb.create_sheet("02_Patterns")
    ws2["A1"] = "Observed Pattern Registry (W3Y)"
    ws2["A1"].font = TITLE_FONT
    h2 = ["Category", "Company", "Registry ID", "Operator Observation"]
    ws2.append(h2)
    style_header_row(ws2, 2, len(h2))
    pattern_data = []
    categories = [
        "Catalog Layout",
        "Product Listing Layout",
        "View Switchers",
        "Filter UX",
        "Information Density",
        "Product Card Structure",
        "Navigation Pattern",
        "Commercial Architecture",
    ]
    for cat in categories:
        marker = f"### {cat}"
        chunk = text.split(marker)
        if len(chunk) < 2:
            continue
        sub = chunk[1].split("###")[0]
        tbl = parse_md_table(sub, "| Company |", None)
        if not tbl:
            tbl = parse_md_table(sub, "| Company | Registry ID |", None)
        for row in tbl[1:] if tbl else []:
            if len(row) >= 3 and row[0].strip() != "—":
                pattern_data.append([cat, row[0], row[1], row[2] if len(row) > 2 else ""])
    for r in pattern_data:
        ws2.append(r)
    write_data_rows(ws2, 2, pattern_data, h2)
    add_table(ws2, f"A2:D{1 + len(pattern_data)}", "Patterns")
    freeze_and_filter(ws2, "A3")
    autofit_columns(ws2)

    # FIM
    ws3 = wb.create_sheet("03_FIM_Registry")
    ws3["A1"] = "Future Investigation Markers (W3Y)"
    ws3["A1"].font = TITLE_FONT
    h3 = ["Marker ID", "Subject", "Registry ID", "Operator Observation", "Suggested Review Surface"]
    fim_rows = parse_md_table(text, "## Future Investigation Markers", "## Program Impact")
    if fim_rows:
        fim_rows = fim_rows[1:]
    ws3.append(h3)
    style_header_row(ws3, 2, len(h3))
    for r in fim_rows:
        ws3.append(r[:5])
    write_data_rows(ws3, 2, [r[:5] for r in fim_rows], h3)
    add_table(ws3, f"A2:E{1 + len(fim_rows)}", "FIMRegistry")
    freeze_and_filter(ws3, "A3")
    autofit_columns(ws3)

    # Benchmark
    ws4 = wb.create_sheet("04_Benchmark_Group")
    ws4["A1"] = "Native Benchmark Group (W3Y)"
    ws4["A1"].font = TITLE_FONT
    h4 = ["Company", "URL", "Registry ID", "Why Operator Attention", "Review Surfaces Touched"]
    bench_rows = parse_md_table(text, "## Native Benchmark Group", "## Future Investigation Markers")
    if bench_rows:
        bench_rows = bench_rows[1:]
    ws4.append(h4)
    style_header_row(ws4, 2, len(h4))
    for r in bench_rows:
        ws4.append(r[:5])
    write_data_rows(ws4, 2, [r[:5] for r in bench_rows], h4)
    add_table(ws4, f"A2:E{1 + len(bench_rows)}", "BenchmarkGroup")
    freeze_and_filter(ws4, "A3")
    autofit_columns(ws4)

    wb.save(OUT / "BZPM-OPERATOR-INSIGHTS.xlsx")
    return {
        "highlights": len(rows1),
        "patterns": len(pattern_data),
        "fim": len(fim_rows),
        "benchmark": len(bench_rows),
    }


def build_manual_review_checklist(entities: list[dict]):
    tier_a = [e for e in entities if e["Status"] == "Approved" and e["Tier"] == "A"]
    strong = [e for e in entities if e["Status"] == "Strong Expansion"]
    benchmark_ids = {
        "COMP-BZPM-007",
        "COMP-BZPM-012",
        "CAN-EXP-005",
        "COMP-BZPM-011",
        "CAN-EXP-003",
        "CAN-EXP-004",
    }
    by_id = {e["ID"]: e for e in entities}
    benchmark = [by_id[i] for i in benchmark_ids if i in by_id]

    seen = set()
    checklist = []
    for group_name, group in [("Tier A OEM", tier_a), ("Strong Expansion", strong), ("Native Benchmark", benchmark)]:
        for e in group:
            key = e["ID"]
            if key in seen:
                continue
            seen.add(key)
            checklist.append(
                {
                    "Company": e["Company"],
                    "Website": e["Website"],
                    "Source Group": group_name,
                }
            )

    wb = Workbook()
    ws = wb.active
    ws.title = "Review_Checklist"
    ws["A1"] = "BZPM Manual Review Checklist"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Populated: Tier A OEM + Strong Expansion + Native Benchmark Group | {GEN_DATE}"
    headers = ["Company", "Website", "Source Group", "Opened", "Interesting", "Keep", "Priority", "Notes"]
    ws.append([])
    ws.append(headers)
    hdr_row = 4
    style_header_row(ws, hdr_row, len(headers))
    for item in checklist:
        ws.append([item["Company"], item["Website"], item["Source Group"], "", "", "", "", ""])
    write_data_rows(ws, hdr_row, [[c["Company"], c["Website"], c["Source Group"], "", "", "", "", ""] for c in checklist], headers)
    end = hdr_row + len(checklist)
    add_table(ws, f"A{hdr_row}:{get_column_letter(len(headers))}{end}", "ReviewChecklist")
    freeze_and_filter(ws, f"A{hdr_row + 1}")

    dv_yes_no = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_priority = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(dv_yes_no)
    ws.add_data_validation(dv_priority)
    for col_letter in ("D", "E", "F"):
        dv_yes_no.add(f"{col_letter}{hdr_row + 1}:{col_letter}{end}")
    dv_priority.add(f"G{hdr_row + 1}:G{end}")
    autofit_columns(ws)
    wb.save(OUT / "BZPM-MANUAL-REVIEW-CHECKLIST.xlsx")
    return len(checklist)


def build_package_summary(entities: list[dict], validation: dict):
    wb = Workbook()
    approved = [e for e in entities if e["Status"] == "Approved"]

    ws = wb.active
    ws.title = "Executive_Summary"
    ws["A1"] = "BZPM Market Intelligence — Package Summary"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Client-facing overview | No internal IDs | Generated " + GEN_DATE
    ws["A4"] = "Market Overview"
    ws["A4"].font = SUBTITLE_FONT
    overview = [
        "Program: BZPM Market Intelligence for neutral commercial kitchen equipment (ЗПМ).",
        "Primary geography: Siberia, Ural, Far East (Barnaul HQ context).",
        f"Research universe: {validation['entity_count']} canonical competitor entities after deduplication.",
        f"Approved registry baseline: {validation['status_counts'].get('Approved', 0)} companies with full field coverage.",
        f"Expansion pipeline: {validation['status_counts'].get('Strong Expansion', 0)} strong + {validation['status_counts'].get('Possible Expansion', 0)} possible candidates.",
        "Waves W1–W3Y completed and approved; consolidation and operator insights captured.",
    ]
    r = 5
    for line in overview:
        ws.cell(row=r, column=1, value=line)
        r += 1

    ws.cell(row=r + 1, column=1, value="Geography Overview").font = SUBTITLE_FONT
    r += 2
    geo = Counter()
    for e in approved:
        g = e["Geography"]
        if g.startswith("Russia"):
            geo["Russia"] += 1
        elif "Kazakhstan" in g:
            geo["Kazakhstan"] += 1
        elif "Belarus" in g:
            geo["Belarus"] += 1
        elif "International" in g:
            geo["International"] += 1
    ws.append(["Region", "Approved Companies"])
    style_header_row(ws, r, 2)
    geo_rows = list(geo.items())
    for gr in geo_rows:
        ws.append(list(gr))
    write_data_rows(ws, r, geo_rows, ["Region", "Approved Companies"])
    r += len(geo_rows) + 2

    ws.cell(row=r, column=1, value="Tier Overview (Approved Registry)").font = SUBTITLE_FONT
    r += 1
    tier_names = {
        "A": "Direct Competitors / OEM",
        "B": "Federal Leaders",
        "C": "Strategic Regional Coverage",
        "D": "Aggregators",
        "E": "Industry Marketplaces",
        "F": "International References",
    }
    ws.append(["Tier", "Description", "Count"])
    style_header_row(ws, r, 3)
    tier_rows = []
    tc = Counter(e["Tier"] for e in approved)
    for t in "ABCDEF":
        tier_rows.append([t, tier_names[t], tc.get(t, 0)])
    for tr in tier_rows:
        ws.append(tr)
    write_data_rows(ws, r, tier_rows, ["Tier", "Description", "Count"])
    r += len(tier_rows) + 2

    ws.cell(row=r, column=1, value="Key Competitors").font = SUBTITLE_FONT
    r += 1
    key = [
        ("Abat", "https://abat.ru/", "Tier A OEM — direct competitor"),
        ("Trapeza", "https://www.trapeza.ru/", "Tier B federal leader — highest SERP visibility"),
        ("КЛЕН", "https://www.klenmarket.ru/", "Tier B federal leader"),
        ("УЗНМ", "https://zavod-uznm.ru/", "Tier A OEM — Ural/Siberia coverage"),
        ("Kobor", "https://kobor.ru/", "Strong expansion — multi-region presence"),
        ("Комплекс Трейд", "https://kompleks-trade.ru/", "Strong expansion — Siberia/Ural/Far East"),
        ("Rational", "https://www.rational-online.com/", "International reference (W1D anchor)"),
        ("Hoshizaki", "https://www.hoshizaki.com/", "International reference (W1D anchor)"),
    ]
    ws.append(["Company", "Website", "Role"])
    style_header_row(ws, r, 3)
    for k in key:
        ws.append(list(k))
    write_data_rows(ws, r, [list(k) for k in key], ["Company", "Website", "Role"])
    r += len(key) + 2

    ws.cell(row=r, column=1, value="Research Coverage").font = SUBTITLE_FONT
    r += 1
    coverage = [
        ("Discovery waves", "W1 Market Mapping through W3S SERP expansion"),
        ("Registry consolidation", "W3X — 126 canonical entities"),
        ("Operator insights", "W3Y — patterns and benchmark group captured"),
        ("Regional focus", "Siberia, Ural, Far East, Kazakhstan, Belarus documented"),
        ("SERP visibility", "Top leaders mapped; regional blind spots identified"),
        ("Next step", "Operator review and W4 competitor intelligence (not in this pack)"),
    ]
    ws.append(["Area", "Coverage"])
    style_header_row(ws, r, 2)
    for c in coverage:
        ws.append(list(c))
    write_data_rows(ws, r, [list(c) for c in coverage], ["Area", "Coverage"])
    autofit_columns(ws)
    wb.save(OUT / "BZPM-MI-PACKAGE-SUMMARY.xlsx")


def main():
    entities = load_registry()
    validation = validate_sources(entities)
    chart_count = build_dashboard(entities, validation)
    build_competitor_registry(entities)
    build_core_research_set(entities)
    insights_stats = build_operator_insights()
    checklist_count = build_manual_review_checklist(entities)
    build_package_summary(entities, validation)

    print("VALIDATION:", validation)
    print("CHECKLIST ROWS:", checklist_count)
    print("INSIGHTS:", insights_stats)
    print("DASHBOARD CHARTS:", chart_count)
    print("OUTPUT:", OUT)


if __name__ == "__main__":
    main()
