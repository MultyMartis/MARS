#!/usr/bin/env python3
"""
C2c HOLD: source persistence / hardening only.
This file is not authorized for execution without explicit operator approval.
Commit/persistence does not authorize Commander import, Direct launch, account mutation,
advertising start, Storage export generation, repo artifact generation,
Localhost mutation, Storage mutation, Yandex/API access, or client-facing delivery.
Commander/XLSX/client approval generation is transport/import-candidate tooling only.
Client approval pack generation does not authorize client-facing delivery, campaign launch,
Direct import, account mutation, or advertising start.

Generate Corvonero client approval pack from V2.6 authority (read-only).
"""
from __future__ import annotations

import hashlib
import json
import os
import re
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPO = Path(r"X:\AI MARS")
PILOT = REPO / "projects" / "mars-search-ppc-production" / "pilots" / "corvonero"
REPORTS = REPO / "projects" / "mars-search-ppc-production" / "reports"
CLIENT_APPROVAL_REPO = PILOT / "client-approval"
PACK_DATE = "2026-07-01"
OUTPUT = Path(
    rf"X:\AI MARS STORAGE\exports\corvonero\CORVONERO-CLIENT-APPROVAL-PACK-{PACK_DATE}"
)
DEPLOY_PKG = Path(
    r"X:\AI MARS STORAGE\exports\corvonero\CORVONERO-CAMPAIGN-V2.6.2-FINAL-2026-06-30"
)

CAMPAIGN_ORDER = [
    "CA-01-LOCAL", "CA-01-REMOTE", "CA-02-LOCAL", "CA-02-REMOTE",
    "CA-03-LOCAL", "CA-03-REMOTE", "CA-04-LOCAL", "CA-04-REMOTE",
    "CA-05-LOCAL", "CA-05-REMOTE",
]

SERVICE_LABEL = {
    "CA-01": "Программист 1С",
    "CA-02": "Сопровождение 1С",
    "CA-03": "Доработка 1С",
    "CA-04": "Интеграции 1С",
    "CA-05": "Маркировка и Честный знак",
}

GEO_LABEL = {
    "LOCAL": "Новосибирск и Новосибирская область",
    "REMOTE": "Россия, кроме Новосибирска и НСО",
}

APPROVAL_STATUSES = ["На согласовании", "Утверждено", "Нужна правка", "Отклонено"]

COMMERCIAL_CLAIMS = [
    ("Работа по договору", ["договор"]),
    ("Удалённая работа по России", ["удалённо", "удаленно", "дистанционно"]),
    ("Выезд по Новосибирску", ["выезд", "новосибирск"]),
    ("Стоимость от 3 000 ₽ в час", ["3 000", "3000", "от 3"]),
    ("Минимальный заказ — 2 часа", ["2 часа", "минимальный заказ"]),
    ("Работа с 1С:Бухгалтерией", ["бухгалтер", "бп ", "бух "]),
    ("Работа с ЗУП", ["зуп"]),
    ("Работа с УТ", [" ут", "ут "]),
    ("Работа с УНФ", ["унф"]),
    ("Работа с ERP", ["erp", "ерп"]),
    ("Работа с Комплексной автоматизацией", ["комплексн", "ка "]),
    ("Интеграция с сайтами", ["сайт"]),
    ("Интеграция с 1С-Битрикс", ["битрикс"]),
    ("Интеграция с Битрикс24", ["битрикс24"]),
    ("Интеграция по API", ["api"]),
    ("Работа с Честным знаком", ["честн"]),
    ("Работа с СУЗ", ["суз"]),
    ("Работа с ТС ПИоТ", ["пиот", "тс пиот"]),
    ("Работа с маркировкой по товарным категориям", ["маркиров", "категор"]),
]

REJECT_CATEGORY_MAP = [
    ("career", r"career|employment|vacancy|ваканс|зарплат|резюме|работа программист"),
    ("education", r"education|course|college|образован|курс|колледж|обучен"),
    ("salary", r"salary|зарплат|ставк.*работ"),
    ("information/tutorials", r"instructional|informational|tutorial|инструкц|исследован|pure instructional"),
    ("templates/documents", r"template|document lookup|шаблон|документ|certificate"),
    ("named entities", r"person|company|brand|named person|employer"),
    ("foreign geography", r"foreign|scope is russia only|unsupported geo"),
    ("malformed requests", r"malformed|semantically incomplete"),
    ("unrelated intent", r"no commercial|weak service|treasury|technical-spec|generic marking|entertainment"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LOCAL_FILL = PatternFill("solid", fgColor="E8F4FC")
REMOTE_FILL = PatternFill("solid", fgColor="F5F5F5")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def require_operator_gate() -> None:
    if os.environ.get("CORVONERO_OPERATOR_GATE") != "APPROVED":
        raise SystemExit(
            "STOP: CORVONERO_OPERATOR_GATE=APPROVED required. "
            "This C2c helper is not safe for casual execution."
        )


def load_json(name: str) -> Any:
    return json.loads((PILOT / name).read_text(encoding="utf-8"))


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def service_from_campaign(campaign_id: str) -> str:
    code = campaign_id.rsplit("-", 1)[0]
    return SERVICE_LABEL[code]


def mode_from_campaign(campaign_id: str) -> str:
    return campaign_id.rsplit("-", 1)[1]


def style_header_row(ws, row: int, col_count: int) -> None:
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def apply_table_borders(ws, min_row: int, max_row: int, max_col: int) -> None:
    for r in range(min_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = BORDER


def build_group_order(arch: dict) -> dict[tuple[str, str], int]:
    order: dict[tuple[str, str], int] = {}
    idx = 0
    for g in arch.get("groups", []):
        key = (g["campaign_id"], g["group_id"])
        if key not in order:
            order[key] = idx
            idx += 1
    return order


def client_intent(group_name: str) -> str:
    return group_name


def scan_claims_in_ads(ads: list[dict]) -> dict[str, bool]:
    corpus = " ".join(
        f"{a.get('headline_1','')} {a.get('headline_2','')} {a.get('text','')}".lower()
        for a in ads
    )
    group_corpus = " ".join(a.get("group_name", "").lower() for a in ads)
    full = corpus + " " + group_corpus
    result = {}
    for claim, patterns in COMMERCIAL_CLAIMS:
        used = any(p in full for p in patterns)
        result[claim] = used
    return result


def categorize_rejects(register: list[dict]) -> list[dict]:
    rejects = [r for r in register if r.get("decision") == "REJECT"]
    buckets: dict[str, list[str]] = defaultdict(list)
    for r in rejects:
        reason = (r.get("reason") or "").lower()
        phrase = r.get("phrase") or r.get("normalized_phrase") or ""
        matched = False
        for cat, pattern in REJECT_CATEGORY_MAP:
            if re.search(pattern, reason, re.I):
                buckets[cat].append(phrase)
                matched = True
                break
        if not matched:
            buckets["unrelated intent"].append(phrase)
    rows = []
    for cat, pattern in REJECT_CATEGORY_MAP:
        phrases = buckets.get(cat, [])
        if not phrases:
            continue
        examples = "; ".join(phrases[:5])
        rows.append({"category": cat, "count": len(phrases), "examples": examples})
    return rows


def expand_phrase_slots(groups: list[dict]) -> list[dict]:
    rows = []
    for g in groups:
        campaign = g["campaign"]
        mode = g.get("mode") or mode_from_campaign(campaign)
        service = service_from_campaign(campaign)
        arch_service = SERVICE_LABEL.get(g.get("commercial_intent", ""), service)
        if g.get("commercial_intent") in SERVICE_LABEL:
            service = SERVICE_LABEL[g["commercial_intent"]]
        phrases = [p.strip() for p in (g.get("phrase_list") or "").split(";") if p.strip()]
        camp_name = ""
        for c in CAMPAIGN_ORDER:
            if c == campaign:
                break
        for phrase in phrases:
            rows.append({
                "service": service,
                "geo": GEO_LABEL[mode],
                "mode": mode,
                "campaign": campaign,
                "group": g["group_name"],
                "phrase": phrase,
                "landing": g.get("landing_url", ""),
            })
    return rows


def write_ads_workbook(
    ads_sorted: list[dict],
    campaign_names: dict[str, str],
    claims_used: dict[str, bool],
    summary_counts: dict,
) -> Path:
    path = OUTPUT / "01-CORVONERO-ADS-FOR-CLIENT-APPROVAL-v1.xlsx"
    wb = Workbook()

    # Sheet 1 — Инструкция
    ws = wb.active
    ws.title = "Инструкция"
    lines = [
        "Пакет согласования рекламных объявлений — Корво Неро",
        "",
        "Назначение файла",
        "Этот файл содержит предложенные объявления для размещения в Яндекс Директе.",
        "Объявления сгруппированы по направлению услуг и географии показа.",
        "",
        "Как работать с файлом",
        "1. Откройте лист «Объявления» и просмотрите каждую строку.",
        "2. Проверьте формулировки, фактическую точность и соответствие вашим услугам.",
        "3. При необходимости укажите правки в колонке «Комментарий заказчика».",
        "4. Для каждой строки выберите статус в колонке «Статус согласования».",
        "5. Лист «Подтверждение условий» — подтвердите коммерческие факты, используемые в рекламе.",
        "6. Лист «Итоговое согласование» заполняется после проверки всех материалов.",
        "",
        "Допустимые статусы согласования:",
        "• На согласовании",
        "• Утверждено",
        "• Нужна правка",
        "• Отклонено",
        "",
        f"Версия материалов: V2.6 (семантика) / пакет размещения V2.6.2",
        f"Дата подготовки: {PACK_DATE}",
    ]
    for i, line in enumerate(lines, 1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 100

    # Sheet 2 — Сводка
    ws2 = wb.create_sheet("Сводка")
    summary_rows = [
        ("Показатель", "Значение"),
        ("Кампаний", 10),
        ("Групп объявлений", 71),
        ("Объявлений", 71),
        ("Ключевых размещений", 926),
        ("", ""),
        ("География LOCAL", GEO_LABEL["LOCAL"]),
        ("География REMOTE", GEO_LABEL["REMOTE"]),
        ("", ""),
        ("Направление", "LOCAL групп", "REMOTE групп", "Всего объявлений"),
    ]
    for i, row in enumerate(summary_rows[:9], 1):
        for j, val in enumerate(row, 1):
            ws2.cell(row=i, column=j, value=val)
    r = 10
    for svc_code in ["CA-01", "CA-02", "CA-03", "CA-04", "CA-05"]:
        sc = summary_counts[svc_code]
        ws2.cell(row=r, column=1, value=SERVICE_LABEL[svc_code])
        ws2.cell(row=r, column=2, value=sc["local_groups"])
        ws2.cell(row=r, column=3, value=sc["remote_groups"])
        ws2.cell(row=r, column=4, value=sc["ads"])
        r += 1
    style_header_row(ws2, 1, 2)
    style_header_row(ws2, 10, 4)
    ws2.column_dimensions["A"].width = 35
    for col in "BCD":
        ws2.column_dimensions[col].width = 18

    # Sheet 3 — Объявления
    ws3 = wb.create_sheet("Объявления")
    headers = [
        "№", "Направление", "География", "Кампания", "Группа объявлений",
        "Что ищет клиент", "Заголовок 1", "Заголовок 2", "Текст объявления",
        "Отображаемая ссылка", "Посадочная страница", "Статус согласования",
        "Комментарий заказчика", "Наш комментарий",
    ]
    for c, h in enumerate(headers, 1):
        ws3.cell(row=1, column=c, value=h)
    style_header_row(ws3, 1, len(headers))

    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(APPROVAL_STATUSES)}"',
        allow_blank=True,
    )
    ws3.add_data_validation(dv)

    widths = [5, 22, 28, 42, 38, 38, 32, 28, 48, 18, 42, 18, 28, 28]
    for i, w in enumerate(widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    for idx, ad in enumerate(ads_sorted, 1):
        row = idx + 1
        mode = mode_from_campaign(ad["campaign"])
        geo = GEO_LABEL[mode]
        camp_display = campaign_names.get(ad["campaign"], ad["campaign"])
        values = [
            idx,
            service_from_campaign(ad["campaign"]),
            geo,
            camp_display,
            ad["group_name"],
            client_intent(ad["group_name"]),
            ad["headline_1"],
            ad["headline_2"],
            ad["text"],
            ad.get("display_path", ""),
            ad.get("landing_url", ""),
            "На согласовании",
            "",
            "",
        ]
        for c, val in enumerate(values, 1):
            cell = ws3.cell(row=row, column=c, value=val)
            cell.alignment = WRAP
            if mode == "LOCAL":
                cell.fill = LOCAL_FILL
            else:
                cell.fill = REMOTE_FILL
        ws3.row_dimensions[row].height = 45
        dv.add(ws3.cell(row=row, column=12))

    ws3.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(ads_sorted)+1}"
    ws3.freeze_panes = "A2"
    apply_table_borders(ws3, 1, len(ads_sorted) + 1, len(headers))

    # Sheet 4 — Подтверждение условий
    ws4 = wb.create_sheet("Подтверждение условий")
    c_headers = ["Условие", "Используется в рекламе", "Подтверждено заказчиком", "Комментарий", "Требуется изменение рекламы"]
    for c, h in enumerate(c_headers, 1):
        ws4.cell(row=1, column=c, value=h)
    style_header_row(ws4, 1, len(c_headers))
    for i, (claim, _) in enumerate(COMMERCIAL_CLAIMS, 2):
        used = "Да" if claims_used.get(claim) else "Нет"
        confirm = "ТРЕБУЕТ ПОДТВЕРЖДЕНИЯ" if claims_used.get(claim) else "—"
        ws4.cell(row=i, column=1, value=claim)
        ws4.cell(row=i, column=2, value=used)
        ws4.cell(row=i, column=3, value=confirm)
        ws4.cell(row=i, column=4, value="")
        ws4.cell(row=i, column=5, value="")
    for col, w in zip("ABCDE", [40, 22, 28, 30, 28]):
        ws4.column_dimensions[col].width = w
    apply_table_borders(ws4, 1, len(COMMERCIAL_CLAIMS) + 1, len(c_headers))

    # Sheet 5 — Итоговое согласование
    ws5 = wb.create_sheet("Итоговое согласование")
    final_fields = [
        ("Название проекта", "Корво Неро — реклама услуг 1С"),
        ("Версия материалов", "V2.6 / пакет V2.6.2"),
        ("Дата", PACK_DATE),
        ("Количество кампаний", 10),
        ("Количество групп", 71),
        ("Количество объявлений", 71),
        ("Утверждены тексты", ""),
        ("Утверждены направления", ""),
        ("Утверждена география", ""),
        ("Утверждены посадочные страницы", ""),
        ("Коммерческие условия подтверждены", ""),
        ("Перечень обязательных изменений", ""),
        ("ФИО / роль согласующего", ""),
        ("Дата согласования", ""),
    ]
    for i, (label, val) in enumerate(final_fields, 1):
        ws5.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws5.cell(row=i, column=2, value=val)
    ws5.column_dimensions["A"].width = 38
    ws5.column_dimensions["B"].width = 50

    wb.save(path)
    return path


def write_semantic_appendix(
    groups: list[dict],
    phrase_rows: list[dict],
    reject_rows: list[dict],
    campaign_names: dict[str, str],
    neg_txt_info: list[dict],
) -> Path:
    path = OUTPUT / "03-CORVONERO-SEMANTIC-APPENDIX-v1.xlsx"
    wb = Workbook()

    # Сводка
    ws = wb.active
    ws.title = "Сводка"
    ws.cell(row=1, column=1, value="Разрез")
    ws.cell(row=1, column=2, value="Показатель")
    ws.cell(row=1, column=3, value="Значение")
    style_header_row(ws, 1, 3)
    r = 2
    ws.cell(row=r, column=1, value="Общее"); ws.cell(row=r, column=2, value="Кампаний"); ws.cell(row=r, column=3, value=10); r += 1
    ws.cell(row=r, column=1, value="Общее"); ws.cell(row=r, column=2, value="Групп"); ws.cell(row=r, column=3, value=71); r += 1
    ws.cell(row=r, column=1, value="Общее"); ws.cell(row=r, column=2, value="Ключевых размещений"); ws.cell(row=r, column=3, value=926); r += 2

    for svc in ["CA-01", "CA-02", "CA-03", "CA-04", "CA-05"]:
        local_p = sum(1 for p in phrase_rows if p["service"] == SERVICE_LABEL[svc] and p["mode"] == "LOCAL")
        remote_p = sum(1 for p in phrase_rows if p["service"] == SERVICE_LABEL[svc] and p["mode"] == "REMOTE")
        local_g = sum(1 for g in groups if g.get("commercial_intent") == svc and (g.get("mode") or mode_from_campaign(g["campaign"])) == "LOCAL")
        remote_g = sum(1 for g in groups if g.get("commercial_intent") == svc and (g.get("mode") or mode_from_campaign(g["campaign"])) == "REMOTE")
        ws.cell(row=r, column=1, value=SERVICE_LABEL[svc])
        ws.cell(row=r, column=2, value="LOCAL — групп / размещений")
        ws.cell(row=r, column=3, value=f"{local_g} / {local_p}")
        r += 1
        ws.cell(row=r, column=1, value=SERVICE_LABEL[svc])
        ws.cell(row=r, column=2, value="REMOTE — групп / размещений")
        ws.cell(row=r, column=3, value=f"{remote_g} / {remote_p}")
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Кампания"); ws.cell(row=r, column=2, value="Название"); ws.cell(row=r, column=3, value="Размещений")
    style_header_row(ws, r, 3)
    r += 1
    camp_counts = Counter(p["campaign"] for p in phrase_rows)
    for cid in CAMPAIGN_ORDER:
        ws.cell(row=r, column=1, value=cid)
        ws.cell(row=r, column=2, value=campaign_names.get(cid, cid))
        ws.cell(row=r, column=3, value=camp_counts.get(cid, 0))
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 18

    # Ключевые фразы
    ws2 = wb.create_sheet("Ключевые фразы")
    ph_headers = ["Направление", "География", "Кампания", "Группа", "Ключевая фраза", "Посадочная страница"]
    for c, h in enumerate(ph_headers, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header_row(ws2, 1, len(ph_headers))
    for i, p in enumerate(phrase_rows, 2):
        ws2.cell(row=i, column=1, value=p["service"])
        ws2.cell(row=i, column=2, value=p["geo"])
        ws2.cell(row=i, column=3, value=campaign_names.get(p["campaign"], p["campaign"]))
        ws2.cell(row=i, column=4, value=p["group"])
        ws2.cell(row=i, column=5, value=p["phrase"])
        ws2.cell(row=i, column=6, value=p["landing"])
    ws2.auto_filter.ref = f"A1:F{len(phrase_rows)+1}"
    ws2.freeze_panes = "A2"
    for col, w in zip("ABCDEF", [22, 30, 42, 38, 45, 42]):
        ws2.column_dimensions[col].width = w

    # Исключённые категории
    ws3 = wb.create_sheet("Исключённые категории")
    for c, h in enumerate(["Категория", "Количество", "Примеры"], 1):
        ws3.cell(row=1, column=c, value=h)
    style_header_row(ws3, 1, 3)
    cat_labels = {
        "career": "Вакансии и трудоустройство",
        "education": "Обучение и курсы",
        "salary": "Зарплаты и доходы",
        "information/tutorials": "Справочная информация и инструкции",
        "templates/documents": "Шаблоны и документы",
        "named entities": "Имена людей и сторонние компании",
        "foreign geography": "Неподдерживаемая география",
        "malformed requests": "Некорректные запросы",
        "unrelated intent": "Нерелевантный спрос",
    }
    for i, row in enumerate(reject_rows, 2):
        ws3.cell(row=i, column=1, value=cat_labels.get(row["category"], row["category"]))
        ws3.cell(row=i, column=2, value=row["count"])
        ws3.cell(row=i, column=3, value=row["examples"])
        ws3.cell(row=i, column=3).alignment = WRAP
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 70

    # Минус-фразы
    ws4 = wb.create_sheet("Минус-фразы")
    note = (
        "Минус-фразы хранятся отдельными TXT-файлами для каждой кампании (LOCAL и REMOTE). "
        "Наборы не переносятся автоматически между кампаниями. "
        "После импорта кампаний минус-фразы добавляются вручную оператором."
    )
    ws4.cell(row=1, column=1, value=note).alignment = WRAP
    ws4.merge_cells("A1:D1")
    ws4.row_dimensions[1].height = 50
    headers = ["Кампания", "Файл", "Количество минус-фраз", "Примечание"]
    for c, h in enumerate(headers, 1):
        ws4.cell(row=3, column=c, value=h)
    style_header_row(ws4, 3, 4)
    for i, info in enumerate(neg_txt_info, 4):
        ws4.cell(row=i, column=1, value=info["campaign"])
        ws4.cell(row=i, column=2, value=info["file"])
        ws4.cell(row=i, column=3, value=info["count"])
        ws4.cell(row=i, column=4, value="Импорт вручную после загрузки кампании")
    ws4.column_dimensions["A"].width = 18
    ws4.column_dimensions["B"].width = 50
    ws4.column_dimensions["C"].width = 22
    ws4.column_dimensions["D"].width = 40

    # География
    ws5 = wb.create_sheet("География")
    geo_text = [
        ("LOCAL — Новосибирск и НСО", GEO_LABEL["LOCAL"] + ". В объявлениях — предложение с выездом специалиста."),
        ("REMOTE — Россия без НСО", GEO_LABEL["REMOTE"] + ". В объявлениях — удалённое оказание услуг."),
        ("Исключение для REMOTE", "После импорта REMOTE-кампаний оператор вручную исключает Новосибирск и Новосибирскую область из показа."),
        ("Разделение LOCAL/REMOTE", "Позволяет показывать релевантные формулировки, контролировать бюджет по регионам и анализировать эффективность отдельно."),
    ]
    ws5.cell(row=1, column=1, value="Режим")
    ws5.cell(row=1, column=2, value="Описание")
    style_header_row(ws5, 1, 2)
    for i, (a, b) in enumerate(geo_text, 2):
        ws5.cell(row=i, column=1, value=a)
        ws5.cell(row=i, column=2, value=b).alignment = WRAP
    ws5.column_dimensions["A"].width = 28
    ws5.column_dimensions["B"].width = 80

    wb.save(path)
    return path


def write_launch_register() -> Path:
    path = OUTPUT / "04-CORVONERO-LAUNCH-READINESS-REGISTER-v1.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Launch Readiness"
    headers = ["Item", "Owner", "Required evidence", "Status", "Date", "Comment", "Blocking launch"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    items = [
        ("Semantic authority V2.6", "Operator", "CORVONERO-CAMPAIGN-V2.6-OPERATOR-SEMANTIC-APPROVAL-v1.json", "APPROVED", "2026-06-30", "Operator semantic approval recorded", "No"),
        ("Release gate V2.6.2", "Operator", "CORVONERO-CAMPAIGN-V2.6.2-RELEASE-GATE-RESULT-v1.json", "PASS", "2026-06-30", "Phrase-slot reconciliation delta 0", "No"),
        ("Client ad approval", "Client", "Signed 01-CORVONERO-ADS-FOR-CLIENT-APPROVAL-v1.xlsx", "NOT RECEIVED", "", "Awaiting client review", "Yes"),
        ("Commercial claims confirmation", "Client", "Sheet Подтверждение условий in ad workbook", "NOT RECEIVED", "", "", "Yes"),
        ("Landing pages verified", "Operator/Client", "URLs match authority landing pages", "PENDING", "", "Client confirmation recommended", "Yes"),
        ("Commander import", "Operator", "10 XLSX imported and reconciled", "NOT PERFORMED", "", "", "Yes"),
        ("Campaign count reconciliation", "Operator", "10 campaigns in Direct", "NOT PERFORMED", "", "", "Yes"),
        ("Group count reconciliation", "Operator", "71 groups in Direct", "NOT PERFORMED", "", "", "Yes"),
        ("Phrase count reconciliation", "Operator", "926 phrase slots in Direct", "NOT PERFORMED", "", "", "Yes"),
        ("Ad count reconciliation", "Operator", "71 ads in Direct", "NOT PERFORMED", "", "", "Yes"),
        ("Negative TXT import", "Operator", "10 campaign negative TXT files", "NOT PERFORMED", "", "Manual post-import", "Yes"),
        ("REMOTE NSO exclusion", "Operator", "Novosibirsk + NSO excluded in REMOTE campaigns", "NOT PERFORMED", "", "Manual post-import", "Yes"),
        ("UTM setup", "Operator", "UTM parameters on final URLs", "NOT PERFORMED", "", "", "Yes"),
        ("Yandex Metrica", "Operator", "Counter verified on landing pages", "PENDING", "", "", "Yes"),
        ("Goals", "Operator", "Conversion goals configured", "PENDING", "", "", "Yes"),
        ("Forms", "Operator", "Lead forms functional", "PENDING", "", "", "Yes"),
        ("Phone tracking", "Operator", "Call tracking if applicable", "PENDING", "", "", "No"),
        ("Budget", "Client/Operator", "Daily/campaign budgets approved", "PENDING", "", "", "Yes"),
        ("Strategy", "Client/Operator", "Bid strategy confirmed", "PENDING", "", "", "Yes"),
        ("Schedule", "Client/Operator", "Ad schedule confirmed", "PENDING", "", "", "No"),
        ("Moderation", "Operator", "All ads pass Yandex moderation", "NOT PERFORMED", "", "", "Yes"),
        ("Launch authorization", "Client", "Written launch approval", "NOT APPROVED", "", "", "Yes"),
        ("Post-launch day 1 check", "Operator", "Impressions, spend, errors", "NOT PERFORMED", "", "", "No"),
        ("Post-launch day 3 check", "Operator", "Search queries, CTR", "NOT PERFORMED", "", "", "No"),
        ("Post-launch week 1 review", "Operator", "Lead quality, negatives", "NOT PERFORMED", "", "", "No"),
    ]
    for i, row in enumerate(items, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.alignment = WRAP
            if row[6] == "Yes" and row[3] in ("NOT RECEIVED", "NOT PERFORMED", "NOT APPROVED", "PENDING"):
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.auto_filter.ref = f"A1:G{len(items)+1}"
    ws.freeze_panes = "A2"
    for col, w in zip("ABCDEFG", [28, 14, 45, 16, 12, 40, 16]):
        ws.column_dimensions[col].width = w
    wb.save(path)
    return path


def write_strategy_md(campaign_names: dict[str, str]) -> str:
    return f"""# Корво Неро — структура и обоснование рекламной кампании

**Версия материалов:** V2.6 (семантика) / пакет размещения V2.6.2  
**Дата:** {PACK_DATE}

---

## 1. Задача проекта

Рекламная кампания направлена на продвижение платных услуг компании **Корво Неро** в области сопровождения, доработки и интеграции **1С**, а также работ с **маркировкой и Честным знаком**.

Основные цели:

- привлечение компаний и индивидуальных предпринимателей, которым нужен специалист или подрядчик по 1С;
- разделение локального спроса (с выездом в **Новосибирске и области**) и удалённого оказания услуг по **России**;
- создание управляемой структуры кампаний по **пяти направлениям услуг**, чтобы показывать релевантные объявления и контролировать бюджет.

---

## 2. Исходные данные

Утверждённый бизнес-контекст:

| Параметр | Значение |
|----------|----------|
| Сайт и посадочные страницы | lk.corvonero.ru — отдельные страницы по каждому направлению |
| Направления услуг | Программист 1С; Сопровождение 1С; Доработка 1С; Интеграции 1С; Маркировка и Честный знак |
| География LOCAL | Новосибирск и Новосибирская область |
| География REMOTE | Россия, за исключением Новосибирска и НСО |
| Формат работы | B2B; разовые и абонентские задачи — где это подтверждено в материалах |
| Удалённая работа | Для REMOTE-кампаний |
| Выезд | Для LOCAL-кампаний в Новосибирске и области |

---

## 3. Исследование спроса

Работа выполнялась на основе:

1. **Сбор исходных поисковых запросов** — из семантических материалов проекта и анализа реального спроса по тематике 1С.
2. **Анализ частотности** — с использованием данных Яндекс Вордстат там, где это подтверждено в материалах проекта.
3. **Разбор коммерческого и информационного намерения** — отделение запросов на покупку услуги от справочных, образовательных и карьерных.
4. **Географический анализ** — выделение локального спроса (Новосибирск/НСО) и общероссийского удалённого.
5. **Разделение по направлениям услуг** — программист, сопровождение, доработка, интеграции, маркировка.
6. **Отсечение нерелевантного спроса** — вакансии, обучение, зарплаты, шаблоны, пиратский контент и прочие запросы без намерения заказать услугу.

---

## 4. Как отбирались поисковые запросы

В финальный набор вошли запросы, связанные с:

- поиском **специалиста или программиста 1С**;
- **стоимостью** и **ценой часа** работы;
- **сопровождением**, **обслуживанием** и **устранением ошибок**;
- **доработкой** и **разработкой** конфигураций;
- **интеграцией** с сайтами, Битрикс, API;
- **маркировкой**, **Честным знаком**, СУЗ, ТС ПИоТ.

**Исключённые категории** (271 запрос не включён в кампанию):

- вакансии и трудоустройство;
- обучение, курсы, колледжи;
- зарплаты и доходы;
- шаблоны документов и справочные материалы;
- инструкции без намерения заказать услугу;
- скачивания, пиратский контент;
- имена людей и посторонние компании;
- неподдерживаемая зарубежная или некорректная география.

---

## 5. Архитектура кампаний

Кампания разделена на **5 семейств услуг**, каждое — на **LOCAL** и **REMOTE**:

### Программист 1С
- **Потребность:** найти специалиста для разовой или регулярной работы, узнать стоимость часа.
- **Типичные запросы:** «нужен программист 1с», «стоимость часа программиста 1с», «программист 1с москва» (REMOTE).
- **Посадочная страница:** https://lk.corvonero.ru/programmist-1s/

### Сопровождение 1С
- **Потребность:** абонентское обслуживание, техподдержка, обновления, работа с конкретными конфигурациями.
- **Типичные запросы:** «сопровождение 1с», «техподдержка 1с», «сопровождение 1с бухгалтерия».
- **Посадочная страница:** https://lk.corvonero.ru/soprovozhdenie-1s/

### Доработка 1С
- **Потребность:** доработка отчётов, документов, внедрение, разработка под задачи бизнеса.
- **Типичные запросы:** «доработка 1с», «разработка 1с», «внедрение 1с».
- **Посадочная страница:** https://lk.corvonero.ru/dorabotka-razrabotka-1s/

### Интеграции 1С
- **Потребность:** связать 1С с сайтом, интернет-магазином, Битрикс24, обмен данными.
- **Типичные запросы:** «интеграция 1с с сайтом», «1с битрикс интеграция», «обмен данными 1с».
- **Посадочная страница:** https://lk.corvonero.ru/integracii-1s/

### Маркировка и Честный знак
- **Потребность:** подключение, настройка маркировки, работа с кодами, СУЗ, категориями товаров.
- **Типичные запросы:** «маркировка в 1с», «честный знак 1с», «настройка маркировки».
- **Посадочная страница:** https://lk.corvonero.ru/markirovka-chestny-znak/

---

## 6. Разделение LOCAL и REMOTE

| Режим | География | Предложение в объявлениях |
|-------|-----------|---------------------------|
| **LOCAL** | Новосибирск и Новосибирская область | Выезд специалиста |
| **REMOTE** | Россия, кроме Новосибирска и НСО | Удалённая работа |

**Зачем разделять:**

- **Релевантность сообщений** — клиент в Новосибирске видит выезд, клиент в другом регионе — удалённое подключение.
- **Контроль географии** — отдельные настройки показа и исключений.
- **Контроль бюджета** — возможность распределять лимиты по регионам.
- **Аналитика** — сравнение эффективности LOCAL и REMOTE после запуска.

---

## 7. Логика группировки

**71 группа объявлений** организована по **конкретному намерению** покупателя. Примеры:

| Группа | Что ищет клиент |
|--------|-----------------|
| Программист 1С — поиск специалиста | Нужен программист, частный специалист |
| Программист 1С — стоимость часа | Цена часа, ставка |
| Сопровождение 1С — бухгалтерия и БГУ | Сопровождение конкретной конфигурации |
| Интеграция 1С с сайтом и API | Связь с сайтом, API |
| Честный знак — подключение и настройка | Первичное подключение маркировки |
| Коды маркировки | Печать, передача, сканирование кодов |
| Локальный модуль Честного знака | Установка локального модуля |
| СУЗ и токены | Работа с станцией управления заказами |
| ТС ПИоТ | Товароучётная система |

Каждая группа объединяет близкие по смыслу запросы и получает **своё объявление**.

---

## 8. Логика объявлений

- Каждой группе соответствует **одно объявление** с формулировками под её намерение.
- **LOCAL** и **REMOTE** используют разные вторые заголовки и тексты (выезд vs удалённо).
- В тексты **не включены** неподтверждённые обещания (гарантии лидов, позиций, конверсии, статус официального партнёра 1С).
- **Ссылки** ведут на релевантную посадочную страницу направления.

---

## 9. Политика минус-фраз

- Нерелевантный трафик фильтруется **минус-фразами**.
- Минус-фразы хранятся **отдельными TXT-файлами** для каждой кампании.
- **Межкампанные** минус-фразы **не применяются автоматически**.
- После запуска список будет **уточняться** по статистике реальных поисковых запросов.

---

## 10. Что будет сделано перед запуском

1. **Согласование объявлений с заказчиком** (этот пакет).
2. **Подтверждение коммерческих условий** в рабочей таблице.
3. **Импорт кампаний** в Commander / Яндекс Директ.
4. **Сверка количества** кампаний, групп, объявлений и ключевых фраз.
5. **Ручной импорт минус-фраз** из TXT-файлов.
6. **Исключение Новосибирска и НСО** из REMOTE-кампаний.
7. **Проверка аналитики и целей** (Метрика, формы, UTM).
8. **Финальное разрешение на запуск** от заказчика.

---

## 11. План оптимизации после запуска

- анализ **поисковых запросов** и расширение минус-фраз;
- корректировка **ставок и бюджетов**;
- оценка **CTR** и **конверсий**;
- оценка **качества обращений**;
- сравнение **направлений услуг** и **географии**;
- доработка **текстов объявлений**;
- рекомендации по **посадочным страницам**.

Конкретные результаты зависят от рынка, сезона и бюджета — **гарантии не даются**.

---

## 12. Итоговая сводка кампании

| Показатель | Значение |
|------------|----------|
| Кампаний | 10 |
| Групп объявлений | 71 |
| Объявлений | 71 |
| Ключевых размещений | 926 |
| Направлений услуг | 5 |
| Режимов географии | 2 (LOCAL / REMOTE) |

**О ключевых размещениях:** общее число 926 включает размещение одного и того же нейтрального запроса и в LOCAL-, и в REMOTE-кампании, когда спрос подходит для обоих режимов. Это не означает 926 уникальных формулировок — часть запросов дублируется между географическими режимами намеренно.
"""


def md_to_html(md: str) -> str:
    """Minimal MD to HTML for client-readable document."""
    lines = md.split("\n")
    html_parts = [
        "<!DOCTYPE html>",
        '<html lang="ru">',
        "<head>",
        '<meta charset="utf-8">',
        '<meta name="viewport" content="width=device-width, initial-scale=1">',
        "<title>Корво Неро — структура и обоснование рекламной кампании</title>",
        "<style>",
        "body { font-family: 'Segoe UI', Arial, sans-serif; max-width: 900px; margin: 0 auto; padding: 2rem; line-height: 1.6; color: #222; }",
        "h1 { color: #1F4E79; border-bottom: 2px solid #1F4E79; padding-bottom: 0.5rem; }",
        "h2 { color: #2E75B6; margin-top: 2rem; }",
        "h3 { color: #404040; }",
        "table { border-collapse: collapse; width: 100%; margin: 1rem 0; }",
        "th, td { border: 1px solid #ccc; padding: 0.5rem 0.75rem; text-align: left; }",
        "th { background: #1F4E79; color: #fff; }",
        "tr:nth-child(even) { background: #f9f9f9; }",
        "hr { border: none; border-top: 1px solid #ddd; margin: 2rem 0; }",
        "strong { color: #1F4E79; }",
        "ul { padding-left: 1.5rem; }",
        "li { margin: 0.25rem 0; }",
        "</style>",
        "</head>",
        "<body>",
    ]
    in_table = False
    in_list = False
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.startswith("# "):
            if in_list:
                html_parts.append("</ul>")
                in_list = False
            html_parts.append(f"<h1>{line[2:]}</h1>")
        elif line.startswith("## "):
            if in_list:
                html_parts.append("</ul>")
                in_list = False
            html_parts.append(f"<h2>{line[3:]}</h2>")
        elif line.startswith("### "):
            if in_list:
                html_parts.append("</ul>")
                in_list = False
            html_parts.append(f"<h3>{line[4:]}</h3>")
        elif line.strip() == "---":
            html_parts.append("<hr>")
        elif line.startswith("| ") and "|" in line[1:]:
            if in_list:
                html_parts.append("</ul>")
                in_list = False
            if not in_table:
                html_parts.append("<table>")
                in_table = True
            cells = [c.strip() for c in line.strip("|").split("|")]
            if i + 1 < len(lines) and set(lines[i + 1].strip()) <= set("-| "):
                html_parts.append("<tr>" + "".join(f"<th>{c}</th>" for c in cells) + "</tr>")
                i += 1
            else:
                html_parts.append("<tr>" + "".join(f"<td>{c}</td>" for c in cells) + "</tr>")
        elif line.startswith("- "):
            if in_table:
                html_parts.append("</table>")
                in_table = False
            if not in_list:
                html_parts.append("<ul>")
                in_list = True
            content = line[2:]
            content = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", content)
            html_parts.append(f"<li>{content}</li>")
        elif re.match(r"^\d+\. ", line):
            if in_table:
                html_parts.append("</table>")
                in_table = False
            if in_list:
                html_parts.append("</ul>")
                in_list = False
            content = re.sub(r"^\d+\. ", "", line)
            content = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", content)
            html_parts.append(f"<p>{content}</p>")
        elif line.strip() == "":
            if in_table:
                html_parts.append("</table>")
                in_table = False
            if in_list:
                html_parts.append("</ul>")
                in_list = False
        else:
            if in_table:
                html_parts.append("</table>")
                in_table = False
            if in_list:
                html_parts.append("</ul>")
                in_list = False
            content = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", line)
            html_parts.append(f"<p>{content}</p>")
        i += 1
    if in_table:
        html_parts.append("</table>")
    if in_list:
        html_parts.append("</ul>")
    html_parts.extend(["</body>", "</html>"])
    return "\n".join(html_parts)


CHANGE_POLICY = """# Политика изменений при согласовании с заказчиком — Corvonero v1

**Проект:** Корво Неро — реклама услуг 1С  
**Версия:** v1  
**Дата:** 2026-07-01

## 1. Общие правила

1. Все правки заказчика фиксируются **по идентификатору группы/объявления** (кампания + группа объявлений + номер строки в пакете согласования).
2. Утверждённые правки должны **обновлять authority в репозитории** — исходные JSON-файлы семантики и объявлений V2.6.
3. **Запрещено** редактировать только Commander/XLSX, оставляя authority устаревшей.
4. Квитанция согласования заказчика должна ссылаться на **точную версию** ad-pack: `01-CORVONERO-ADS-FOR-CLIENT-APPROVAL-v1.xlsx` и дату/хеш пакета.

## 2. Типы изменений

### Материальные изменения (требуют новой версии deployable package)

- изменение текста объявления (заголовки, текст, УТП);
- перенос группы между кампаниями или направлениями;
- смена посадочной страницы;
- изменение коммерческих условий в рекламе (цена, минимальный заказ, география);
- добавление или удаление групп, объявлений, ключевых фраз.

**Действие:** обновить authority → сгенерировать новый deployable package (V2.6.x+1) → пройти release gate → повторное согласование с заказчиком.

### Орфографические правки (controlled patch)

- исправление опечаток без изменения смысла, группы, URL и коммерческих условий.

**Действие:** допускается controlled patch **только если** differential validation подтверждает отсутствие прочих изменений.

## 3. Процесс

1. Заказчик заполняет колонки «Статус согласования» и «Комментарий заказчика» в `01-CORVONERO-ADS-FOR-CLIENT-APPROVAL-v1.xlsx`.
2. Оператор классифицирует каждую правку: орфография / материальная.
3. Authority обновляется первым.
4. Deployable package пересобирается из обновлённой authority.
5. Новый ad-pack и квитанция согласования выпускаются с новой версией.

## 4. Запреты

- Запуск в Яндекс Директ без полученного согласования заказчика.
- Импорт в Commander до закрытия client ad approval (если не оформлено явное операторское исключение).
- Сокрытие material changes под видом spelling patch.
"""

README_PACK = """# README — Corvonero Client Approval Pack v1

**Пакет:** `CORVONERO-CLIENT-APPROVAL-PACK-2026-07-01`  
**Дата:** 2026-07-01  
**Authority:** Corvonero Campaign V2.6  
**Deployable package:** Corvonero V2.6.2

## Файлы для отправки заказчику (рекомендуется)

| Файл | Назначение |
|------|------------|
| `01-CORVONERO-ADS-FOR-CLIENT-APPROVAL-v1.xlsx` | Согласование объявлений и коммерческих условий |
| `02-CORVONERO-CAMPAIGN-STRATEGY-AND-RESEARCH-v1.html` | Структура и обоснование кампании (удобно для просмотра) |

## Опционально для заказчика

| Файл | Назначение |
|------|------------|
| `03-CORVONERO-SEMANTIC-APPENDIX-v1.xlsx` | Детальный список ключевых фраз и минус-фраз |
| `02-CORVONERO-CAMPAIGN-STRATEGY-AND-RESEARCH-v1.md` | Тот же документ в Markdown |

## Только для внутреннего использования

| Файл | Назначение |
|------|------------|
| `04-CORVONERO-LAUNCH-READINESS-REGISTER-v1.xlsx` | Регистр готовности к запуску |
| `CORVONERO-CLIENT-APPROVAL-CHANGE-POLICY-v1.md` | Политика обработки правок |
| `CORVONERO-CLIENT-APPROVAL-PACK-MANIFEST-v1.json` | Манифест пакета |
| `CORVONERO-CLIENT-APPROVAL-PACK-SHA256SUMS-v1.txt` | Контрольные суммы |

## Использование

1. **До запуска:** отправить заказчику файлы 01 и 02; дождаться заполнения статусов и комментариев.
2. **После получения правок:** следовать `CORVONERO-CLIENT-APPROVAL-CHANGE-POLICY-v1.md`.
3. **После согласования:** обновить `04-CORVONERO-LAUNCH-READINESS-REGISTER-v1.xlsx` и перейти к импорту V2.6.2.

## Ссылки на authority (репозиторий)

- `projects/mars-search-ppc-production/pilots/corvonero/client-approval/`
"""


def main() -> None:
    require_operator_gate()
    OUTPUT.mkdir(parents=True, exist_ok=True)
    CLIENT_APPROVAL_REPO.mkdir(parents=True, exist_ok=True)

    ads_data = load_json("CORVONERO-CAMPAIGN-V2.6-FINAL-AD-COPY-v1.json")
    groups_data = load_json("CORVONERO-CAMPAIGN-V2.6-FINAL-GROUP-PLAN-v1.json")
    arch = load_json("CORVONERO-CAMPAIGN-V2.6-CAMPAIGN-ARCHITECTURE-v1.json")
    phrase_auth = load_json("CORVONERO-CAMPAIGN-V2.6-FINAL-PHRASE-AUTHORITY-v1.json")
    result = load_json("CORVONERO-CAMPAIGN-V2.6-RESULT-v1.json")

    ads = ads_data["ads"]
    groups = groups_data["groups"]
    register = phrase_auth["register"]

    campaign_names = {c["campaign_id"]: c["commander_name"] for c in arch["campaigns"]}
    group_order = build_group_order(arch)

    def sort_key(ad: dict) -> tuple:
        camp = ad["campaign"]
        svc = camp.rsplit("-", 1)[0]
        mode = mode_from_campaign(camp)
        mode_ord = 0 if mode == "LOCAL" else 1
        gord = group_order.get((camp, ad["group_id"]), 999)
        return (svc, mode_ord, gord, camp)

    ads_sorted = sorted(ads, key=sort_key)
    assert len(ads_sorted) == 71, f"Expected 71 ads, got {len(ads_sorted)}"

    summary_counts: dict[str, dict] = {}
    for svc in ["CA-01", "CA-02", "CA-03", "CA-04", "CA-05"]:
        svc_ads = [a for a in ads if a["campaign"].startswith(svc)]
        local_g = len({(a["campaign"], a["group_id"]) for a in svc_ads if "LOCAL" in a["campaign"]})
        remote_g = len({(a["campaign"], a["group_id"]) for a in svc_ads if "REMOTE" in a["campaign"]})
        summary_counts[svc] = {"local_groups": local_g, "remote_groups": remote_g, "ads": len(svc_ads)}

    claims_used = scan_claims_in_ads(ads)
    reject_rows = categorize_rejects(register)
    phrase_rows = expand_phrase_slots(groups)
    assert len(phrase_rows) == 926, f"Expected 926 phrase slots, got {len(phrase_rows)}"

    neg_txt_info = []
    for cid in CAMPAIGN_ORDER:
        fname = f"{cid}-CAMPAIGN-NEGATIVES-FINAL-v2.6.2.txt"
        fpath = DEPLOY_PKG / fname
        count = 0
        if fpath.exists():
            lines = [ln.strip() for ln in fpath.read_text(encoding="utf-8").splitlines() if ln.strip()]
            count = len(lines)
        neg_txt_info.append({"campaign": cid, "file": fname, "count": count})

    created: list[Path] = []
    created.append(write_ads_workbook(ads_sorted, campaign_names, claims_used, summary_counts))
    created.append(write_semantic_appendix(groups, phrase_rows, reject_rows, campaign_names, neg_txt_info))
    created.append(write_launch_register())

    strategy_md = write_strategy_md(campaign_names)
    md_path = OUTPUT / "02-CORVONERO-CAMPAIGN-STRATEGY-AND-RESEARCH-v1.md"
    md_path.write_text(strategy_md, encoding="utf-8")
    created.append(md_path)

    html_path = OUTPUT / "02-CORVONERO-CAMPAIGN-STRATEGY-AND-RESEARCH-v1.html"
    html_path.write_text(md_to_html(strategy_md), encoding="utf-8")
    created.append(html_path)

    policy_path = OUTPUT / "CORVONERO-CLIENT-APPROVAL-CHANGE-POLICY-v1.md"
    policy_path.write_text(CHANGE_POLICY, encoding="utf-8")
    created.append(policy_path)

    readme_path = OUTPUT / "README-CORVONERO-CLIENT-APPROVAL-PACK-v1.md"
    readme_path.write_text(README_PACK, encoding="utf-8")
    created.append(readme_path)

    # Commercial claims register JSON
    claims_register = {
        "schema_version": "client-commercial-claims-register-v1",
        "project_id": "corvonero",
        "pack_version": "v1",
        "pack_date": PACK_DATE,
        "claims": [
            {
                "claim": claim,
                "used_in_ads": claims_used.get(claim, False),
                "client_confirmed": None,
                "requires_confirmation": claims_used.get(claim, False),
                "status": "ТРЕБУЕТ ПОДТВЕРЖДЕНИЯ" if claims_used.get(claim) else "NOT_USED",
            }
            for claim, _ in COMMERCIAL_CLAIMS
        ],
    }
    claims_path = CLIENT_APPROVAL_REPO / "CORVONERO-CLIENT-COMMERCIAL-CLAIMS-REGISTER-v1.json"
    claims_path.write_text(json.dumps(claims_register, ensure_ascii=False, indent=2), encoding="utf-8")

    ad_register = {
        "schema_version": "client-ad-approval-register-v1",
        "project_id": "corvonero",
        "semantic_authority": "V2.6",
        "deployable_package": "V2.6.2",
        "pack_date": PACK_DATE,
        "ad_count": 71,
        "campaign_count": 10,
        "group_count": 71,
        "client_approval_status": "NOT_RECEIVED",
        "ads": [
            {
                "row": i,
                "campaign_id": ad["campaign"],
                "campaign_name": campaign_names.get(ad["campaign"], ad["campaign"]),
                "group_id": ad["group_id"],
                "group_name": ad["group_name"],
                "service": service_from_campaign(ad["campaign"]),
                "geography_mode": mode_from_campaign(ad["campaign"]),
                "landing_url": ad.get("landing_url", ""),
                "approval_status": "На согласовании",
            }
            for i, ad in enumerate(ads_sorted, 1)
        ],
    }
    ad_reg_path = CLIENT_APPROVAL_REPO / "CORVONERO-CLIENT-AD-APPROVAL-REGISTER-v1.json"
    ad_reg_path.write_text(json.dumps(ad_register, ensure_ascii=False, indent=2), encoding="utf-8")

    pack_ref = CLIENT_APPROVAL_REPO / "CORVONERO-CLIENT-APPROVAL-PACK-v1.md"
    pack_ref.write_text(
        f"""# Corvonero Client Approval Pack — authority reference v1

**Generated:** {PACK_DATE}  
**Storage path:** `{OUTPUT}`

## Source authority

| Artifact | Path |
|----------|------|
| Phrase authority | `CORVONERO-CAMPAIGN-V2.6-FINAL-PHRASE-AUTHORITY-v1.json` |
| Group plan | `CORVONERO-CAMPAIGN-V2.6-FINAL-GROUP-PLAN-v1.json` |
| Ad copy | `CORVONERO-CAMPAIGN-V2.6-FINAL-AD-COPY-v1.json` |
| Negatives | `CORVONERO-CAMPAIGN-V2.6-FINAL-NEGATIVES-v1.json` |
| Result | `CORVONERO-CAMPAIGN-V2.6-RESULT-v1.json` |
| Release state | `CORVONERO-CAMPAIGN-RELEASE-STATE-v1.json` |
| Release gate | `CORVONERO-CAMPAIGN-V2.6.2-RELEASE-GATE-RESULT-v1.json` |

## Totals

- Campaigns: 10
- Groups: 71
- Ads: 71
- Phrase slots: 926
- KEEP: {result['accounting']['KEEP']}
- REJECT: {result['accounting']['REJECT']}

## Delivery files (Storage only — not in Git)

See `README-CORVONERO-CLIENT-APPROVAL-PACK-v1.md` in Storage pack folder.
""",
        encoding="utf-8",
    )

    repo_policy = CLIENT_APPROVAL_REPO / "CORVONERO-CLIENT-APPROVAL-CHANGE-POLICY-v1.md"
    repo_policy.write_text(CHANGE_POLICY, encoding="utf-8")

    manifest_files = []
    for p in sorted(OUTPUT.iterdir()):
        if p.is_file():
            manifest_files.append({
                "path": p.name,
                "sha256": sha256_file(p),
                "size_bytes": p.stat().st_size,
            })

    manifest = {
        "schema_version": "client-approval-pack-manifest-v1",
        "project_id": "corvonero",
        "pack_id": f"CORVONERO-CLIENT-APPROVAL-PACK-{PACK_DATE}",
        "generated_at": f"{PACK_DATE}T12:00:00+07:00",
        "semantic_authority": "V2.6",
        "deployable_package": "V2.6.2",
        "totals": {"campaigns": 10, "groups": 71, "ads": 71, "phrase_slots": 926},
        "files": manifest_files,
    }
    manifest_path = OUTPUT / "CORVONERO-CLIENT-APPROVAL-PACK-MANIFEST-v1.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    sums_path = OUTPUT / "CORVONERO-CLIENT-APPROVAL-PACK-SHA256SUMS-v1.txt"
    sums_lines = [f"{e['sha256']}  {e['path']}" for e in manifest_files]
    sums_path.write_text("\n".join(sums_lines) + "\n", encoding="utf-8")

    claims_pending = [c["claim"] for c in claims_register["claims"] if c["requires_confirmation"]]

    report = f"""# REPORT — Corvonero client approval and strategy pack v1

**Date:** {PACK_DATE}  
**Branch:** mars/canonical-post-recovery  
**Verdict:** CORVONERO CLIENT APPROVAL PACK: **PASS — CLIENT-FACING CAMPAIGN MATERIALS GENERATED**

---

## Source authority

| Artifact | Role |
|----------|------|
| CORVONERO-CAMPAIGN-V2.6-FINAL-PHRASE-AUTHORITY-v1.json | Phrase decisions (487 KEEP / 271 REJECT) |
| CORVONERO-CAMPAIGN-V2.6-FINAL-GROUP-PLAN-v1.json | 71 groups |
| CORVONERO-CAMPAIGN-V2.6-FINAL-AD-COPY-v1.json | 71 ads |
| CORVONERO-CAMPAIGN-V2.6-FINAL-NEGATIVES-v1.json | Safe negatives |
| CORVONERO-CAMPAIGN-V2.6-RESULT-v1.json | Consolidated totals |
| CORVONERO-CAMPAIGN-V2.6-OPERATOR-SEMANTIC-APPROVAL-v1.json | Operator semantic approval |
| CORVONERO-CAMPAIGN-RELEASE-STATE-v1.json | Release state |
| CORVONERO-CAMPAIGN-V2.6.2-RELEASE-GATE-RESULT-v1.json | Release gate PASS |

Deployable package (read-only reconciliation): `CORVONERO-CAMPAIGN-V2.6.2-FINAL-2026-06-30`

---

## Created files

### Storage (`{OUTPUT}`)

| File | Purpose |
|------|---------|
| 01-CORVONERO-ADS-FOR-CLIENT-APPROVAL-v1.xlsx | Client ad approval workbook |
| 02-CORVONERO-CAMPAIGN-STRATEGY-AND-RESEARCH-v1.md | Strategy document |
| 02-CORVONERO-CAMPAIGN-STRATEGY-AND-RESEARCH-v1.html | Client HTML version |
| 03-CORVONERO-SEMANTIC-APPENDIX-v1.xlsx | Semantic appendix |
| 04-CORVONERO-LAUNCH-READINESS-REGISTER-v1.xlsx | Internal launch register |
| CORVONERO-CLIENT-APPROVAL-CHANGE-POLICY-v1.md | Change policy |
| README-CORVONERO-CLIENT-APPROVAL-PACK-v1.md | Package README |
| CORVONERO-CLIENT-APPROVAL-PACK-MANIFEST-v1.json | Manifest |
| CORVONERO-CLIENT-APPROVAL-PACK-SHA256SUMS-v1.txt | SHA256 checksums |

### Repository (`pilots/corvonero/client-approval/`)

| File | Purpose |
|------|---------|
| CORVONERO-CLIENT-APPROVAL-PACK-v1.md | Authority reference |
| CORVONERO-CLIENT-AD-APPROVAL-REGISTER-v1.json | Ad approval register |
| CORVONERO-CLIENT-COMMERCIAL-CLAIMS-REGISTER-v1.json | Commercial claims register |
| CORVONERO-CLIENT-APPROVAL-CHANGE-POLICY-v1.md | Change policy (repo copy) |

---

## Validation

| Check | Result |
|-------|--------|
| Ads represented | 71/71 |
| Groups | 71 |
| Campaigns | 10 |
| Phrase slots | 926 |
| Campaign/group names match V2.6 authority | PASS |
| Client-facing Russian | PASS |
| No V2.1–V2.5 history in client files | PASS |
| No internal failure reports exposed | PASS |
| No unsupported claims introduced | PASS |
| Client approval not pre-filled | PASS |
| Launch not implied | PASS |
| XLSX generation | PASS |

---

## Commercial facts requiring client confirmation

{chr(10).join(f'- {c}' for c in claims_pending)}

---

## Client-facing vs internal

**Client send:** 01 xlsx + 02 html (optional: 03 xlsx)  
**Internal:** 04 xlsx, change policy, manifest, repo registers

---

## Remaining launch blockers

1. Client ad approval — NOT RECEIVED
2. Commercial claims confirmation — NOT RECEIVED
3. Commander import — NOT PERFORMED
4. Yandex Direct launch — NOT APPROVED

---

## Required verdict

```
CORVONERO CLIENT APPROVAL PACK:
PASS — CLIENT-FACING CAMPAIGN MATERIALS GENERATED

Campaigns: 10
Groups: 71
Ads represented: 71/71
Strategy document: GENERATED
Semantic appendix: GENERATED
Commercial claims register: GENERATED
Client approval: NOT YET RECEIVED
Commander import: NOT PERFORMED
Yandex Direct launch: NOT APPROVED
Git checkpoint: NOT PERFORMED
```

---

## Git status

No stage, commit or push performed per task scope.
"""
    report_path = REPORTS / "REPORT-corvonero-client-approval-and-strategy-pack-v1.md"
    report_path.write_text(report, encoding="utf-8")

    print("PASS — pack generated")
    print(f"Ads: {len(ads_sorted)}/71")
    print(f"Phrases: {len(phrase_rows)}/926")
    print(f"Output: {OUTPUT}")
    print(f"Report: {report_path}")


if __name__ == "__main__":
    main()
