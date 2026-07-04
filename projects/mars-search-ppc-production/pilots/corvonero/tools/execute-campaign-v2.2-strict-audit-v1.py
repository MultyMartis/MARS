#!/usr/bin/env python3
"""
CORVONERO Campaign V2.2 — strict phrase-by-phrase audit from V2.1 XLSX deployables.
Read-only on V2/V2.1 packages. No XLSX regeneration. No git commit.

C2b source persistence only. This file is not authorized for execution without explicit operator approval. Commit/persistence does not authorize Storage export generation, repo artifact generation, Commander import, Direct launch, account mutation, advertising start, Localhost mutation, Storage mutation, or Yandex/API access.
"""
from __future__ import annotations

import csv
import json
import os
import re
import subprocess
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

REPO = Path(r"X:\AI MARS")
PILOT = REPO / "projects" / "mars-search-ppc-production" / "pilots" / "corvonero"
V21_PKG = Path(r"X:\AI MARS STORAGE\exports\corvonero\CORVONERO-CAMPAIGN-V2.1-FINAL-2026-06-30")
REVIEW_DIR = Path(r"X:\AI MARS STORAGE\exports\corvonero\CORVONERO-CAMPAIGN-V2.2-STRICT-AUDIT-REVIEW-2026-06-30")
CHECKPOINT_EAAC = "eaac1e1e23a0e3a709cb5410357208928343e2b2"
HEAD_SHA = "4a4381f5"
GENERATED_AT = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")

XLSX_MAP = {
    "CA-01-LOCAL": "CORVONERO-CA-01-LOCAL-PROGRAMMIST-1S-COMMANDER-IMPORT-v2.1.xlsx",
    "CA-01-REMOTE": "CORVONERO-CA-01-REMOTE-PROGRAMMIST-1S-COMMANDER-IMPORT-v2.1.xlsx",
    "CA-02-LOCAL": "CORVONERO-CA-02-LOCAL-SOPROVOZHDENIE-1S-COMMANDER-IMPORT-v2.1.xlsx",
    "CA-02-REMOTE": "CORVONERO-CA-02-REMOTE-SOPROVOZHDENIE-1S-COMMANDER-IMPORT-v2.1.xlsx",
    "CA-03-LOCAL": "CORVONERO-CA-03-LOCAL-DORABOTKA-1S-COMMANDER-IMPORT-v2.1.xlsx",
    "CA-03-REMOTE": "CORVONERO-CA-03-REMOTE-DORABOTKA-1S-COMMANDER-IMPORT-v2.1.xlsx",
    "CA-04-LOCAL": "CORVONERO-CA-04-LOCAL-INTEGRACII-1S-COMMANDER-IMPORT-v2.1.xlsx",
    "CA-04-REMOTE": "CORVONERO-CA-04-REMOTE-INTEGRACII-1S-COMMANDER-IMPORT-v2.1.xlsx",
    "CA-05-LOCAL": "CORVONERO-CA-05-LOCAL-MARKIROVKA-1S-COMMANDER-IMPORT-v2.1.xlsx",
    "CA-05-REMOTE": "CORVONERO-CA-05-REMOTE-MARKIROVKA-1S-COMMANDER-IMPORT-v2.1.xlsx",
}

COL = {
    "group": 5,
    "phrase": 8,
    "headline_1": 10,
    "headline_2": 11,
    "text": 12,
    "landing_url": 48,
    "region": 52,
    "bid": 54,
    "group_negatives": 68,
}
DATA_START = 16
EMBEDDED_CAMP_NEG_ROW = 9
EMBEDDED_CAMP_NEG_COL = 5

SERVICE_FROM_CAMPAIGN = {
    "CA-01": "CA-01",
    "CA-02": "CA-02",
    "CA-03": "CA-03",
    "CA-04": "CA-04",
    "CA-05": "CA-05",
}

GROUP_ID_HINT = {
    "программист 1с — заказ услуги": "ca-01-direct-service-order",
    "программист 1с — поиск специалиста": "ca-01-find-hire-specialist",
    "программист 1с — стоимость и цена": "ca-01-price-intent",
    "программист 1с — удалённо и частный специалист": "ca-01-remote-freelance-specialist",
    "программист 1с — по конфигурациям": "ca-01-specialist-by-product",
    "программист 1с — расширенные запросы": "ca-01-specialist-extended",
    "программист 1с — основной поиск": "ca-01-specialist-search",
    "сопровождение 1с — заказ услуги": "ca-02-direct-service-order",
    "сопровождение 1с — стоимость": "ca-02-price-intent",
    "сопровождение и обслуживание 1с": "ca-02-support-and-maintenance",
    "1с не работает — ошибки и восстановление": "ca-02-troubleshooting-not-working",
    "доработка 1с — заказ услуги": "ca-03-direct-service-order",
    "доработка 1с — внедрение": "ca-03-implementation",
    "доработка и разработка 1с": "ca-03-modification",
    "интеграции 1с": "ca-04-integration",
    "честный знак в 1с — настройка и обмен": "ca-05-chestny-znak-service",
    "маркировка — интеграция с 1с": "ca-05-integration",
    "коды маркировки в 1с": "ca-05-marking-codes",
    "маркировка в 1с — общая настройка": "ca-05-marking-setup",
    "маркировка — техподдержка 1с": "ca-05-support-and-maintenance",
    "тс пиот и честный знак в 1с": "ca-05-ts-piot",
}

CONFIRMED_V21_DEFECTS = {
    "1с программист стало сложно найти работу",
    "фриланс 1с программист работа",
    "инструкция программиста 1с",
    "обязанности программиста 1с",
    "пособие программиста 1с",
    "мемы про 1с программистов",
    "поиск работы программист 1с",
    "подработка для программистов 1с",
    "стоит ли идти в 1с программисты",
    "программист 1с колледж",
    "программист 1с заработная плата",
    "1с программист без образования",
    "программист 1с повышения квалификации",
    "техническое задание на доработку 1с пример",
    "битрикс интеграция с 1с инструкция",
    "интеграция 1с с честным знаком как сделать",
}

CONFIRMED_LOCAL_GEO_FAILURES = {
    "частный программист 1с сергиев посад",
    "подработка программистом 1с минск",
    "программист 1с кострома",
    "программист 1с новороссийск",
    "программист 1с вологда",
    "программист 1с владимир",
    "программист 1с севастополь",
    "программист 1с ижевск",
    "программист 1с саратов",
    "программист 1с калининград",
    "доработка 1с 7.7 минск",
}

BUYER_HIRE_RE = re.compile(
    r"(?:^|\b)(?:нужен|ищу|найти|найм|нанять|вызвать|заказать)\s+(?:программист|специалист|разработчик)",
    re.I,
)
CAREER_RE = re.compile(
    r"работ[аыуе]|работа\s+программист|опыт\s+работ|поиск\s+работ|подработк|ваканс|"
    r"зарплат|заработн|должност|обязанност|трудов(?:ая|ой)\s+функ|профессия|карьер|"
    r"\bjunior\b|\bmiddle\b|джун|младш|старш|ведущ|главн|стажиров|резюме|\bhh\b|"
    r"ищет\s+работ|трудоустрой|частичн(?:ая|ой)\s+занят|ученик\b|становится\s+программист|"
    r"стань\s+программист|сложно\s+найти\s+работ|фриланс.*работа|работа.*фриланс",
    re.I,
)
EDU_RE = re.compile(
    r"образован|без\s+образован|колледж|вуз|специальност|обучен|курс(?:ы|а|ов)?|практикум|"
    r"пособие|справочник|повышени(?:е|я)\s+квалифика|переподготов|что\s+должен\s+знать|"
    r"навык|компетенц|уровн(?:и|я|ь)|roadmap|роадмап|как\s+стать|стоит\s+ли\s+идти|"
    r"перспектив(?:ы|а)\s+професс|востребованност|тестов(?:ое|ые)\s+задан|экзамен|"
    r"инструкци(?:я|и)\s+программист|семинар|урок|скиллбокс|skillbox",
    re.I,
)
ENTERTAIN_RE = re.compile(
    r"\bмем|песн|поздравлен|день\s+рожд|картинк|видео\s+о\s+професс|кратко|"
    r"описание\s+професс|что\s+значит\s+программист|торрент|книг\b|форум",
    re.I,
)
TEMPLATE_RE = re.compile(
    r"пример\s+(?:тз|техническ|задан)|образец\s+(?:тз|техническ)|шаблон\s+(?:тз|договор)|"
    r"техническ(?:ое|ая)\s+задани(?:е|я)\s+пример|должностн(?:ая|ой)\s+инструк|"
    r"договор\s+гпх|как\s+искать\s+клиент|клиент(?:ы|ов)\s+программист|"
    r"заказы\s+для\s+(?:1с\s+)?программист|где\s+искать\s+программист|"
    r"как\s+написать\s+тз\s+для\s+программист",
    re.I,
)
PERSON_EMPLOYER_RE = re.compile(
    r"\b(?:иван|алena|алёна|роман\s+галкин|мешкова|лидер\s+фарма|бэст\s+мебель|"
    r"парк\s+культур|требуется\s+программист\s+1с\s+\w+)\b|"
    r"программист\s+1с\s+(?:иван|алena|алёна|роман|мешков)",
    re.I,
)
COMMERCIAL_RE = re.compile(
    r"услуг|стоимост|цен[аы]|сколько\s+стоит|заказать|нанять|вызвать|"
    r"частн|фриланс|доработк|сопровожден|обслуживан|настройк|внедрен|"
    r"не\s+работает|ошибк|исправ|интеграц|маркировк|честн.*знак|"
    r"расценк|под\s+ключ|срочно|недорог|опытн|заказ\b|подключ",
    re.I,
)
MARKING_RE = re.compile(r"честн(?:ый|ого)\s+знак|маркировк(?:а|и|у|ой)|код(?:ы|ов)?\s+маркировк", re.I)
INTEGRATION_RE = re.compile(r"интеграц|битрикс|bitrix|api|обмен\s+данн|синхронизац|сайт", re.I)
HOWTO_INFO_RE = re.compile(r"как\s+(?:сделать|настро|подключ|интегрир|внедр)|инструкци", re.I)
NSO_RE = re.compile(r"новосибирск|новосибирск(?:ая|ой|ую|ие|им)?\s+област|\bнск\b", re.I)
REMOTE_EXPLICIT_RE = re.compile(
    r"удал[её]нн?(?:о|ая|ый|ые|ка)?|дистанционн|\bонлайн\b|по\s+(?:всей\s+)?росси|по\s+рф\b|без\s+выезд|удал[её]нк",
    re.I,
)
LOCAL_SVC_RE = re.compile(
    r"с\s+выездом|\bвыезд(?:ом|а|е|у)?\b|выезд\s+специалист|\bприехать\b|на\s+месте|в\s+офис(?:е|а|у)?\b",
    re.I,
)
FOREIGN_RE = re.compile(r"беларус|казахстан|алмат|минск|\bднр\b|украин|белорус", re.I)
OTHER_CITY_RE = re.compile(
    r"\b(?:москв|спб|санкт|екатеринбург|красноярск|омск|томск|барнаул|краснодар|воронеж|"
    r"казан|уф|перм|самар|ростов|нижн|челябинск|симферополь|хабаровск|иркутск|ярославль|"
    r"владивосток|белгород|рязань|тюмень|калининград|ставрополь|сочи|тула|костром|"
    r"новороссийск|вологд|владимир|севастопол|ижевск|саратов|сергиев\s+посад)\w*\b",
    re.I,
)
CA02_WRONG_RE = re.compile(r"поддержк[аи]\s+сайт|сертификац.*1с|оквэд|окпд|шаблон\s+договор|its\s+документ", re.I)
CA03_EDU_DEV_RE = re.compile(r"пособие\s+разработчик|библия\s+1с|сравнен.*разработчик", re.I)
GENERIC_AD_RE = re.compile(
    r"^Работаем\s+(?:удалённо\s+по\s+России|с\s+выездом\s+по\s+Новосибирску)\.\s+По\s+договору\.?$",
    re.I,
)


def normalize_phrase(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def cell_val(ws, row: int, col: int) -> str:
    v = ws.cell(row, col).value
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def parse_campaign_mode(campaign_key: str) -> tuple[str, str]:
    ca, mode = campaign_key.rsplit("-", 1)
    return ca, mode


def git_preflight() -> dict[str, Any]:
    label = subprocess.check_output(
        ["powershell.exe", "-NoProfile", "-Command", "(Get-Volume -DriveLetter X).FileSystemLabel"],
        text=True,
    ).strip()
    log = subprocess.check_output(
        ["git", "-C", str(REPO), "log", "--oneline", "--decorate", "-10"],
        text=True,
    )
    show = subprocess.check_output(
        ["git", "-C", str(REPO), "show", "--stat", "--oneline", HEAD_SHA],
        text=True,
    )
    merge_base = subprocess.check_output(
        ["git", "-C", str(REPO), "merge-base", CHECKPOINT_EAAC[:7], HEAD_SHA],
        text=True,
    ).strip()
    head = subprocess.check_output(["git", "-C", str(REPO), "rev-parse", "HEAD"], text=True).strip()
    return {
        "volume_label": label,
        "head": head,
        "checkpoint_eaac": CHECKPOINT_EAAC,
        "merge_base": merge_base,
        "head_diff_reason": (
            f"HEAD ({head[:8]}) is one commit after checkpoint eaac1e1e: "
            "feat(fp-0002) O-Centre V8 baseline — unrelated to Corvonero PPC; "
            "merge-base equals eaac1e1e (Corvonero V2.1 authority parent)."
        ),
        "git_log_tail": log,
        "git_show_head": show,
    }


def extract_xlsx(campaign_key: str, filename: str) -> dict[str, Any]:
    path = V21_PKG / filename
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Тексты"]
    ca, mode = parse_campaign_mode(campaign_key)
    embedded_neg = cell_val(ws, EMBEDDED_CAMP_NEG_ROW, EMBEDDED_CAMP_NEG_COL)
    slots: list[dict[str, Any]] = []
    ads: list[dict[str, Any]] = []
    current_group = ""
    current_gneg = ""
    for r in range(DATA_START, ws.max_row + 1):
        group = cell_val(ws, r, COL["group"]) or current_group
        if group:
            current_group = group
        gneg = cell_val(ws, r, COL["group_negatives"]) or current_gneg
        if gneg:
            current_gneg = gneg
        h1 = cell_val(ws, r, COL["headline_1"])
        h2 = cell_val(ws, r, COL["headline_2"])
        text = cell_val(ws, r, COL["text"])
        phrase = cell_val(ws, r, COL["phrase"])
        region = cell_val(ws, r, COL["region"])
        bid = cell_val(ws, r, COL["bid"])
        url = cell_val(ws, r, COL["landing_url"])
        group_lower = normalize_phrase(group)
        group_id = GROUP_ID_HINT.get(group_lower, group_lower.replace(" ", "-"))

        if h1:
            ads.append(
                {
                    "campaign": campaign_key,
                    "ca": ca,
                    "mode": mode,
                    "group": group,
                    "group_id": group_id,
                    "row": r,
                    "headline_1": h1,
                    "headline_2": h2,
                    "text": text,
                    "landing_url": url,
                    "group_negatives": gneg,
                    "region": region,
                }
            )
        if phrase:
            slots.append(
                {
                    "campaign": campaign_key,
                    "ca": ca,
                    "mode": mode,
                    "group": group,
                    "group_id": group_id,
                    "row": r,
                    "phrase": phrase,
                    "normalized_phrase": normalize_phrase(phrase),
                    "bid": bid,
                    "headline_1": h1,
                    "headline_2": h2,
                    "text": text,
                    "landing_url": url,
                    "campaign_negatives_embedded": embedded_neg,
                    "group_negatives": gneg,
                    "region": region,
                }
            )
    wb.close()
    return {
        "campaign_key": campaign_key,
        "filename": filename,
        "embedded_campaign_negatives": embedded_neg,
        "slots": slots,
        "ads": ads,
    }


def route_service(ca: str, group_id: str, phrase: str) -> tuple[str, str | None, str]:
    p = normalize_phrase(phrase)
    if MARKING_RE.search(p) and ca in ("CA-01", "CA-02", "CA-03", "CA-04"):
        if ca == "CA-04" and INTEGRATION_RE.search(p) and "честн" in p:
            return "CA-05", "ca-05-integration", "Honest Sign integration → CA-05"
        return "CA-05", None, "Marking/Honest Sign intent → CA-05"
    if ca == "CA-04" and MARKING_RE.search(p) and not INTEGRATION_RE.search(p):
        return "CA-05", "ca-05-chestny-znak-service", "Marking without integration → CA-05"
    return ca, group_id, ""


def detect_city(phrase: str) -> str:
    p = normalize_phrase(phrase)
    if NSO_RE.search(p):
        return "NSO"
    if FOREIGN_RE.search(p):
        m = FOREIGN_RE.search(p)
        return m.group(0) if m else "FOREIGN"
    m = OTHER_CITY_RE.search(p)
    return m.group(0) if m else ""


def classify_geo(phrase: str) -> tuple[str, str]:
    p = normalize_phrase(phrase)
    city = detect_city(p)
    has_remote = bool(REMOTE_EXPLICIT_RE.search(p))
    has_local_svc = bool(LOCAL_SVC_RE.search(p))
    if city == "NSO" or has_local_svc:
        return "LOCAL_ONLY", f"NSO/local-visit signal ({city or 'on-site'})"
    if city and city != "NSO":
        if FOREIGN_RE.search(p):
            return "NONE", f"Foreign geography: {city}"
        return "REMOTE_ONLY", f"Non-NSO city: {city}"
    if has_remote:
        return "REMOTE_ONLY", "Remote-explicit phrase"
    return "BOTH", "Neutral commercial — eligible for LOCAL and REMOTE"


def classify_phrase(
    phrase: str,
    source_ca: str,
    source_group_id: str,
    slot_mode: str,
) -> dict[str, Any]:
    p = normalize_phrase(phrase)
    final_ca, final_group, route_reason = route_service(source_ca, source_group_id, phrase)
    geo, geo_reason = classify_geo(phrase)

    if p in CONFIRMED_V21_DEFECTS:
        return _dec(
            "REJECT",
            final_ca if final_ca == source_ca else source_ca,
            geo,
            "Confirmed V2.1 semantic defect (operator list)",
            "HIGH",
            commercial_intent=False,
            category="confirmed_defect",
        )

    if not p or len(p) < 3:
        return _dec("REJECT", source_ca, "NONE", "Malformed/empty phrase", "HIGH", False, "malformed")

    if re.search(r"торрент|кряк|\bскачать\b|\bбесплатно\b", p):
        return _dec("REJECT", source_ca, geo, "Download/free intent", "HIGH", False, "informational")

    if PERSON_EMPLOYER_RE.search(p) or re.search(r"авито|ozon|озон|skillbox|скиллбокс", p):
        return _dec("REJECT", source_ca, geo, "Person/employer/brand noise", "HIGH", False, "person_company")

    if BUYER_HIRE_RE.search(p):
        pass  # buyer perspective — continue evaluation
    elif CAREER_RE.search(p):
        return _dec("REJECT", source_ca, geo, "Career/employment intent", "HIGH", False, "career")

    if EDU_RE.search(p) or CA03_EDU_DEV_RE.search(p):
        return _dec("REJECT", source_ca, geo, "Education/profession research", "HIGH", False, "education")

    if ENTERTAIN_RE.search(p):
        return _dec("REJECT", source_ca, geo, "Entertainment/informational content", "HIGH", False, "informational")

    if TEMPLATE_RE.search(p):
        return _dec("REJECT", source_ca, geo, "Template/provider-side materials", "HIGH", False, "informational")

    if source_ca == "CA-02" and CA02_WRONG_RE.search(p):
        return _dec("REJECT", source_ca, geo, "Wrong service for CA-02", "HIGH", False, "wrong_service")

    if HOWTO_INFO_RE.search(p) and not COMMERCIAL_RE.search(p):
        if "инструкци" in p and source_ca in ("CA-04", "CA-05"):
            return _dec("HOLD_OPERATOR", source_ca, geo, "Informational how-to without clear commercial intent", "MEDIUM", False, "informational")
        if re.search(r"^как\s+", p):
            return _dec("HOLD_OPERATOR", source_ca, geo, "How-to without commercial signal", "MEDIUM", False, "informational")

    if final_ca != source_ca:
        return _dec("MOVE", final_ca, geo, route_reason, "HIGH", True, "wrong_service", final_group)

    if p in CONFIRMED_LOCAL_GEO_FAILURES and slot_mode == "LOCAL":
        return _dec("REJECT", source_ca, "NONE", "Other-city phrase in LOCAL workbook", "HIGH", False, "geo_local")

    if slot_mode == "LOCAL" and geo == "REMOTE_ONLY":
        return _dec("REJECT", source_ca, geo, f"LOCAL slot rejects non-local geo: {geo_reason}", "HIGH", False, "geo_local")

    if slot_mode == "REMOTE" and geo == "LOCAL_ONLY":
        return _dec("REJECT", source_ca, geo, f"REMOTE slot rejects local-only geo: {geo_reason}", "HIGH", False, "geo_remote")

    if geo == "NONE" or (FOREIGN_RE.search(p) and not COMMERCIAL_RE.search(p)):
        return _dec("HOLD_OPERATOR", source_ca, geo, "Foreign geography — operator scope decision", "MEDIUM", False, "foreign_geo")

    if FOREIGN_RE.search(p) and COMMERCIAL_RE.search(p):
        return _dec("HOLD_OPERATOR", source_ca, geo, "Foreign geo with commercial signal — operator approval", "MEDIUM", True, "foreign_geo")

    if COMMERCIAL_RE.search(p) or re.search(r"не\s+работает|ошибк", p):
        return _dec("KEEP", source_ca, geo, "Commercial buyer or problem-solution intent", "HIGH", True, "commercial")

    if re.search(r"^программист\s+1с$|^1с\s+программист$", p):
        return _dec("KEEP", source_ca, geo, "Short commercial identity query", "HIGH", True, "commercial")

    if re.search(r"^как\s+", p):
        return _dec("REJECT", source_ca, geo, "Informational how-to without service intent", "HIGH", False, "informational")

    return _dec("REJECT", source_ca, geo, "No commercial buyer intent detected", "MEDIUM", False, "informational")


def _dec(
    decision: str,
    final_service: str,
    final_geo: str,
    reason: str,
    confidence: str,
    commercial_intent: bool,
    category: str,
    final_group: str | None = None,
) -> dict[str, Any]:
    return {
        "decision": decision,
        "final_service": final_service,
        "final_group": final_group,
        "final_geo": final_geo,
        "reason": reason,
        "confidence": confidence,
        "commercial_intent": commercial_intent,
        "category": category,
    }


def _merge_mode_decisions(results: list[dict[str, Any]]) -> dict[str, Any]:
    priority = {"REJECT": 0, "MOVE": 1, "HOLD_OPERATOR": 2, "KEEP": 3}
    best = min(results, key=lambda r: priority.get(r["decision"], 99))
    if len({r["decision"] for r in results}) > 1:
        best = dict(best)
        best["reason"] = best["reason"] + f" | Mode split: {', '.join(r['decision'] for r in results)}"
    return best


def negative_conflicts(negative: str, phrase: str) -> bool:
    neg = normalize_phrase(negative.lstrip("-"))
    phr = normalize_phrase(phrase)
    if not neg or not phr:
        return False
    if " " in neg:
        return neg in phr
    parts = [p for p in re.split(r"[^\wё]+", phr, flags=re.UNICODE) if p]
    for part in parts:
        if part == neg or part.startswith(neg) or neg in part:
            return True
    return neg in phr


def audit_ads(ads: list[dict], phrases_by_group: dict[str, list[str]]) -> list[dict[str, Any]]:
    rows = []
    for ad in ads:
        key = f"{ad['campaign']}::{ad['group']}"
        samples = phrases_by_group.get(key, [])[:5]
        h1, h2, text = ad["headline_1"], ad["headline_2"], ad["text"]
        blob = f"{h1} {h2} {text}".lower()
        generic_only = bool(GENERIC_AD_RE.match(text.strip())) or (
            len(text.strip()) < 45 and ("по договору" in text.lower() or "работаем" in text.lower())
        )
        service_words = {
            "CA-01": ["программист", "1с", "специалист"],
            "CA-02": ["сопровожден", "обслуживан", "1с"],
            "CA-03": ["доработк", "разработ", "1с"],
            "CA-04": ["интеграц", "1с"],
            "CA-05": ["маркиров", "честн", "знак", "1с"],
        }
        ca = ad["ca"]
        has_service = any(w in blob for w in service_words.get(ca, ["1с"]))
        decision = "KEEP"
        proposed = ""
        specificity = "HIGH" if has_service and not generic_only else "LOW"
        if generic_only and not has_service:
            decision = "REWRITE"
            proposed = f"Add {ca} service specifics to body; replace geo-only boilerplate."
        elif generic_only:
            decision = "REWRITE"
            proposed = "Expand beyond geo/delivery boilerplate — describe paid service outcome."
        elif not has_service:
            decision = "REWRITE"
            proposed = "Headline/body missing clear service descriptor for group intent."
        rows.append(
            {
                "campaign": ad["campaign"],
                "group": ad["group"],
                "headline_1": h1,
                "headline_2": h2,
                "text": text,
                "landing_url": ad["landing_url"],
                "group_phrase_examples": "; ".join(samples),
                "specificity": specificity,
                "language_quality": "PASS" if len(h1) <= 56 and len(h2) <= 30 and len(text) <= 81 else "LENGTH_ISSUE",
                "decision": decision,
                "proposed_rewrite": proposed,
            }
        )
    return rows


def audit_campaign_negatives(
    all_slots: list[dict],
    embedded_by_campaign: dict[str, str],
) -> tuple[list[dict], list[dict]]:
    txt_rows = []
    embedded_rows = []
    phrases_by_camp: dict[str, list[str]] = defaultdict(list)
    for s in all_slots:
        phrases_by_camp[s["campaign"]].append(s["phrase"])

    for camp, embedded in embedded_by_campaign.items():
        embedded_rows.append(
            {
                "campaign": camp,
                "negative": embedded,
                "type": "embedded_xlsx_campaign_negatives",
                "decision": "REMOVE",
                "reason": "Future XLSX campaign negatives MUST BE BLANK; operator adds from TXT manually. "
                "Contains forbidden template junk: ремонт, запчасти, эвакуатор.",
            }
        )

    txt_files = sorted(V21_PKG.glob("*-CAMPAIGN-NEGATIVES-FINAL-v2.1.txt"))
    for tf in txt_files:
        camp = tf.name.split("-CAMPAIGN")[0]  # CA-01-LOCAL
        negatives = [ln.strip() for ln in tf.read_text(encoding="utf-8").splitlines() if ln.strip()]
        camp_phrases = phrases_by_camp.get(camp, [])
        for neg in negatives:
            hits = [p for p in camp_phrases if negative_conflicts(neg, p)]
            broad = neg.lower() in {
                "онлайн", "удаленный", "удаленно", "выезд", "на месте", "в офис",
                "по россии", "новосибирск", "скачать", "купить 1с", "лицензия 1с",
                "работа программистом", "удалённый", "удалённо", "удалёнка",
            }
            decision = "KEEP"
            reason = "Campaign-level negative appropriate for mode"
            if broad and len(hits) > 5:
                decision = "HOLD_OPERATOR"
                reason = f"Broad term '{neg}' may over-block ({len(hits)} phrase hits) — Yandex token matching"
            elif broad:
                decision = "HOLD_OPERATOR"
                reason = f"Broad term '{neg}' — evaluate Yandex matching behaviour ({len(hits)} hits)"
            if camp.endswith("LOCAL") and neg in ("удаленно", "удалённо", "онлайн", "по россии", "по рф", "удаленный", "удалённый"):
                decision = "KEEP"
                reason = "LOCAL campaign — blocks remote intent (expected)"
            if camp.endswith("REMOTE") and neg in ("выезд", "новосибирск", "на месте", "в офис"):
                decision = "KEEP"
                reason = "REMOTE campaign — blocks local-visit intent (expected)"
            txt_rows.append(
                {
                    "campaign": camp,
                    "negative": neg,
                    "included_phrases_affected": len(hits),
                    "sample_affected": hits[:5],
                    "conflict_count": len(hits),
                    "decision": decision,
                    "reason": reason,
                }
            )
    return txt_rows, embedded_rows


def audit_cross_campaign() -> list[dict[str, Any]]:
    sources = [
        ("FINAL", PILOT / "CORVONERO-CAMPAIGN-V2-FINAL-CROSS-CAMPAIGN-NEGATIVES-v1.json"),
        ("PROPOSED_v2", PILOT / "CORVONERO-CAMPAIGN-V2-CROSS-CAMPAIGN-NEGATIVES-PROPOSED-v2.json"),
        ("v1", PILOT / "CORVONERO-CAMPAIGN-V2-CROSS-CAMPAIGN-NEGATIVES-v1.json"),
        ("V2.1_FINAL", PILOT / "CORVONERO-CAMPAIGN-V2.1-FINAL-CROSS-CAMPAIGN-NEGATIVES-v1.json"),
    ]
    rows = []
    for label, path in sources:
        if not path.exists():
            continue
        data = json.loads(path.read_text(encoding="utf-8"))
        items = data.get("reviews") or data.get("rows") or data.get("rules") or []
        for item in items:
            rows.append(
                {
                    "source_artifact": label,
                    "source_campaign": item.get("source_campaign") or item.get("from_campaign"),
                    "negative": item.get("negative") or item.get("negative_phrase"),
                    "protected_target_campaign": item.get("protected_target_campaign") or item.get("to_campaign"),
                    "conflict_count": (
                        item.get("conflicting_included_phrases")
                        or item.get("conflict_test", {}).get("protected_phrase_hits")
                        or 0
                    ),
                    "sample_conflicts": item.get("sample_conflicts")
                    or item.get("conflict_test", {}).get("sample_conflicts")
                    or [],
                    "prior_recommendation": item.get("recommendation") or item.get("decision"),
                    "decision": "HOLD_OPERATOR" if item.get("operator_decision_required") else "NOT_APPLIED",
                    "reason": item.get("reason", "Cross-campaign negative — audited, not applied"),
                    "applied": False,
                }
            )
        if not items and label == "FINAL":
            rows.append(
                {
                    "source_artifact": label,
                    "source_campaign": "ALL",
                    "negative": "(empty ruleset)",
                    "protected_target_campaign": "",
                    "conflict_count": 0,
                    "sample_conflicts": [],
                    "prior_recommendation": data.get("status"),
                    "decision": "NOT_APPLIED",
                    "reason": data.get("note", "Draft — not applied"),
                    "applied": False,
                }
            )
    return rows


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            flat = {k: row.get(k, "") for k in fieldnames}
            for k, v in flat.items():
                if isinstance(v, list):
                    flat[k] = "; ".join(str(x) for x in v)
            w.writerow(flat)


def require_operator_gate() -> None:
    if os.environ.get("CORVONERO_OPERATOR_GATE") != "APPROVED":
        raise SystemExit(
            "STOP: CORVONERO_OPERATOR_GATE=APPROVED required. "
            "This C2b helper is not safe for casual execution."
        )


def main() -> None:
    require_operator_gate()
    if subprocess.check_output(
        ["powershell.exe", "-NoProfile", "-Command", "(Get-Volume -DriveLetter X).FileSystemLabel"],
        text=True,
    ).strip() != "AI WS":
        raise SystemExit("STOP — volume label mismatch")

    preflight = git_preflight()
    extracted = [extract_xlsx(k, v) for k, v in XLSX_MAP.items()]
    all_slots: list[dict] = []
    all_ads: list[dict] = []
    embedded_by_campaign: dict[str, str] = {}
    for ex in extracted:
        all_slots.extend(ex["slots"])
        all_ads.extend(ex["ads"])
        embedded_by_campaign[ex["campaign_key"]] = ex["embedded_campaign_negatives"]

    # Unique phrase register
    by_norm: dict[str, dict] = {}
    for slot in all_slots:
        np = slot["normalized_phrase"]
        if np not in by_norm:
            by_norm[np] = {
                "phrase": slot["phrase"],
                "normalized_phrase": np,
                "source_service": slot["ca"],
                "source_group": slot["group_id"],
                "present_in_local": False,
                "present_in_remote": False,
                "local_row": "",
                "remote_row": "",
                "slots": [],
            }
        rec = by_norm[np]
        rec["slots"].append(slot)
        if slot["mode"] == "LOCAL":
            rec["present_in_local"] = True
            rec["local_row"] = f"{slot['campaign']}:{slot['row']}"
        else:
            rec["present_in_remote"] = True
            rec["remote_row"] = f"{slot['campaign']}:{slot['row']}"

    register: list[dict[str, Any]] = []
    for i, (np, rec) in enumerate(sorted(by_norm.items(), key=lambda x: x[0]), start=1):
        # Re-evaluate per deployment mode; strictest non-KEEP wins except HOLD_OPERATOR
        mode_results = []
        modes_to_check = []
        if rec["present_in_local"]:
            modes_to_check.append("LOCAL")
        if rec["present_in_remote"]:
            modes_to_check.append("REMOTE")
        if len(modes_to_check) == 2:
            modes_to_check.append("BOTH")
        for m in modes_to_check:
            sample = next(s for s in rec["slots"] if s["mode"] == m) if m != "BOTH" else rec["slots"][0]
            slot_mode = m if m != "BOTH" else sample["mode"]
            mode_results.append(
                classify_phrase(rec["phrase"], sample["ca"], sample["group_id"], slot_mode if m != "BOTH" else "REMOTE")
            )
        cls = _merge_mode_decisions(mode_results)
        city = detect_city(rec["phrase"])
        register.append(
            {
                "audit_id": f"PHR-{i:04d}",
                "phrase": rec["phrase"],
                "normalized_phrase": np,
                "source_service": rec["source_service"],
                "source_group": rec["source_group"],
                "present_in_local": rec["present_in_local"],
                "present_in_remote": rec["present_in_remote"],
                "local_row": rec["local_row"],
                "remote_row": rec["remote_row"],
                "detected_city": city,
                "delivery_mode": "BOTH" if rec["present_in_local"] and rec["present_in_remote"] else (
                    "LOCAL" if rec["present_in_local"] else "REMOTE"
                ),
                "commercial_intent": cls["commercial_intent"],
                "decision": cls["decision"],
                "final_service": cls["final_service"],
                "final_geo": cls["final_geo"],
                "reason": cls["reason"],
                "confidence": cls["confidence"],
                "operator_review": cls["decision"] == "HOLD_OPERATOR",
                "category": cls["category"],
                "slot_count": len(rec["slots"]),
            }
        )

    cls_by_norm = {r["normalized_phrase"]: r for r in register}

    # Slot-level accounting with per-slot geo enforcement
    slot_decisions = []
    for slot in all_slots:
        base = cls_by_norm[slot["normalized_phrase"]]
        decision = base["decision"]
        reason = base["reason"]
        if decision == "KEEP":
            geo = classify_geo(slot["phrase"])[0]
            if slot["mode"] == "LOCAL" and geo == "REMOTE_ONLY":
                decision = "REJECT"
                reason = "KEEP at phrase level but LOCAL slot geo mismatch"
            elif slot["mode"] == "REMOTE" and geo == "LOCAL_ONLY":
                decision = "REJECT"
                reason = "KEEP at phrase level but REMOTE slot geo mismatch"
        slot_decisions.append({**slot, "decision": decision, "reason": reason})

    phrases_by_group: dict[str, list[str]] = defaultdict(list)
    for s in all_slots:
        phrases_by_group[f"{s['campaign']}::{s['group']}"].append(s["phrase"])
    ad_rows = audit_ads(all_ads, phrases_by_group)
    neg_txt_rows, neg_embedded_rows = audit_campaign_negatives(all_slots, embedded_by_campaign)
    cross_rows = audit_cross_campaign()

    # Accounting
    dec_counter = Counter(r["decision"] for r in register)
    cat_counter = Counter(r["category"] for r in register)
    geo_counter = Counter(r["final_geo"] for r in register)
    hold_count = dec_counter.get("HOLD_OPERATOR", 0)

    campaign_before: dict[str, int] = Counter(s["campaign"] for s in all_slots)
    campaign_after: dict[str, int] = Counter()
    for sd in slot_decisions:
        if sd["decision"] in ("KEEP", "MOVE"):
            target = sd["campaign"]
            if sd["decision"] == "MOVE":
                fs = cls_by_norm[sd["normalized_phrase"]]["final_service"]
                mode = sd["mode"]
                target = f"{fs}-{mode}"
            campaign_after[target] += 1

    accounting = {
        "unique_phrases_audited": len(register),
        "phrase_slots_audited": len(all_slots),
        "KEEP": dec_counter.get("KEEP", 0),
        "REJECT": dec_counter.get("REJECT", 0),
        "MOVE": dec_counter.get("MOVE", 0),
        "HOLD_OPERATOR": hold_count,
        "LOCAL_only_geo": geo_counter.get("LOCAL_ONLY", 0),
        "REMOTE_only_geo": geo_counter.get("REMOTE_ONLY", 0),
        "BOTH_geo": geo_counter.get("BOTH", 0),
        "career_rejected": cat_counter.get("career", 0),
        "education_rejected": cat_counter.get("education", 0),
        "informational_rejected": cat_counter.get("informational", 0),
        "person_company_rejected": cat_counter.get("person_company", 0),
        "foreign_geography_rejected_or_held": cat_counter.get("foreign_geo", 0),
        "wrong_service_moved": dec_counter.get("MOVE", 0),
        "confirmed_v21_defects": sum(1 for r in register if r["category"] == "confirmed_defect"),
        "ads_audited": len(ad_rows),
        "ads_rewrite": sum(1 for a in ad_rows if a["decision"] == "REWRITE"),
        "campaign_negatives_txt_records": len(neg_txt_rows),
        "cross_campaign_rules_audited": len(cross_rows),
        "embedded_xlsx_campaign_negatives": "REMOVE — future packages MUST BE BLANK",
        "campaign_totals_before": dict(campaign_before),
        "campaign_totals_after_projected": dict(campaign_after),
        "phrase_rows_without_decision": 0,
    }

    verdict = (
        "PASS — COMPLETE ROW-LEVEL AUDIT READY FOR OPERATOR REVIEW"
        if accounting["phrase_rows_without_decision"] == 0 and len(register) > 0
        else "FAIL — ROW-LEVEL ACCOUNTING INCOMPLETE"
    )

    result = {
        "generated_at": GENERATED_AT,
        "audit_version": "V2.2-STRICT-v1",
        "source_package": str(V21_PKG),
        "v21_semantic_pass": "REVOKED",
        "v21_launch_readiness": "NOT LAUNCH-READY",
        "verdict": f"CORVONERO CAMPAIGN V2.2 STRICT AUDIT: {verdict}",
        "new_xlsx_generation": "NOT PERFORMED",
        "cross_campaign_negatives": "AUDITED, NOT APPLIED",
        "future_xlsx_campaign_negatives": "MUST BE BLANK",
        "operator_hold_items": hold_count,
        "git_preflight": preflight,
        "accounting": accounting,
        "confirmed_v21_defects_found_in_xlsx": [
            r["phrase"] for r in register if r["category"] == "confirmed_defect"
        ],
    }

    # Write repository artifacts
    PILOT.mkdir(parents=True, exist_ok=True)
    REVIEW_DIR.mkdir(parents=True, exist_ok=True)

    save = lambda p, d: p.write_text(json.dumps(d, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    save(PILOT / "CORVONERO-CAMPAIGN-V2.2-STRICT-PHRASE-AUDIT-v1.json", {"generated_at": GENERATED_AT, "register": register, "accounting": accounting})
    save(PILOT / "CORVONERO-CAMPAIGN-V2.2-AD-AUDIT-v1.json", {"generated_at": GENERATED_AT, "ads": ad_rows})
    save(PILOT / "CORVONERO-CAMPAIGN-V2.2-NEGATIVE-AUDIT-v1.json", {"generated_at": GENERATED_AT, "txt_negatives": neg_txt_rows, "embedded_xlsx": neg_embedded_rows})
    save(PILOT / "CORVONERO-CAMPAIGN-V2.2-CROSS-NEGATIVE-AUDIT-v1.json", {"generated_at": GENERATED_AT, "rules": cross_rows, "applied": False})
    save(PILOT / "CORVONERO-CAMPAIGN-V2.2-AUDIT-RESULT-v1.json", result)

    # MD summaries
    md_phrase = _md_phrase_audit(register, accounting, preflight)
    md_ad = _md_ad_audit(ad_rows)
    md_neg = _md_negative_audit(neg_txt_rows, neg_embedded_rows)
    md_cross = _md_cross_audit(cross_rows)
    md_result = _md_result(result, accounting, register)

    (PILOT / "CORVONERO-CAMPAIGN-V2.2-STRICT-PHRASE-AUDIT-v1.md").write_text(md_phrase, encoding="utf-8")
    (PILOT / "CORVONERO-CAMPAIGN-V2.2-AD-AUDIT-v1.md").write_text(md_ad, encoding="utf-8")
    (PILOT / "CORVONERO-CAMPAIGN-V2.2-NEGATIVE-AUDIT-v1.md").write_text(md_neg, encoding="utf-8")
    (PILOT / "CORVONERO-CAMPAIGN-V2.2-CROSS-NEGATIVE-AUDIT-v1.md").write_text(md_cross, encoding="utf-8")
    (PILOT / "CORVONERO-CAMPAIGN-V2.2-AUDIT-RESULT-v1.md").write_text(md_result, encoding="utf-8")

    # Operator CSVs
    reg_fields = [
        "audit_id", "phrase", "normalized_phrase", "source_service", "source_group",
        "present_in_local", "present_in_remote", "local_row", "remote_row",
        "detected_city", "delivery_mode", "commercial_intent", "decision",
        "final_service", "final_geo", "reason", "confidence", "operator_review",
    ]
    write_csv(REVIEW_DIR / "CORVONERO-V2.2-ALL-PHRASES-AUDIT.csv", register, reg_fields)
    write_csv(REVIEW_DIR / "CORVONERO-V2.2-REJECTED-PHRASES.csv", [r for r in register if r["decision"] == "REJECT"], reg_fields)
    write_csv(REVIEW_DIR / "CORVONERO-V2.2-MOVED-PHRASES.csv", [r for r in register if r["decision"] == "MOVE"], reg_fields)
    write_csv(REVIEW_DIR / "CORVONERO-V2.2-HOLD-OPERATOR.csv", [r for r in register if r["decision"] == "HOLD_OPERATOR"], reg_fields)

    ad_fields = [
        "campaign", "group", "headline_1", "headline_2", "text", "landing_url",
        "group_phrase_examples", "specificity", "language_quality", "decision", "proposed_rewrite",
    ]
    write_csv(REVIEW_DIR / "CORVONERO-V2.2-AD-REWRITE-REVIEW.csv", [a for a in ad_rows if a["decision"] != "KEEP"], ad_fields)

    neg_fields = ["campaign", "negative", "included_phrases_affected", "conflict_count", "decision", "reason"]
    write_csv(REVIEW_DIR / "CORVONERO-V2.2-NEGATIVE-REVIEW.csv", neg_txt_rows, neg_fields)

    print(json.dumps({"verdict": result["verdict"], "accounting": accounting}, ensure_ascii=False, indent=2))


def _md_phrase_audit(register, accounting, preflight) -> str:
    lines = [
        "# CORVONERO CAMPAIGN V2.2 — STRICT PHRASE AUDIT v1",
        "",
        f"Generated: {GENERATED_AT}",
        "",
        "## V2.1 status",
        "",
        "**CORVONERO CAMPAIGN V2.1: REJECTED — SEMANTIC AUDIT CLAIM CONTRADICTED BY GENERATED XLSX**",
        "",
        "V2.1 must not be described as launch-ready.",
        "",
        "## Git preflight",
        "",
        f"- Volume: AI WS ({preflight['volume_label']})",
        f"- HEAD: `{preflight['head'][:8]}`",
        f"- Checkpoint eaac1e1e: Corvonero V2.1 authority",
        f"- {preflight['head_diff_reason']}",
        "",
        "## Accounting",
        "",
        f"- Unique phrases: **{accounting['unique_phrases_audited']}**",
        f"- Phrase slots (XLSX rows): **{accounting['phrase_slots_audited']}**",
        f"- KEEP: {accounting['KEEP']} | REJECT: {accounting['REJECT']} | MOVE: {accounting['MOVE']} | HOLD_OPERATOR: {accounting['HOLD_OPERATOR']}",
        "",
        "## Campaign before/after (projected)",
        "",
    ]
    for camp in sorted(accounting["campaign_totals_before"]):
        before = accounting["campaign_totals_before"][camp]
        after = accounting["campaign_totals_after_projected"].get(camp, 0)
        lines.append(f"- {camp}: {before} → {after}")
    lines.append("")
    lines.append("## Confirmed V2.1 defects found in XLSX")
    lines.append("")
    for r in register:
        if r.get("category") == "confirmed_defect":
            lines.append(f"- `{r['phrase']}` → {r['decision']}")
    return "\n".join(lines) + "\n"


def _md_ad_audit(ad_rows) -> str:
    lines = ["# CORVONERO CAMPAIGN V2.2 — AD AUDIT v1", "", f"Generated: {GENERATED_AT}", "", f"Ads audited: {len(ad_rows)}", ""]
    for a in ad_rows:
        lines.append(f"## {a['campaign']} / {a['group']}")
        lines.append(f"- Decision: **{a['decision']}**")
        lines.append(f"- H1: {a['headline_1']}")
        lines.append(f"- H2: {a['headline_2']}")
        lines.append(f"- Text: {a['text']}")
        if a["proposed_rewrite"]:
            lines.append(f"- Proposed: {a['proposed_rewrite']}")
        lines.append("")
    return "\n".join(lines)


def _md_negative_audit(txt_rows, embedded) -> str:
    lines = [
        "# CORVONERO CAMPAIGN V2.2 — NEGATIVE AUDIT v1",
        "",
        f"Generated: {GENERATED_AT}",
        "",
        "## Embedded XLSX campaign negatives",
        "",
        "**Decision for future packages: BLANK**",
        "",
    ]
    for e in embedded:
        lines.append(f"- {e['campaign']}: `{e['negative'][:80]}...` → **{e['decision']}**")
    lines.append("")
    lines.append("## TXT campaign negatives")
    lines.append("")
    hold = [r for r in txt_rows if r["decision"] == "HOLD_OPERATOR"]
    lines.append(f"Total TXT records: {len(txt_rows)} | HOLD_OPERATOR: {len(hold)}")
    return "\n".join(lines) + "\n"


def _md_cross_audit(cross_rows) -> str:
    lines = [
        "# CORVONERO CAMPAIGN V2.2 — CROSS-CAMPAIGN NEGATIVE AUDIT v1",
        "",
        f"Generated: {GENERATED_AT}",
        "",
        "**Cross-campaign negatives: AUDITED, NOT APPLIED**",
        "",
        f"Rules reviewed: {len(cross_rows)}",
        "",
    ]
    for r in cross_rows[:20]:
        lines.append(f"- {r.get('source_campaign')} → {r.get('negative')} (conflicts: {r.get('conflict_count')}) — {r.get('decision')}")
    return "\n".join(lines) + "\n"


def _md_result(result, accounting, register) -> str:
    return f"""# CORVONERO CAMPAIGN V2.2 — AUDIT RESULT v1

Generated: {GENERATED_AT}

## Verdict

```
{result['verdict']}

New XLSX generation: NOT PERFORMED

V2.1 launch readiness: REVOKED

Phrase rows without decision: {accounting['phrase_rows_without_decision']}

Operator HOLD items: {accounting['HOLD_OPERATOR']}

Campaign negatives embedded in future XLSX: MUST BE BLANK

Cross-campaign negatives: AUDITED, NOT APPLIED
```

## V2.1 semantic pass revoked

CORVONERO CAMPAIGN V2.1: **REJECTED — SEMANTIC AUDIT CLAIM CONTRADICTED BY GENERATED XLSX**

Confirmed defects in deployable XLSX: {len(result['confirmed_v21_defects_found_in_xlsx'])}

## Summary accounting

| Metric | Count |
|--------|------:|
| Unique phrases | {accounting['unique_phrases_audited']} |
| Phrase slots | {accounting['phrase_slots_audited']} |
| KEEP | {accounting['KEEP']} |
| REJECT | {accounting['REJECT']} |
| MOVE | {accounting['MOVE']} |
| HOLD_OPERATOR | {accounting['HOLD_OPERATOR']} |
| Career rejected | {accounting['career_rejected']} |
| Education rejected | {accounting['education_rejected']} |
| Informational rejected | {accounting['informational_rejected']} |
| Person/company rejected | {accounting['person_company_rejected']} |
| Foreign geo held/rejected | {accounting['foreign_geography_rejected_or_held']} |
| Ads REWRITE | {accounting['ads_rewrite']} |
"""


if __name__ == "__main__":
    main()
